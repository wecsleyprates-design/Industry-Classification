"""
Worth AI — KYB Portal Intelligence Hub
7-tab app: one tab per KYB sub-tab + API Fields + AI Agent
Every card, every field, every badge fully documented with source-code proof.
"""
import io, os, base64, json
import streamlit as st
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from kyb_rag_builder import load_or_build, search as rag_search

st.set_page_config(
    page_title="Worth AI — KYB Intelligence Hub",
    page_icon="🔍",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Colours ───────────────────────────────────────────────────────────────────
BG="#0F172A";GRID="#1E293B";RED="#F87171";AMBER="#FCD34D";GREEN="#34D399"
BLUE="#60A5FA";TEAL="#2DD4BF";GREY="#94A3B8";PURPLE="#A78BFA";TEXT="#E2E8F0"

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""<style>
html,body,[class*="css"]{font-family:'Inter','Segoe UI',sans-serif;}
section[data-testid="stSidebar"]{background:#0B1120!important;}
section[data-testid="stSidebar"] *{color:#E2E8F0!important;}
.card{background:#1A2744;border:1px solid #2D4070;border-left:5px solid #3B82F6;border-radius:8px;padding:14px 18px;margin-bottom:10px;color:#CBD5E1;line-height:1.65;}
.card *{color:#CBD5E1!important;} .card b,.card strong{color:#E2E8F0!important;}
.card-red{background:#2A1010;border-left:5px solid #EF4444;color:#FCA5A5!important;}
.card-red *{color:#FCA5A5!important;} .card-red b{color:#FEE2E2!important;}
.card-green{background:#0C2218;border-left:5px solid #10B981;color:#6EE7B7!important;}
.card-green *{color:#6EE7B7!important;} .card-green b{color:#A7F3D0!important;}
.card-amber{background:#221A06;border-left:5px solid #F59E0B;color:#FCD34D!important;}
.card-amber *{color:#FCD34D!important;} .card-amber b{color:#FDE68A!important;}
.card-purple{background:#180D30;border-left:5px solid #8B5CF6;color:#C4B5FD!important;}
.card-purple *{color:#C4B5FD!important;} .card-purple b{color:#DDD6FE!important;}
.card-teal{background:#061B1B;border-left:5px solid #2DD4BF;color:#99F6E4!important;}
.card-teal *{color:#99F6E4!important;} .card-teal b{color:#CCFBF1!important;}
.badge{display:inline-block;padding:3px 11px;border-radius:12px;font-size:.73rem;font-weight:700;margin:2px;letter-spacing:.01em;}
.b-blue{background:#1E3A6E;color:#93C5FD;border:1px solid #2563EB;}
.b-green{background:#064E3B;color:#6EE7B7;border:1px solid #059669;}
.b-amber{background:#451A03;color:#FCD34D;border:1px solid #D97706;}
.b-red{background:#450A0A;color:#FCA5A5;border:1px solid #DC2626;}
.b-grey{background:#1E293B;color:#94A3B8;border:1px solid #475569;}
.b-teal{background:#042F2E;color:#5EEAD4;border:1px solid #0D9488;}
.metric-block{background:#1A2744;border:1px solid #2D4070;border-radius:10px;padding:16px 18px;text-align:center;}
.metric-num{font-size:2rem;font-weight:700;color:#60A5FA;}
.metric-label{font-size:.78rem;color:#94A3B8;margin-top:4px;}
.t{width:100%;border-collapse:collapse;font-size:.87rem;}
.t th{background:#0F2040;color:#93C5FD;font-weight:700;padding:9px 13px;text-align:left;border-bottom:2px solid #2563EB;}
.t td{padding:8px 13px;border-bottom:1px solid #1E293B;vertical-align:top;color:#CBD5E1;background:#0F172A;}
.t tr:nth-child(even) td{background:#141F35;}
.t code{background:#1E3A5F;color:#93C5FD;padding:1px 6px;border-radius:4px;font-size:.82em;}
.sh{background:linear-gradient(135deg,#0F2040 0%,#1D4ED8 100%);color:#E2E8F0;padding:11px 20px;border-radius:8px;font-weight:700;font-size:1.08rem;margin-bottom:12px;border:1px solid #2563EB;}
.cb{background:#0A1628;color:#93C5FD;border:1px solid #1E3A5F;border-radius:6px;padding:12px 16px;font-family:'Courier New',monospace;font-size:.82rem;white-space:pre-wrap;word-break:break-all;line-height:1.55;}
.src-ref{background:#0A1628;border:1px solid #1E3A5F;border-left:4px solid #60A5FA;border-radius:6px;padding:8px 12px;margin:6px 0;font-size:.8rem;color:#93C5FD;}
.field-row{background:#0E1E38;border:1px solid #1E3A5F;border-radius:8px;padding:12px 16px;margin-bottom:6px;}
.badge-demo{display:inline-flex;align-items:center;gap:6px;padding:4px 12px;border-radius:9999px;font-size:.8rem;font-weight:600;margin:3px;}
.badge-verified{background:#DBEAFE;color:#1E40AF;border:1px solid #93C5FD;}
.badge-unverified{background:#FEF9C3;color:#854D0E;border:1px solid #FDE047;}
.badge-missing{background:#FEE2E2;color:#991B1B;border:1px solid #FCA5A5;}
.badge-deliverable{background:#D1FAE5;color:#065F46;border:1px solid #6EE7B7;}
.badge-nohits{background:#D1FAE5;color:#065F46;border:1px solid #6EE7B7;}
.badge-hits{background:#FEE2E2;color:#991B1B;border:1px solid #FCA5A5;}
.badge-online{background:#DBEAFE;color:#1E40AF;border:1px solid #93C5FD;}
.badge-offline{background:#FEE2E2;color:#991B1B;border:1px solid #FCA5A5;}
.badge-unknown{background:#F3F4F6;color:#374151;border:1px solid #D1D5DB;}
.badge-match{background:#D1FAE5;color:#065F46;border:1px solid #6EE7B7;}
.badge-nomatch{background:#FEE2E2;color:#991B1B;border:1px solid #FCA5A5;}
</style>""", unsafe_allow_html=True)

REPO_URLS = {
    "ADMIN_PORTAL": "https://github.com/wecsleyprates-design/Admin-Portal-KYB/blob/main",
    "SIC_UK": "https://github.com/wecsleyprates-design/SIC-UK-Codes/blob/main",
}

# ── Helpers ───────────────────────────────────────────────────────────────────
def sh(t): st.markdown(f'<div class="sh">{t}</div>', unsafe_allow_html=True)
def card(t, s=""): st.markdown(f'<div class="card {s}">{t}</div>', unsafe_allow_html=True)
def src(stype, path, l1, l2, label=""):
    url = f"{REPO_URLS.get(stype,'')}/{path}#L{l1}-L{l2}"
    tag = "🖥️ Frontend" if stype=="ADMIN_PORTAL" else "⚙️ Backend"
    return f'<div class="src-ref">{tag} <a href="{url}" target="_blank" style="color:#60A5FA;">🔗 {path} L{l1}–{l2}</a>{"  — "+label if label else ""}</div>'

def badge(text, kind):
    cls = {"green":"b-green","blue":"b-blue","amber":"b-amber","red":"b-red","grey":"b-grey","teal":"b-teal"}.get(kind,"b-grey")
    return f'<span class="badge {cls}">{text}</span>'

def portal_badge(text, cls):
    return f'<span class="badge-demo {cls}">{text}</span>'

def metric(label, value, colour=BLUE):
    return f'<div class="metric-block"><div class="metric-num" style="color:{colour};">{value}</div><div class="metric-label">{label}</div></div>'

@st.cache_resource
def get_rag():
    return load_or_build()

rag_index = get_rag()

# ── Load UCM spreadsheet fields ───────────────────────────────────────────────
@st.cache_data
def load_ucm_fields():
    import openpyxl, os
    # Try both locations — Admin-Portal-KYB repo (cloned) and workspace backup
    for path in [
        # Primary: same folder as this app (uploaded to Industry-Classification/Admin-Portal-KYB-App/)
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "[GPN version] Worth Field Outputs_Working Session Disussions_020426.xlsx"),
        # Fallback: Admin-Portal-KYB cloned repo
        "/tmp/Admin-Portal-KYB/[GPN version] Worth Field Outputs_Working Session Disussions_020426.xlsx",
        # Fallback: workspace NAICS classifier folder
        "/workspace/AI-Powered-NAICS-Industry-Classification-Agent/naics_mcc_classifier/[GPN version] Worth Field Outputs_Working Session Disussions_020426.xlsx",
    ]:
        try:
            wb = openpyxl.load_workbook(path)
            ws = wb['WS UCM QA']
            fields = []
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                if not any(row[:16]): continue
                if row[4]:
                    fields.append({
                        'section': str(row[0] or '').strip(), 'data_type': str(row[1] or '').strip(),
                        'internal_name': str(row[2] or '').strip(), 'api_endpoint': str(row[3] or '').strip(),
                        'api_field': str(row[4]).strip(), 'requires_transform': str(row[5] or '').strip(),
                        'gpn_questions': str(row[6] or '').strip()[:600],
                        'decisions': str(row[7] or '').strip()[:600],
                        'w360': str(row[8] or '').strip(), 'description': str(row[9] or '').strip()[:400],
                        'display_ucm': str(row[10] or '').strip(), 'decisional': str(row[11] or '').strip(),
                    })
            if fields:
                return fields
        except Exception:
            continue
    return []

UCM_FIELDS = load_ucm_fields()

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style="background:linear-gradient(135deg,#0B0F2A 0%,#1a1150 100%);border-radius:10px;
         padding:14px 18px;margin-bottom:8px;border:1px solid #2D2080;text-align:center;">
      <div style="font-size:1.2rem;font-weight:900;color:#E2E8F0;">
        <span style="color:#B57BFF;">W</span>ORTH <span style="color:#60A5FA;">AI</span>
      </div>
      <div style="font-size:.58rem;color:#8B8FBF;letter-spacing:.15em;margin-top:2px;">KYB INTELLIGENCE HUB</div>
    </div>""", unsafe_allow_html=True)

    tab = st.radio("KYB Sub-Tab", [
        "📋 Background",
        "🏛️ Business Registration",
        "📬 Contact Information",
        "🌐 Website",
        "🔍 Watchlists",
        "🔑 KYB API Fields",
        "🤖 AI Agent",
    ], label_visibility="collapsed")
    st.markdown("---")
    st.caption("Source: Admin-Portal-KYB · SIC-UK-Codes · April 2026")


# ════════════════════════════════════════════════════════════════════════════
# SHARED: field card renderer
# ════════════════════════════════════════════════════════════════════════════
# ── Shared badge library — full details for every badge used in the app ───────
BADGE_LIBRARY = {
    # TIN / Business Registration
    "tin_verified": (
        "✅ Verified", "badge-verified",
        "IRS confirmed the submitted EIN and legal business name are a valid registered match.",
        "tin_match_boolean.value === true",
        "Depends on: tin_match fact (Middesk IRS TIN Match task result). Middesk submits EIN+name to IRS TIN Matching program and returns {status:'success'}. tin_match_boolean is then computed as (status==='success'). Source: TinBadge.tsx L22-25.",
        "No action required. Business is verified with the IRS."
    ),
    "tin_unverified": (
        "⚠️ Unverified", "badge-unverified",
        "IRS could not confirm the EIN and legal name are a registered match.",
        "tin_match_boolean.value === false  OR  tin_match.value.status === 'failure'",
        "Depends on: tin_match.value.status from Middesk IRS TIN Match task. Possible causes: (1) incorrect EIN submitted, (2) business name does not match IRS records, (3) sole proprietor using personal SSN instead of EIN, (4) EIN too new (not yet in IRS records). Source: TinBadge.tsx L27-35.",
        "UCM Rule: Unverified → FAIL. Analyst should request correct EIN or IRS EIN confirmation letter. Check if applicant uploaded a shareholder document for manual verification."
    ),
    # SoS Filings
    "sos_verified": (
        "✅ Verified", "badge-verified",
        "An active Secretary of State filing was found for this business and is in good standing.",
        "sos_match_boolean.value === true  AND  sosFiling.active === true",
        "Two independent facts must both be true: (1) sos_match_boolean=true (Middesk found a SoS filing for this TIN+name), (2) sosFiling.active=true (that filing's status === 'active'). Source: SOSBadge.tsx L17-28. Tooltip shown to analyst: 'An active filing was found and is in good standing.'",
        "No action required. SoS registration is confirmed and active."
    ),
    "sos_missing_active": (
        "🔴 Missing Active Filing", "badge-missing",
        "A SoS filing was found but it is not currently active (inactive/dissolved), OR no filing was found at all.",
        "sosFiling.active === false  OR  sos_match_boolean === false",
        "Two sub-cases: (a) Filing found but inactive: sos_match_boolean=true AND sosFiling.active=false — Middesk found a record but status='inactive'/'dissolved'. (b) No filing at all: sos_match_boolean=false — Middesk searched all 50 US SoS databases by TIN+name and found nothing. Source: SOSBadge.tsx L30-41, SOSBadges.ts constants.",
        "UCM Rule: Review required. Sub-case (a): check if business has re-filed or is operating illegally. Sub-case (b): verify EIN, check if sole proprietor (not required to file with SoS in most states), or if business name/address variant was used at registration."
    ),
    "sos_invalidated": (
        "⚠️ Invalidated", "badge-unverified",
        "This SoS filing has been marked as invalid by Worth AI (e.g. data quality issue flagged).",
        "isInvalidated === true  (prop passed to SOSBadge component)",
        "Depends on: isInvalidated prop from BusinessRegistrationTab. This is set when Worth detects a data inconsistency in the SoS record. Source: SOSBadge.tsx L14-16.",
        "Analyst review required. Manually verify the SoS record via the state registry URL (Articles of Incorporation link in the SoS card)."
    ),
    "no_registry_data": (
        "No Registry Data", "badge-unknown",
        "Middesk searched all SoS databases and found zero registrations for this EIN+name combination.",
        "isNonEmptyArray(sos_filings.value) === false  (sos_filings array is empty)",
        "Depends on: sos_filings.value (array from facts table). If empty, 'No Registry Data to Display' card is shown instead of filing cards. This is NOT a confidence threshold — Worth shows exactly what Middesk returned. Source: BusinessRegistrationTab.tsx — shouldHideSosForIntegration check.",
        "No immediate action if business is a sole proprietor (not required to file in most states). If business claims to be an LLC/Corp, investigate why no SoS record exists — may indicate misrepresentation."
    ),
    # Addresses
    "address_verified": (
        "✅ Business Registration", "badge-verified",
        "The submitted address was confirmed in public business registration records (Middesk SoS data) AND matches at least one verified address.",
        "address_verification_boolean.value === true  OR  address_match_boolean.value === true",
        "Depends on: address_verification (Middesk address task result) and address_match_boolean. Middesk compares submitted address against SoS-registered address. enrichAddressesWithStatusFor360ReportParity() applies status to each address. Source: AddressesCard.tsx L240-244.",
        "No action required. Address confirmed in public records."
    ),
    "address_google_verified": (
        "✅ Google Profile", "badge-verified",
        "Google Business Profile confirms the same address AND business name as submitted.",
        "googleProfileMatch === true  (business_match='match found' AND address_match='match')",
        "Depends on: googleProfileData.data.business_match AND address_match from Google Places API (via SERP). Both must match. Source: AddressesCard.tsx L223-227. The Google Place ID is found by SERP scraping first, then Google Places API is called.",
        "No action required. Independent Google verification passed."
    ),
    "address_unverified": (
        "⚠️ Business Registration", "badge-unverified",
        "The submitted address could not be confirmed in public business registration records.",
        "address_verification_boolean.value === false  AND  address_match_boolean.value === false",
        "Depends on: address_verification.value.sublabel (shown as message under badge) and address_match_boolean. Address not found in Middesk SoS data or does not match registered address. Source: AddressesCard.tsx, KnowYourBusiness/index.tsx L413-444.",
        "Analyst should verify address against state SoS website directly. May indicate: (1) business uses a virtual office, (2) address changed since SoS filing, (3) DBA address differs from registered address."
    ),
    "address_google_unverified": (
        "⚠️ Google Profile", "badge-unverified",
        "Google Business Profile shows a different address, or no Google Business Profile was found for this business.",
        "googleProfileMatch === false  (no profile found OR address/name mismatch)",
        "Depends on: Google Places API result. googleProfileMatch = (business_match==='match found') AND (address_match==='match'). If either fails, badge shows Unverified. Many legitimate businesses have no Google listing. Source: AddressesCard.tsx L223-228.",
        "Low risk if business is new or brick-and-mortar without Google presence. Higher risk if business claimed to be an established restaurant/retail but has no Google listing."
    ),
    "address_deliverable": (
        "✅ Deliverable", "badge-deliverable",
        "USPS confirms this address is a valid mailable postal address.",
        "address in addresses_deliverable.value[]  (USPS deliverability check passed)",
        "Depends on: addresses_deliverable fact (array of USPS-verified addresses). addressesDeliverable() in KnowYourBusiness/index.tsx L82-106 compares each address against the deliverable set. Source: addresses_deliverable fact from integration-service.",
        "Informational. Deliverable does NOT mean the business actually operates there — only that mail can be delivered. Combine with Business Registration badge for stronger verification."
    ),
    # Business Names
    "name_verified": (
        "✅ Verified", "badge-verified",
        "The submitted business name (and/or DBA) was found in public business registration records.",
        "name_match_boolean.value === true",
        "Depends on: name_match_boolean fact (boolean derived from Middesk name task). Middesk's name task checks if submitted name matches SoS/IRS records. Source: BusinessNamesCard.tsx, name_match_boolean fact from kyb/index.ts.",
        "No action required. Business name confirmed in public records."
    ),
    "name_failure": (
        "❌ Failure", "badge-missing",
        "The submitted business name was NOT found in public business registration records.",
        "name_match_boolean.value === false  AND  name_match.value.status !== 'warning'",
        "Depends on: name_match_boolean=false. Middesk could not confirm submitted name in SoS records. Source: KnowYourBusiness/index.tsx L499-505.",
        "UCM Rule: Fail. Analyst should verify: (1) is the submitted name a DBA (not the legal name)? (2) Is there a spelling variant? (3) Did the business recently change names?"
    ),
    "name_warning": (
        "⚠️ [sublabel]", "badge-unverified",
        "Partial name match found — business found but under a different name variant.",
        "name_match.value.status === 'warning'  (sublabel shows the specific reason)",
        "Depends on: name_match.value.status and name_match.value.sublabel. Middesk returns a warning when the name partially matches (e.g. 'Unregistered' or 'DBA not found'). The sublabel text is shown to the analyst. Source: KnowYourBusiness/index.tsx L493-505.",
        "Review the specific sublabel message. May be acceptable if DBA differs from legal name — check the Legal Business Name field for the registered name."
    ),
    # Watchlists
    "watchlist_no_hits": (
        "✅ No Hits", "badge-nohits",
        "This entity (business name or officer name) was screened against all 14 watchlists and no matches were found.",
        "watchlist.value.metadata.length === 0  (after deduplication by combineWatchlistMetadata)",
        "Depends on: watchlist fact. ALL hits (business + person) are now consolidated in watchlist.value.metadata via Trulioo advanced watchlist screening (since BEST-65). No separate /people/watchlist endpoint needed. Source: WatchlistsTab.tsx L207, combineWatchlistMetadata rule in rules.ts L273-344.",
        "No action required for this entity. Normal expected state."
    ),
    "watchlist_hits": (
        "🔴 N Hits Found", "badge-hits",
        "This entity was found on one or more government watchlists (OFAC, BIS, State Dept).",
        "watchlist.value.metadata.length > 0  (N = count after deduplication)",
        "Depends on: watchlist.value.metadata array. Each hit has: {type, metadata:{title,agency}, entity_name, url, entity_type:'BUSINESS'|'PERSON', list_country}. Hits deduplicated by type|title|entity_name|url key. ADVERSE MEDIA is filtered out (it goes to Public Records tab). Source: WatchlistsTab.tsx L207, WatchlistHitCard.tsx.",
        "UCM Rule: IMMEDIATE REVIEW required. Check each hit: (a) Review the source URL/agency, (b) Determine if it is the same entity or a false positive (common name), (c) Escalate to compliance team if confirmed match. Do NOT auto-approve cases with active watchlist hits."
    ),
    # Website
    "website_online": (
        "✅ Online", "badge-online",
        "The business website is live and accessible. SERP/Verdata successfully loaded the URL.",
        "websiteData.status === 'online'",
        "Depends on: websiteData.status from GET /verification/businesses/:id/website-data endpoint. SERP scraper or Verdata checks if the URL responds with HTTP 200. Source: useWebsiteNonEditableFields.tsx L105-115.",
        "Positive signal. Website confirms business is operational. Review the website content for consistency with submitted industry/NAICS."
    ),
    "website_offline": (
        "❌ Offline", "badge-offline",
        "The business website URL was found but is not currently accessible or returning errors.",
        "websiteData.status === 'offline'  (or HTTP error response)",
        "Depends on: websiteData.status. SERP/Verdata checked the URL but received an error or no response. Source: useWebsiteNonEditableFields.tsx L108.",
        "Risk signal. Could indicate: (1) website temporarily down, (2) business no longer operating, (3) domain expired. Check domain expiration date. Cross-reference with SoS filing status."
    ),
    "website_unknown": (
        "Unknown", "badge-unknown",
        "Website status could not be determined — either no URL was found or the status check did not complete.",
        "websiteData === null  OR  websiteData.status is absent  OR  isWebsiteDirty === true",
        "Depends on: websiteData.status and isWebsiteDirty state. Shows 'Unknown' when: (1) no website URL was found/submitted, (2) URL was just edited by analyst (isWebsiteDirty=true clears dependent fields), (3) SERP/Verdata integration not yet complete. Source: useWebsiteNonEditableFields.tsx L56-65.",
        "If URL was just edited: wait for re-verification to complete. If no URL found: risk signal for established businesses — most legitimate businesses have a web presence."
    ),
    "website_match": (
        "✅ Match", "badge-match",
        "Google Business Profile confirms both the business name and address match the submitted information.",
        "googleProfileData.data.business_match === 'match found'  AND  address_match === 'match'",
        "Depends on: Google Places API result via SERP-found google_place_id. Both name and address must match. Source: AddressesCard.tsx L221-228.",
        "Strong positive signal. Independent third-party (Google) confirms business identity and location."
    ),
    "website_no_match": (
        "❌ No Match", "badge-nomatch",
        "Google Business Profile found but shows different business name or address than submitted.",
        "googleProfileData.data.business_match !== 'match found'  OR  address_match !== 'match'",
        "Depends on: Google Places API. Either name or address (or both) do not match. Source: AddressesCard.tsx L221-228.",
        "Review discrepancy: check if the Google listing is outdated, if the business recently moved, or if there is a different business at the same address."
    ),
    # Case status
    "case_pending": (
        "Pending", "badge-unknown",
        "Integrations are still running. Vendor API calls have not all completed yet.",
        "data_integration_tasks_progress.is_complete === false",
        "Depends on: is_integration_complete flag from case-management.ts getCaseByIDQuery L1570-1574. The 'Integrations are currently processing' orange banner is shown while is_complete=false. Transitions to complete when all vendor tasks finish. Source: case-management.ts L1570.",
        "Wait for integrations to complete before reviewing. Worth Score shows '-' until all data is collected."
    ),
    "case_under_review": (
        "Under Review", "badge-unverified",
        "One or more UCM decisioning rules failed. Analyst manual review is required.",
        "data_cases.status → core_case_statuses.code = 'under_review'",
        "Depends on: Worth Score calculation and UCM rule evaluation. Triggered when: (a) TIN Unverified, (b) watchlist hits found, (c) IDV failed for an owner, (d) Worth Score below auto-approve threshold, (e) other configured rules fail. Source: case-management.ts.",
        "Analyst must review each failed rule and make a determination. Case will remain Under Review until manually moved to Approved or Archived."
    ),
    "case_auto_approved": (
        "Auto-Approved", "badge-verified",
        "All UCM decisioning rules passed AND the Worth Score is above the configured auto-approval threshold.",
        "data_cases.status → core_case_statuses.code = 'auto_approved'",
        "Depends on: UCM rule engine + Worth Score threshold (configured per customer/integration). All checks must pass: TIN verified, no watchlist hits, IDV passed for all required owners, address verified, score above threshold. Source: worth score engine + case-service.",
        "Case is approved for processing. No analyst action required unless exception review is triggered."
    ),
    "case_archived": (
        "Archived", "badge-unknown",
        "The case has been closed and is no longer active.",
        "data_cases.status → core_case_statuses.code = 'archived'",
        "Depends on: manual analyst action (PATCH /cases/:id) or automated archival rule. Archived cases cannot be reopened without explicit action. Source: case-management.ts.",
        "Case is closed. No further action unless reopened. Historical data is preserved."
    ),
    # Ownership flags
    "minority_yes": (
        "Yes", "badge-verified",
        "Equifax data indicates this business is a Minority Business Enterprise (MBE).",
        "efx_mbe = 'Y'  in warehouse.equifax_us_latest",
        "Depends on: efx_mbe column from Equifax (pid=17, w=0.7). Equifax sets this flag based on business registration data and self-reported information. Source: businessDetails/index.ts minority_owned fact, equifax_us_latest.efx_mbe.",
        "Informational. Confirm via official MBE certification if required for compliance."
    ),
    "minority_no": (
        "No", "badge-missing",
        "Equifax data indicates this business is NOT a Minority Business Enterprise.",
        "efx_mbe ≠ 'Y'  in warehouse.equifax_us_latest",
        "Depends on: efx_mbe column from Equifax. Value is 'N' or absent.",
        "Informational. If merchant claims MBE status but field shows No, request official MBE certification documentation."
    ),
    "minority_na": (
        "N/A", "badge-unknown",
        "Equifax does not have MBE data for this business, or the business was not matched in Equifax.",
        "minority_owned fact value is null/empty  OR  no Equifax match found",
        "Depends on: Equifax entity match (efx_matches_custom_inc_ml). If business has no EFX match, the minority_owned fact will be null. Source: businessDetails/index.ts minority_owned, woman_owned, veteran_owned facts.",
        "Cannot determine ownership status from Equifax. Consider requesting self-certification documentation from the merchant if MBE/WBE/VBE status is required."
    ),
    # IDV
    "idv_verified": (
        "✅ Verified (IDV)", "badge-verified",
        "Identity verification (Plaid IDV) passed for this owner — Plaid confirmed SSN, DOB, and address against government/credit records.",
        "idv_passed_boolean.value === true  (idv_status.SUCCESS > 0)",
        "Depends on: idv_status fact (aggregated counts: {SUCCESS:N, FAILED:N, PENDING:N}) and idv_passed_boolean = idv_status.SUCCESS > 0. Plaid IDV checks: SSN via credit header, DOB match, name match, address match. Source: kyb/index.ts L493-552.",
        "No action required. Owner identity confirmed."
    ),
    "idv_failed": (
        "❌ Failed (IDV)", "badge-missing",
        "Identity verification (Plaid IDV) failed — Plaid could not confirm the owner's identity against records.",
        "idv_status.FAILED > 0  AND  idv_status.SUCCESS === 0",
        "Depends on: idv_status.FAILED count. Plaid IDV fails when: (1) SSN not found in credit header, (2) DOB mismatch, (3) name mismatch, (4) identity appears synthetic or stolen. Source: kyb/index.ts L493-527, plaid IDV response.",
        "UCM Rule: FAIL. Escalate to fraud/compliance team. Do NOT approve. Consider requesting government-issued photo ID for manual verification."
    ),
    "idv_pending": (
        "⏳ Pending (IDV)", "badge-unknown",
        "Owner has not yet completed the Plaid IDV flow — the IDV link was sent but not clicked, or is in progress.",
        "idv_status.PENDING > 0  AND  idv_status.SUCCESS === 0",
        "Depends on: idv_status.PENDING count. Pending when: (1) IDV email/SMS link not yet clicked, (2) owner started but did not complete, (3) link expired. Source: kyb/index.ts L493-527.",
        "Resend the IDV link if expired. Follow up with the merchant to ensure all owners complete IDV before approval."
    ),
}


def enrich_badges(badges):
    """Convert any badge text in BADGE_LIBRARY to the full 6-tuple format."""
    result = []
    for b in badges:
        if len(b) >= 6:
            result.append(b)
        else:
            btext, bclass, bexplain = b[0], b[1], b[2]
            # Try to find a matching library entry by class+text similarity
            match = None
            for key, lib_entry in BADGE_LIBRARY.items():
                if lib_entry[1] == bclass and lib_entry[0].lower() in btext.lower():
                    match = lib_entry
                    break
            if match:
                result.append(match)
            else:
                result.append((btext, bclass, bexplain, None, None, None))
    return result


RS_CONN = (
    "import psycopg2, json\n\n"
    "conn = psycopg2.connect(\n"
    "    host='your-cluster.redshift.amazonaws.com',\n"
    "    port=5439, dbname='dev',\n"
    "    user='your_username', password='your_password',\n"
    "    sslmode='require'\n"
    ")\n"
    "BID = 'YOUR-BUSINESS-UUID-HERE'  # replace with your business UUID\n"
)

# ── Master SQL/Python generator — used by render_field() for EVERY field ──────
ARRAY_FACTS = {
    'sos_filings','addresses','addresses_deliverable','addresses_found',
    'people','names_found','names_submitted','dba_found',
    'watchlist','watchlist_raw','business_addresses_submitted_strings',
    'classification_codes','owners',
}
SKIP_NAMES = {
    'rds_warehouse_public','rds_cases_public','integration_data','datascience',
    'data_cases','core_case_statuses','core_naics_code','core_mcc_code',
    'core_business_industries','websiteData','customer_files','calculated',
}
PID_MAP_COMMENT = "-- winning_vendor: 16=Middesk 24=ZoomInfo 23=OC 17=Equifax 38=Trulioo 22=SERP 31=AI"

# Extra SQL for facts with nested value objects
NESTED_EXTRAS = {
    'tin_match': (
        "\n-- tin_match: value is a NESTED object {status, message, sublabel}\n"
        "SELECT\n"
        "    JSON_EXTRACT_PATH_TEXT(value,'value','status')   AS tin_status,\n"
        "    JSON_EXTRACT_PATH_TEXT(value,'value','message')  AS tin_message,\n"
        "    JSON_EXTRACT_PATH_TEXT(value,'value','sublabel') AS tin_sublabel\n"
        "FROM rds_warehouse_public.facts\n"
        "WHERE business_id = 'YOUR-BUSINESS-UUID-HERE' AND name = 'tin_match';\n\n"
        "-- tin_match_boolean: true=Verified badge, false=Unverified badge\n"
        "SELECT JSON_EXTRACT_PATH_TEXT(value,'value') AS tin_verified\n"
        "FROM rds_warehouse_public.facts\n"
        "WHERE business_id = 'YOUR-BUSINESS-UUID-HERE' AND name = 'tin_match_boolean';\n\n"
        "-- tin: raw masked EIN submitted by applicant\n"
        "SELECT JSON_EXTRACT_PATH_TEXT(value,'value') AS tin_masked\n"
        "FROM rds_warehouse_public.facts\n"
        "WHERE business_id = 'YOUR-BUSINESS-UUID-HERE' AND name = 'tin';"
    ),
    'naics_code': (
        "\n-- naics_code: fetch value column to inspect all vendor alternatives[]\n"
        "SELECT name, value FROM rds_warehouse_public.facts\n"
        "WHERE business_id = 'YOUR-BUSINESS-UUID-HERE' AND name = 'naics_code';\n"
        "-- In Python: fact=json.loads(value); alts=fact.get('alternatives',[])\n"
        "-- Each alt has: platformId, value, confidence -> sorted desc = winner selection"
    ),
    'mcc_code': (
        "\n-- All MCC facts: mcc_code = mcc_code_found ?? mcc_code_from_naics\n"
        "SELECT name, JSON_EXTRACT_PATH_TEXT(value,'value') AS mcc_value\n"
        "FROM rds_warehouse_public.facts\n"
        "WHERE business_id = 'YOUR-BUSINESS-UUID-HERE'\n"
        "  AND name IN ('mcc_code','mcc_code_found','mcc_code_from_naics','mcc_description');"
    ),
    'address_verification': (
        "\n-- address_verification: value is a nested object {status, sublabel, message}\n"
        "SELECT\n"
        "    JSON_EXTRACT_PATH_TEXT(value,'value','status')   AS addr_status,\n"
        "    JSON_EXTRACT_PATH_TEXT(value,'value','sublabel') AS addr_sublabel,\n"
        "    JSON_EXTRACT_PATH_TEXT(value,'value','message')  AS addr_message\n"
        "FROM rds_warehouse_public.facts\n"
        "WHERE business_id = 'YOUR-BUSINESS-UUID-HERE' AND name = 'address_verification';"
    ),
    'name_match': (
        "\n-- name_match: value is a nested object {status, message, sublabel}\n"
        "SELECT\n"
        "    JSON_EXTRACT_PATH_TEXT(value,'value','status')   AS name_status,\n"
        "    JSON_EXTRACT_PATH_TEXT(value,'value','message')  AS name_message,\n"
        "    JSON_EXTRACT_PATH_TEXT(value,'value','sublabel') AS name_sublabel\n"
        "FROM rds_warehouse_public.facts\n"
        "WHERE business_id = 'YOUR-BUSINESS-UUID-HERE' AND name = 'name_match';"
    ),
    'idv_status': (
        "\n-- idv_status: value is a nested object {SUCCESS:N, FAILED:N, PENDING:N}\n"
        "SELECT\n"
        "    JSON_EXTRACT_PATH_TEXT(value,'value','SUCCESS') AS idv_success,\n"
        "    JSON_EXTRACT_PATH_TEXT(value,'value','FAILED')  AS idv_failed,\n"
        "    JSON_EXTRACT_PATH_TEXT(value,'value','PENDING') AS idv_pending\n"
        "FROM rds_warehouse_public.facts\n"
        "WHERE business_id = 'YOUR-BUSINESS-UUID-HERE' AND name = 'idv_status';"
    ),
}

# Python array parsers per fact
ARRAY_PARSERS = {
    'sos_filings': (
        "filings = fact.get('value', [])\n"
        "    print(f'  {len(filings)} SoS filing(s):')\n"
        "    for fil in filings:\n"
        "        print(f'    State:{fil.get(\"state\")} Active:{fil.get(\"active\")} '\n"
        "              f'Entity:{fil.get(\"entity_type\")} Date:{fil.get(\"filing_date\")} '\n"
        "              f'Name:{fil.get(\"filing_name\")}')"
    ),
    'addresses': (
        "addrs = fact.get('value', [])\n"
        "    print(f'  {len(addrs)} address(es):')\n"
        "    for a in addrs: print(f'    {a}')"
    ),
    'addresses_deliverable': (
        "deliverable = fact.get('value', [])\n"
        "    print(f'  Deliverable addresses: {deliverable}')"
    ),
    'people': (
        "people = fact.get('value', [])\n"
        "    print(f'  {len(people)} person(s):')\n"
        "    for p in people:\n"
        "        print(f'    {p.get(\"name\")} — titles:{p.get(\"titles\")}')"
    ),
    'watchlist': (
        "hits = fact.get('value', {}).get('metadata', [])\n"
        "    print(f'  Total watchlist hits: {len(hits)}')\n"
        "    for h in hits:\n"
        "        print(f'    {h.get(\"entity_name\")} — {h.get(\"metadata\",{}).get(\"title\")}')"
    ),
    'names_found': (
        "names_list = fact.get('value', [])\n"
        "    print(f'  Names found: {names_list}')"
    ),
    'dba_found': (
        "dbas = fact.get('value', [])\n"
        "    print(f'  DBA names: {dbas}')"
    ),
}

# Python extra prints for nested scalar facts
SCALAR_EXTRAS_PY = {
    'tin_match': (
        "    tin_val = fact.get('value', {})\n"
        "    if isinstance(tin_val, str): tin_val = json.loads(tin_val)\n"
        "    print(f'  tin_status:  {tin_val.get(\"status\")}  (success=Verified badge)')\n"
        "    print(f'  tin_message: {tin_val.get(\"message\")}')"
    ),
    'naics_code': (
        "    pid_map={'16':'Middesk','23':'OC','24':'ZoomInfo','17':'EFX','38':'Trulioo','22':'SERP','31':'AI'}\n"
        "    alts=fact.get('alternatives',[])\n"
        "    print(f'  ALL {len(alts)} vendor responses (sorted by confidence):')\n"
        "    for a in sorted(alts, key=lambda x: float(x.get('confidence') or 0), reverse=True):\n"
        "        v=pid_map.get(str(a.get('platformId','')),f'pid={a.get(\"platformId\")}')\n"
        "        print(f'    {v:12} conf={str(a.get(\"confidence\",\"N/A\")):6} naics={a.get(\"value\")}')"
    ),
    'address_verification': (
        "    av=fact.get('value',{})\n"
        "    if isinstance(av,str): av=json.loads(av)\n"
        "    print(f'  status:{av.get(\"status\")} sublabel:{av.get(\"sublabel\")}')"
    ),
    'idv_status': (
        "    counts=fact.get('value',{})\n"
        "    if isinstance(counts,str): counts=json.loads(counts)\n"
        "    print(f'  SUCCESS:{counts.get(\"SUCCESS\")} FAILED:{counts.get(\"FAILED\")} PENDING:{counts.get(\"PENDING\")}')"
    ),
    'name_match': (
        "    nm=fact.get('value',{})\n"
        "    if isinstance(nm,str): nm=json.loads(nm)\n"
        "    status=nm.get('status') if isinstance(nm,dict) else nm\n"
        "    print(f'  name_match status:{status}')\n"
        "    if isinstance(nm,dict): print(f'  message:{nm.get(\"message\")}')"
    ),
}


import re as _re

def _parse_fact_names(raw):
    """Extract real Redshift fact names from the fact field string."""
    tokens = _re.split(r'[/+,]', raw)
    names = []
    for t in tokens:
        c = t.strip().split('[')[0].split('.')[0].split(' ')[0].strip("'\" ")
        if c and _re.match(r'^[a-z_][a-z0-9_]*$', c) and c not in SKIP_NAMES:
            if c not in names:
                names.append(c)
    return names


def generate_sql(f):
    """Generate the confirmed-working Redshift SQL for any field."""
    label   = f.get("label", "")
    raw     = f.get("fact", "")
    names   = _parse_fact_names(raw)
    BID     = "'YOUR-BUSINESS-UUID-HERE'"

    if not names:
        # Website/domain data — lives in /website-data endpoint, not facts table
        return (
            f"-- {label}\n"
            "-- This field comes from GET /verification/businesses/:id/website-data\n"
            "-- It is NOT stored in rds_warehouse_public.facts\n"
            "-- No SQL query — fetch via the API endpoint or check integration_data.request_response:\n"
            "SELECT response FROM integration_data.request_response\n"
            f"WHERE business_id = {BID} AND platform_id = 22\n"
            "ORDER BY requested_at DESC LIMIT 1;"
        )

    is_arr  = any(n in ARRAY_FACTS for n in names)
    names_q = ', '.join(f"'{n}'" for n in names[:5])
    extra   = next((NESTED_EXTRAS[n] for n in names if n in NESTED_EXTRAS), '')

    if is_arr:
        return (
            f"-- {label} (JSON array stored in value column)\n"
            f"-- Returns rows AND full JSON value — parse 'value' column with Python below\n"
            f"SELECT name, value, received_at\n"
            f"FROM rds_warehouse_public.facts\n"
            f"WHERE business_id = {BID}\n"
            f"  AND name IN ({names_q});\n"
            f"-- Copy the 'value' text output and parse with: json.loads(value)"
        )

    return (
        f"-- {label}\n"
        f"{PID_MAP_COMMENT}\n"
        f"SELECT\n"
        f"    name,\n"
        f"    JSON_EXTRACT_PATH_TEXT(value, 'value')                AS fact_value,\n"
        f"    JSON_EXTRACT_PATH_TEXT(value, 'source', 'platformId') AS winning_vendor,\n"
        f"    JSON_EXTRACT_PATH_TEXT(value, 'source', 'confidence') AS vendor_conf,\n"
        f"    JSON_EXTRACT_PATH_TEXT(value, 'override', 'value')    AS analyst_override,\n"
        f"    JSON_EXTRACT_PATH_TEXT(value, 'override', 'userId')   AS overridden_by,\n"
        f"    received_at\n"
        f"FROM rds_warehouse_public.facts\n"
        f"WHERE business_id = {BID}\n"
        f"  AND name IN ({names_q});"
        + extra
    )


def generate_python(f):
    """Generate field-specific Python for any field."""
    label  = f.get("label", "")
    raw    = f.get("fact", "")
    names  = _parse_fact_names(raw)
    BID    = "'YOUR-BUSINESS-UUID-HERE'"

    if not names:
        return (
            "import psycopg2, json\n\n"
            "conn = psycopg2.connect(\n"
            "    host='your-cluster.redshift.amazonaws.com', port=5439,\n"
            "    dbname='dev', user='your_username', password='your_password', sslmode='require'\n"
            ")\n"
            f"# {label} — fetch from /website-data API endpoint\n"
            "# or check raw SERP response:\n"
            "cur = conn.cursor()\n"
            f"cur.execute('SELECT response FROM integration_data.request_response '\n"
            f"            'WHERE business_id = %s AND platform_id = 22 '\n"
            f"            'ORDER BY requested_at DESC LIMIT 1', ({BID},))\n"
            "row = cur.fetchone()\n"
            "if row: print(json.loads(row[0]))\n"
            "cur.close(); conn.close()"
        )

    is_arr  = any(n in ARRAY_FACTS for n in names)
    names_q = ', '.join(f"'{n}'" for n in names[:5])

    conn_block = (
        "import psycopg2, json\n\n"
        "conn = psycopg2.connect(\n"
        "    host='your-cluster.redshift.amazonaws.com', port=5439,\n"
        "    dbname='dev', user='your_username', password='your_password', sslmode='require'\n"
        ")\n"
        f"BID = {BID}  # replace\n"
        "cur = conn.cursor()\n"
    )

    fetch = (
        f"cur.execute(\n"
        f"    'SELECT name, value, received_at FROM rds_warehouse_public.facts '\n"
        f"    'WHERE business_id = %s AND name IN ({names_q})',\n"
        f"    (BID,)\n"
        f")\n"
        f"rows = cur.fetchall()\n"
        f"print(f'--- {label}: {{len(rows)}} row(s) found ---')\n"
    )

    if is_arr:
        arr_name = next((n for n in names if n in ARRAY_PARSERS), None)
        parse = ARRAY_PARSERS.get(arr_name, "print(fact.get('value'))")
        return (
            conn_block + fetch +
            "for name, value_text, ts in rows:\n"
            "    fact = json.loads(value_text)\n"
            "    print(f'  === {name} (received {ts})')\n"
            f"    {parse}\n"
            "cur.close(); conn.close()"
        )

    extra_blocks = ''.join(
        f"\n    # --- {n} nested fields ---\n{SCALAR_EXTRAS_PY[n]}"
        for n in names if n in SCALAR_EXTRAS_PY
    )
    return (
        conn_block + fetch +
        "for name, value_text, ts in rows:\n"
        "    fact = json.loads(value_text)  # value is varchar — always json.loads()\n"
        "    val = fact.get('value')\n"
        "    src = fact.get('source', {})\n"
        "    pid_map = {'16':'Middesk','24':'ZoomInfo','23':'OC','17':'Equifax',\n"
        "               '38':'Trulioo','22':'SERP','31':'AI'}\n"
        "    vendor = pid_map.get(str(src.get('platformId','')), f'pid={src.get(\"platformId\")}')\n"
        "    print(f'  {name}:')\n"
        "    print(f'    fact_value    : {val}')\n"
        "    print(f'    winning_vendor: {vendor}  (pid={src.get(\"platformId\")})')\n"
        "    print(f'    confidence    : {src.get(\"confidence\")}')\n"
        "    print(f'    override      : {fact.get(\"override\")}')\n"
        "    print(f'    received_at   : {ts}')"
        + extra_blocks + "\n"
        "cur.close(); conn.close()"
    )

def build_fact_specific_cols(f):
    """Extra SQL block for fields that have special structure."""
    fact = f.get("fact","").lower()
    label = f.get("label","")
    if "naics" in fact and "code" in fact:
        return (
            "\n-- All vendor NAICS alternatives (what every vendor returned before winner chosen):\n"
            "-- Fetch raw value and parse in Python — see Python snippet below"
        )
    if "mcc" in fact:
        return (
            "\n-- mcc_code_found (AI direct) vs mcc_code_from_naics (calculated):\n"
            "SELECT name, JSON_EXTRACT_PATH_TEXT(value, 'value') AS mcc\n"
            "FROM rds_warehouse_public.facts\n"
            "WHERE business_id = 'YOUR-BUSINESS-UUID-HERE'\n"
            "  AND name IN ('mcc_code', 'mcc_code_found', 'mcc_code_from_naics', 'mcc_description');"
        )
    if "tin_match" in fact:
        return (
            "\n-- tin_match has a NESTED object as its value (not a simple string):\n"
            "-- fact_value will show: {\"status\":\"success\",\"message\":\"...\",\"sublabel\":\"...\"}\n"
            "-- Use this query to unpack the nested fields:\n"
            "SELECT\n"
            "    JSON_EXTRACT_PATH_TEXT(value, 'value', 'status')   AS tin_status,\n"
            "    JSON_EXTRACT_PATH_TEXT(value, 'value', 'message')  AS tin_message,\n"
            "    JSON_EXTRACT_PATH_TEXT(value, 'value', 'sublabel') AS tin_sublabel,\n"
            "    JSON_EXTRACT_PATH_TEXT(value, 'source', 'platformId') AS source_pid\n"
            "FROM rds_warehouse_public.facts\n"
            "WHERE business_id = 'YOUR-BUSINESS-UUID-HERE'\n"
            "  AND name = 'tin_match';\n\n"
            "-- tin_match_boolean = simpler: just true or false\n"
            "-- true  → admin portal shows ✅ Verified badge\n"
            "-- false → admin portal shows ⚠️ Unverified badge\n"
            "SELECT JSON_EXTRACT_PATH_TEXT(value, 'value') AS tin_verified\n"
            "FROM rds_warehouse_public.facts\n"
            "WHERE business_id = 'YOUR-BUSINESS-UUID-HERE'\n"
            "  AND name = 'tin_match_boolean';\n\n"
            "-- tin = the raw masked EIN submitted by the applicant\n"
            "SELECT JSON_EXTRACT_PATH_TEXT(value, 'value') AS tin_masked\n"
            "FROM rds_warehouse_public.facts\n"
            "WHERE business_id = 'YOUR-BUSINESS-UUID-HERE'\n"
            "  AND name = 'tin';"
        )
    return ""


def build_redshift_sql(f, BID):
    """Build Redshift SQL for datascience.customer_files fields."""
    label = f.get("label","")
    if "naics" in label.lower():
        return (
            f"-- {label} — Pipeline B (Redshift winner-takes-all)\n"
            f"SELECT business_id,\n"
            f"       primary_naics_code,\n"
            f"       zi_match_confidence,\n"
            f"       efx_match_confidence,\n"
            f"       CASE WHEN zi_match_confidence > efx_match_confidence\n"
            f"            THEN 'ZoomInfo (pid=24)'\n"
            f"            ELSE 'Equifax (pid=17)'\n"
            f"       END AS pipeline_b_winner\n"
            f"FROM datascience.customer_files\n"
            f"WHERE business_id = {BID};"
        )
    return (
        f"-- {label} — Redshift customer_files\n"
        f"SELECT business_id, primary_naics_code, mcc_code, worth_score,\n"
        f"       zi_match_confidence, efx_match_confidence,\n"
        f"       employee_count, year_established\n"
        f"FROM datascience.customer_files\n"
        f"WHERE business_id = {BID};"
    )


def build_python_for_field(f, fact_names, is_array, is_redshift, bid_placeholder):
    """Build field-specific Python that fetches and parses the right fact(s)."""
    label  = f.get("label","this field")
    fact   = f.get("fact","")
    names_tuple = ", ".join(f"'{n}'" for n in fact_names[:3] if n)

    conn_block = (
        "conn = psycopg2.connect(\n"
        "    host='your-cluster.redshift.amazonaws.com',\n"
        "    port=5439, dbname='dev',\n"
        "    user='your_username', password='your_password',\n"
        "    sslmode='require'\n"
        ")"
    )

    if is_redshift:
        if "naics" in label.lower():
            return (
                RS_CONN + conn_block + "\n\n"
                "cur = conn.cursor()\n"
                f"cur.execute(\n"
                f"    'SELECT business_id, primary_naics_code, mcc_code, '\n"
                f"    'zi_match_confidence, efx_match_confidence, '\n"
                f"    'CASE WHEN zi_match_confidence > efx_match_confidence '\n"
                f"    'THEN \\'ZoomInfo\\' ELSE \\'Equifax\\' END AS winner '\n"
                f"    'FROM datascience.customer_files WHERE business_id = %s',\n"
                f"    (BID,)\n"
                f")\n"
                "row = cur.fetchone()\n"
                "if row: print(dict(zip([d[0] for d in cur.description], row)))\n"
                "cur.close(); conn.close()"
            )
        return (
            RS_CONN + conn_block + "\n\n"
            "import pandas as pd\n"
            "df = pd.read_sql(\n"
            "    'SELECT * FROM datascience.customer_files WHERE business_id = %s',\n"
            "    conn, params=(BID,)\n"
            ")\n"
            "print(df.to_string())\n"
            "conn.close()"
        )

    if is_array:
        # Array fact — fetch raw, parse with json.loads
        array_parse = {
            "sos_filings": (
                "filings = fact.get('value', [])\n"
                "for f in filings:\n"
                "    print(f'  State: {f.get(\"state\")}  Active: {f.get(\"active\")}'\n"
                "          f'  Entity: {f.get(\"entity_type\")}  Date: {f.get(\"filing_date\")}')"
            ),
            "addresses": (
                "addrs = fact.get('value', [])\n"
                "for a in addrs:\n"
                "    print(f'  {a}')"
            ),
            "people": (
                "people = fact.get('value', [])\n"
                "for p in people:\n"
                "    print(f'  {p.get(\"name\")} — {p.get(\"titles\")}')"
            ),
            "watchlist": (
                "hits = fact.get('value', {}).get('metadata', [])\n"
                "print(f'Total hits: {len(hits)}')\n"
                "for h in hits:\n"
                "    print(f'  {h.get(\"entity_name\")} — {h.get(\"metadata\",{}).get(\"title\")}')"
            ),
            "names": (
                "names = fact.get('value', [])\n"
                "for n in names:\n"
                "    print(f'  {n}')"
            ),
        }
        parse_block = "# Parse the JSON array\nprint(fact)"
        for key, block in array_parse.items():
            if key in fact.lower():
                parse_block = block
                break

        return (
            RS_CONN + conn_block + "\n\n"
            f"# Fetch: {label}\n"
            f"cur = conn.cursor()\n"
            f"cur.execute(\n"
            f"    'SELECT name, value, received_at FROM rds_warehouse_public.facts '\n"
            f"    'WHERE business_id = %s AND name IN ({names_tuple})',\n"
            f"    (BID,)\n"
            f")\n"
            f"for row in cur.fetchall():\n"
            f"    name, value_text, ts = row\n"
            f"    fact = json.loads(value_text)\n"
            f"    print(f'=== {{name}} (received {{ts}}) ===')\n"
            f"    {parse_block.replace(chr(10), chr(10)+'    ')}\n"
            f"cur.close(); conn.close()"
        )
    else:
        # Scalar fact — JSON_EXTRACT_PATH_TEXT for SQL, json.loads for Python
        # Special handling per fact type
        extra_print = ""
        if "naics" in fact.lower() and "code" in fact.lower():
            extra_print = (
                "    # Show all vendor alternatives (what every vendor returned):\n"
                "    alts = fact.get('alternatives', [])\n"
                "    print(f'  All vendor alternatives ({len(alts)} vendors):')\n"
                "    for a in alts:\n"
                "        pid_map = {'16':'Middesk','23':'OC','24':'ZoomInfo','17':'Equifax','38':'Trulioo','31':'AI','22':'SERP'}\n"
                "        vendor = pid_map.get(str(a.get('platformId','')), f'pid={a.get(\"platformId\")}')\n"
                "        print(f'    {vendor:12} conf={str(a.get(\"confidence\",\"N/A\")):6} naics={a.get(\"value\")}')\n"
            )
        elif "tin_match" in fact.lower():
            extra_print = (
                "    # TIN match has nested object in value field:\n"
                "    tin_val = fact.get('value', {})\n"
                "    if isinstance(tin_val, str): tin_val = json.loads(tin_val)\n"
                "    print(f'  TIN status:  {tin_val.get(\"status\")}')\n"
                "    print(f'  TIN message: {tin_val.get(\"message\")}')\n"
            )
        elif "watchlist_hits" in fact.lower():
            extra_print = (
                "    print(f'  Hit count: {fact.get(\"value\")}')\n"
            )

        return (
            RS_CONN + conn_block + "\n\n"
            f"# Fetch: {label}\n"
            f"cur = conn.cursor()\n"
            f"cur.execute(\n"
            f"    'SELECT name, value, received_at FROM rds_warehouse_public.facts '\n"
            f"    'WHERE business_id = %s AND name IN ({names_tuple})',\n"
            f"    (BID,)\n"
            f")\n"
            f"for row in cur.fetchall():\n"
            f"    name, value_text, ts = row\n"
            f"    fact = json.loads(value_text)  # value is varchar — parse with json.loads()\n"
            f"    print(f'=== {{name}} ===')\n"
            f"    print(f'  Value:     {{fact.get(\"value\")}}')\n"
            f"    print(f'  PID:       {{fact.get(\"source\",{{}}).get(\"platformId\")}}  (winning vendor)')\n"
            f"    print(f'  Conf:      {{fact.get(\"source\",{{}}).get(\"confidence\")}}  (vendor confidence)')\n"
            f"    print(f'  Override:  {{fact.get(\"override\")}}  (analyst manual edit if any)')\n"
            f"    print(f'  Received:  {{ts}}')\n"
            f"    {extra_print}"
            f"cur.close(); conn.close()"
        )


def render_field(f):
    """Render a single field card with full deep lineage in an expander."""
    edit_icon = "✏️ Editable" if f.get("editable") else "🔒 Read-Only"
    col = GREEN if f.get("editable") else GREY
    st.markdown(
        f'<div class="field-row">'
        f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:4px;">'
        f'<span style="color:{col};font-weight:700;">{edit_icon}</span>'
        f'<span style="color:#E2E8F0;font-weight:700;font-size:.95rem;">{f["label"]}</span>'
        f'{"".join(badge(t, "grey") for t in f.get("tags",[]))}'
        f'</div>'
        f'<div style="font-size:.78rem;color:#475569;">'
        f'Fact name in rds_warehouse_public.facts: <code>{f["fact"]}</code> · API endpoint: <code>{f["api"]}</code>'
        f'</div>'
        f'</div>', unsafe_allow_html=True)

    with st.expander(f"🔍 Full lineage — {f['label']}", expanded=False):

        # ── Column legend (always shown) ──────────────────────────────
        st.markdown(
            '<div class="card card-teal" style="font-size:.8rem;padding:8px 14px;margin-bottom:8px;">'
            '<b>📖 Column meanings:</b><br>'
            '<b>Source</b> = which vendor/system provides this value.<br>'
            '<b>PID</b> = platform_id in integration_data.request_response table — the vendor identifier used throughout Worth AI.<br>'
            '<b>Weight</b> = Worth\'s static trust score for this vendor (set in sources.ts). Used ONLY as tie-break when two sources have confidence within 5% (WEIGHT_THRESHOLD=0.05). NOT from XGBoost.<br>'
            '<b>Confidence</b> = how sure the vendor is about its match quality. Each vendor computes this differently: '
            'Middesk=task-based (0.15+0.20×tasks), ZI/OC/EFX=match.index/55 (XGBoost entity-match), AI=GPT self-reported.<br>'
            '<b>Data field</b> = the exact path in the vendor\'s response object where this value is read from.<br>'
            '<b>Winner rule</b> = factWithHighestConfidence() compares confidences; if gap ≤5% → weightedFactSelector() uses Weight as tie-break. manualOverride() always runs first. Rule 4: no minimum confidence cutoff.'
            '</div>', unsafe_allow_html=True)

        c1, c2 = st.columns([3, 2])
        with c1:
            if f.get("sources"):
                src_html = ('<table class="t"><tr>'
                            '<th title="Vendor or system that provides this value">Source</th>'
                            '<th title="platform_id in integration_data.request_response — vendor identifier">PID</th>'
                            '<th title="Static trust score (sources.ts). Used for tie-breaking when confidence gap ≤5%. NOT from XGBoost.">Weight ⓘ</th>'
                            '<th title="Exact field path in vendor response + how confidence is computed">Data field + Confidence model</th>'
                            '</tr>')
                for name, pid, w, detail in f["sources"]:
                    wn = float(w.replace("w=","").replace("w≈","")) if any(str(w).startswith(p) for p in ["w=","w≈"]) else 0
                    wc = "#FCD34D" if wn>=2 else ("#6EE7B7" if wn>=0.8 else ("#FCD34D" if wn>=0.5 else "#94A3B8"))
                    pid_note = " (vendor)" if pid and pid.startswith("pid=") else (" (calculated — no vendor)" if "calculated" in str(pid).lower() else "")
                    src_html += (f"<tr><td style='color:#E2E8F0;font-weight:600'>{name}</td>"
                                 f"<td><code>{pid}</code><span style='color:#475569;font-size:.72rem'>{pid_note}</span></td>"
                                 f"<td><b style='color:{wc}'>{w}</b></td>"
                                 f"<td style='color:#94A3B8;font-size:.78rem'>{detail}</td></tr>")
                st.markdown(src_html + "</table>", unsafe_allow_html=True)

            if f.get("rule"):
                st.markdown("**🏆 Winner selection — full mechanics:**")
                card(f"{f['rule']}", "card-green")

            if f.get("editable") is not None:
                editable_explain = (
                    "<b>✏️ Editable = TRUE:</b> Analyst can override this value in the admin portal. "
                    "Override stored in facts JSONB: override: {value, userId, timestamp}. "
                    "manualOverride() rule fires first and wins unconditionally over all vendor data."
                ) if f.get("editable") else (
                    "<b>🔒 Read-Only = TRUE:</b> This field is NOT editable by analysts. "
                    "Value is either auto-derived (calculated fact) or comes directly from vendor verification. "
                    "Source: editable=false in fieldConfigs.tsx."
                )
                card(editable_explain)

            if f.get("storage"):
                st.markdown("**🗄️ Where this value is stored:**")
                for s in f["storage"]:
                    st.markdown(f"- `{s}`")
                st.markdown(
                    '<div style="font-size:.78rem;color:#475569;margin-top:4px;">'
                    'The facts table uses JSONB: <code>{"value":..., "source":{"platformId":N,"confidence":X}, "alternatives":[...]}</code><br>'
                    'alternatives[] contains ALL vendor responses before the winner was selected — full audit trail.'
                    '</div>', unsafe_allow_html=True)

            if f.get("null_cause"):
                st.markdown("**⚠️ When blank/N/A — all scenarios:**")
                card(f"{f['null_cause']}", "card-amber")

            if f.get("badges"):
                st.markdown("**🏷️ Possible badges/states shown in admin portal:**")
                for badge_entry in enrich_badges(f["badges"]):
                    # Support both old 3-tuple and new 5-tuple with extra details
                    if len(badge_entry) == 3:
                        btext, bclass, bexplain = badge_entry
                        bcondition = bdependency = baction = None
                    else:
                        btext, bclass, bexplain, bcondition, bdependency, baction = badge_entry

                    st.markdown(
                        f'<div style="background:#0A1628;border:1px solid #1E3A5F;border-radius:8px;'
                        f'padding:10px 14px;margin-bottom:8px;">'
                        f'<div style="display:flex;align-items:center;gap:10px;margin-bottom:6px;">'
                        f'{portal_badge(btext, bclass)}'
                        f'<span style="color:#E2E8F0;font-size:.88rem;font-weight:600;">{bexplain}</span>'
                        f'</div>'
                        + (f'<div style="color:#94A3B8;font-size:.8rem;margin-bottom:3px;">'
                           f'📐 <b style="color:#CBD5E1;">Condition:</b> <code>{bcondition}</code></div>'
                           if bcondition else '')
                        + (f'<div style="color:#94A3B8;font-size:.8rem;margin-bottom:3px;">'
                           f'🔗 <b style="color:#CBD5E1;">Depends on:</b> {bdependency}</div>'
                           if bdependency else '')
                        + (f'<div style="color:#FCD34D;font-size:.8rem;">'
                           f'💡 <b>Analyst action:</b> {baction}</div>'
                           if baction else '')
                        + '</div>',
                        unsafe_allow_html=True
                    )

            if f.get("react_src"):
                st.markdown("**📍 Source code that renders this field:**")
                st.markdown(src(*f["react_src"]), unsafe_allow_html=True)

        with c2:
            # ── DB connection note ──────────────────────────────────────
            st.markdown(
                '<div class="card card-purple" style="font-size:.78rem;padding:8px 12px;margin-bottom:8px;">'
                '<b>📌 Before running these queries, substitute:</b><br>'
                "<code>'abc123-de45-...'</code> → your actual business UUID "
                "(find it in the admin portal URL: <code>admin.joinworth.com/cases/&lt;caseId&gt;</code> "
                "→ case detail → business.id field via GET /cases/:id)<br>"
                '<b>PostgreSQL tables</b> (rds_warehouse_public, rds_cases_public): '
                'use psycopg2 with your RDS credentials.<br>'
                '<b>Redshift tables</b> (datascience.customer_files): '
                'use your Redshift endpoint with psycopg2 or AWS Data Wrangler.'
                '</div>', unsafe_allow_html=True)

            # ── SQL + Python — generated per field using confirmed Redshift pattern ──
            field_label = f["label"]

            st.markdown("**🔍 SQL — field-specific (Redshift):**")
            st.code(generate_sql(f), language="sql")

            st.markdown(f"**🐍 Python — {field_label} (field-specific):**")
            st.code(generate_python(f), language="python")



            if f.get("gpn_q"):
                st.markdown("**❓ UCM Working Session Question:**")
                card(f"{f['gpn_q']}", "card-amber")
            if f.get("gpn_a"):
                st.markdown("**✅ Confirmed Decision/Answer:**")
                card(f"{f['gpn_a']}", "card-green")


# ════════════════════════════════════════════════════════════════════════════
# TAB 1 — BACKGROUND
# ════════════════════════════════════════════════════════════════════════════
if tab == "📋 Background":
    sh("📋 Background Sub-Tab — Every Card, Every Field, Full Lineage")

    card("""<b>Cards in the Background sub-tab:</b><br>
    1. <b>Business Details card</b> — company info, address, financials, contact, ownership flags<br>
    2. <b>Industry card</b> — NAICS, MCC, industry name<br>
    3. <b>Google Maps pin</b> — rendered from primary_address fact<br><br>
    <b>API calls:</b> <code>GET /facts/business/:id/details</code> + <code>GET /facts/business/:id/kyb</code><br>
    <b>Source:</b> <code>fieldConfigs.tsx</code> lines 99–469 — definitive field config
    """)
    st.markdown(src("ADMIN_PORTAL",
        "microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/config/BackgroundTab/fieldConfigs.tsx",
        99, 469, "BACKGROUND_TAB_FIELD_CONFIGS"), unsafe_allow_html=True)

    bg_section = st.radio("Card", ["🏢 Business Details Card", "🏭 Industry Card"], horizontal=True)

    if bg_section == "🏢 Business Details Card":
        st.markdown("### 🏢 Business Details Card")
        card("""<b>Workflow: Submitted → Vendors → Fact Engine → Facts Table → API → React → Card</b><br>
        All fields in this card come from <code>factsBusinessDetails</code> (GET /facts/business/:id/details)
        except: <code>legal_name</code>, <code>formation_date</code>, <code>email</code>, <code>corporation</code>,
        <code>minority_owned</code>, <code>woman_owned</code>, <code>veteran_owned</code>
        which come from <code>getFactsKybData</code> (GET /facts/business/:id/kyb).
        """)

        BDET = [
            {"label":"Provided Business Name","fact":"business_name","api":"/facts/business/:id/details","editable":True,
             "tags":["✏️ Editable","📝 Submitted"],
             "sources":[("Applicant","pid=0","w=1.0","Submitted at onboarding — factKey: 'businessDetails'"),
                        ("ZoomInfo","pid=24","w=0.8","zi_c_name from zoominfo.comp_standard_global"),
                        ("Equifax","pid=17","w=0.7","efx_name from warehouse.equifax_us_latest")],
             "rule":"factWithHighestConfidence(). Applicant submission confidence=1.0 usually wins unless overridden.",
             "storage":["rds_warehouse_public.facts name='business_name'","rds_cases_public.data_businesses.name"],
             "null_cause":"Empty only if applicant skipped — required field.",
             "badges":[],
             "react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/config/BackgroundTab/fieldConfigs.tsx",101,115,"fieldKey: business_name"),
             "sql":"SELECT JSON_EXTRACT_PATH_TEXT(value, 'value') AS name, JSON_EXTRACT_PATH_TEXT(value, 'source', 'platformId') AS pid,\n       JSON_EXTRACT_PATH_TEXT(value, 'override', 'value') AS override\nFROM rds_warehouse_public.facts\nWHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name='business_name';",
             "python":"import psycopg2, pandas as pd\nconn = psycopg2.connect(host='<host>',dbname='<db>',user='<u>',password='<pw>')\ndf = pd.read_sql(\"SELECT value FROM rds_warehouse_public.facts WHERE business_id=%s AND name='business_name'\",conn,params=['YOUR-BUSINESS-UUID-HERE'])\nprint(df['value'].iloc[0])"},
            {"label":"Legal Business Name","fact":"legal_name","api":"/facts/business/:id/kyb","editable":False,
             "tags":["🔒 Read-Only","🔍 Vendor-Discovered"],
             "sources":[("Middesk","pid=16","w=2.0","businessEntityVerification.name — name found in SoS via IRS match"),
                        ("OpenCorporates","pid=23","w=0.9","firmographic.name from OC global registry"),
                        ("Trulioo","pid=38","w=0.8","clientData.businessName (UK/Canada)")],
             "rule":"factWithHighestConfidence() — Middesk wins (weight=2.0). NOT editable — verified name from public records.",
             "storage":["rds_warehouse_public.facts name='legal_name'"],
             "null_cause":"Empty when Middesk still processing or could not confirm entity.",
             "badges":[],
             "react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/config/BackgroundTab/fieldConfigs.tsx",116,130,"fieldKey: legal_name"),
             "sql":"SELECT JSON_EXTRACT_PATH_TEXT(value, 'value') AS legal_name FROM rds_warehouse_public.facts\nWHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name='legal_name';"},
            {"label":"DBA (Doing Business As)","fact":"dba / dba_found","api":"/facts/business/:id/details","editable":True,
             "tags":["✏️ Editable","📝 Multi-Source Array"],
             "sources":[("Middesk","pid=16","w=2.0","names[] where submitted=false"),
                        ("ZoomInfo","pid=24","w=0.8","zi_c_tradename"),
                        ("OpenCorporates","pid=23","w=0.9","alternate_names"),
                        ("Applicant","pid=0","w=1.0","DBA submitted at onboarding")],
             "rule":"combineFacts() — merges and deduplicates DBA names from ALL sources.",
             "storage":["rds_warehouse_public.facts name='dba_found' (discovered)","rds_warehouse_public.facts name='dba' (submitted)"],
             "null_cause":"Empty = no DBA registered. Very common — NOT an error.",
             "badges":[],
             "react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/config/BackgroundTab/fieldConfigs.tsx",131,145,"fieldKey: dba"),
             "sql":"SELECT name, JSON_EXTRACT_PATH_TEXT(value, 'value') FROM rds_warehouse_public.facts\nWHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name IN ('dba','dba_found');"},
            {"label":"Business Address","fact":"primary_address","api":"/facts/business/:id/details","editable":True,
             "tags":["✏️ Editable","🗺️ Google Maps Pin","💡 Suggestions"],
             "sources":[("Middesk","pid=16","w=2.0","businessEntityVerification.addresses + addressSources"),
                        ("OpenCorporates","pid=23","w=0.9","firmographic.registered_address"),
                        ("ZoomInfo","pid=24","w=0.8","zi_c_street, zi_c_city, zi_c_state, zi_c_zip"),
                        ("Equifax","pid=17","w=0.7","efx_address1, efx_city, efx_state"),
                        ("Applicant","pid=0","w=1.0","address submitted at onboarding")],
             "rule":"factWithHighestConfidence() for primary_address. combineFacts for addresses[].",
             "storage":["rds_warehouse_public.facts name='primary_address'","rds_warehouse_public.facts name='primary_address_string' (formatted)"],
             "null_cause":"Empty if applicant skipped address or vendors returned nothing.",
             "badges":[],
             "react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/config/BackgroundTab/fieldConfigs.tsx",146,171,"fieldKey: primary_address"),
             "sql":"SELECT name, JSON_EXTRACT_PATH_TEXT(value, 'value') FROM rds_warehouse_public.facts\nWHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name IN ('primary_address','primary_address_string','addresses');"},
            {"label":"Mailing Address","fact":"mailing_address","api":"/facts/business/:id/details","editable":True,
             "tags":["✏️ Editable","📝 Applicant Submitted"],
             "sources":[("Applicant","pid=0","w=1.0","Mailing address submitted if different from primary")],
             "rule":"Direct read from applicant submission. No vendor competition.",
             "storage":["rds_warehouse_public.facts name='mailing_address'"],
             "null_cause":"N/A if same as primary or not provided.",
             "badges":[],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/config/BackgroundTab/fieldConfigs.tsx",172,197,"fieldKey: mailing_address"),
             "sql":"SELECT JSON_EXTRACT_PATH_TEXT(value, 'value') FROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name='mailing_address';"},
            {"label":"Business Age (Formation Date)","fact":"formation_date","api":"/facts/business/:id/kyb","editable":True,
             "tags":["✏️ Editable","📅 Date → Age Computed"],
             "sources":[("Middesk","pid=16","w=2.0","businessEntityVerification.formation_date from SoS"),
                        ("OpenCorporates","pid=23","w=0.9","incorporation_date"),
                        ("ZoomInfo","pid=24","w=0.8","zi_c_year_founded"),
                        ("Equifax","pid=17","w=0.7","efx_yrest (year established)")],
             "rule":"factWithHighestConfidence(). Display: formatFormationDateWithAge() → 'MM/DD/YYYY (N years)'.",
             "storage":["rds_warehouse_public.facts name='formation_date'","rds_warehouse_public.facts name='year_established'"],
             "null_cause":"N/A when no SoS filing found and no vendor returned formation date.",
             "badges":[],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/config/BackgroundTab/fieldConfigs.tsx",198,213,"fieldKey: formation_date"),
             "sql":"SELECT name, JSON_EXTRACT_PATH_TEXT(value, 'value') FROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name IN ('formation_date','year_established');"},
            {"label":"Annual Revenue","fact":"revenue","api":"/facts/business/:id/financials","editable":True,
             "tags":["✏️ Editable","💰 Currency","ℹ️ Tooltip: most reputable source"],
             "sources":[("ZoomInfo","pid=24","w=0.8","zi_c_revenue"),
                        ("Equifax","pid=17","w=0.7","efx_locamount or efx_corpamount×1000"),
                        ("Accounting/Rutter","—","—","Connected accounting integration")],
             "rule":"factWithHighestConfidence(). getCurrencyDisplayValue() formats display.",
             "storage":["rds_warehouse_public.facts name='revenue'"],
             "null_cause":"N/A when no vendor has revenue data.",
             "badges":[],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/config/BackgroundTab/fieldConfigs.tsx",214,240,"fieldKey: revenue"),
             "sql":"SELECT JSON_EXTRACT_PATH_TEXT(value, 'value') AS revenue, JSON_EXTRACT_PATH_TEXT(value, 'source', 'platformId') AS pid\nFROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name='revenue';"},
            {"label":"Avg. Annual Revenue","fact":"revenue_equally_weighted_average","api":"/facts/business/:id/financials","editable":False,
             "tags":["🔒 Read-Only","🔢 Calculated","ℹ️ Tooltip"],
             "sources":[("Calculated","—","—","Equal-weight average across all revenue sources")],
             "rule":"Calculated fact — average of all revenue values.",
             "storage":["rds_warehouse_public.facts name='revenue_equally_weighted_average'"],
             "null_cause":"N/A when no revenue sources available.",
             "badges":[],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/config/BackgroundTab/fieldConfigs.tsx",241,270,"fieldKey: revenue_equally_weighted_average"),
             "sql":"SELECT JSON_EXTRACT_PATH_TEXT(value, 'value') FROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name='revenue_equally_weighted_average';"},
            {"label":"Net Income","fact":"net_income","api":"/facts/business/:id/financials","editable":True,
             "tags":["✏️ Editable","💰 Currency","🔗 Accounting Only"],
             "sources":[("Accounting/Rutter","—","—","From connected accounting integration only")],
             "rule":"Primarily from accounting integration.",
             "storage":["rds_warehouse_public.facts name='net_income'"],
             "null_cause":"N/A for most businesses — only when accounting integration connected.",
             "badges":[],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/config/BackgroundTab/fieldConfigs.tsx",271,287,"fieldKey: net_income"),
             "sql":"SELECT JSON_EXTRACT_PATH_TEXT(value, 'value') FROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name IN ('net_income','is_net_income');"},
            {"label":"Corporation Type","fact":"corporation","api":"/facts/business/:id/kyb","editable":True,
             "tags":["✏️ Editable","📋 Dropdown"],
             "sources":[("Middesk","pid=16","w=2.0","registrations[0].entity_type normalised from SoS"),
                        ("OpenCorporates","pid=23","w=0.9","company_type → llc/corporation/llp/lp"),
                        ("Applicant","pid=0","w=1.0","corporation_type from onboarding form")],
             "rule":"factWithHighestConfidence(). Options: LLC, Corporation, Partnership, Sole Proprietorship, etc.",
             "storage":["rds_warehouse_public.facts name='corporation'"],
             "null_cause":"N/A when no SoS filing found and applicant did not submit.",
             "badges":[],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/config/BackgroundTab/fieldConfigs.tsx",288,302,"fieldKey: corporation"),
             "sql":"SELECT JSON_EXTRACT_PATH_TEXT(value, 'value') FROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name='corporation';"},
            {"label":"Number of Employees","fact":"num_employees","api":"/facts/business/:id/details","editable":True,
             "tags":["✏️ Editable","🔢 Numeric"],
             "sources":[("Equifax","pid=17","w=0.7","efx_corpempcnt"),
                        ("ZoomInfo","pid=24","w=0.8","zi_c_employees"),
                        ("OpenCorporates","pid=23","w=0.9","number_of_employees")],
             "rule":"factWithHighestConfidence(). Shows N/A in display mode if empty.",
             "storage":["rds_warehouse_public.facts name='num_employees'"],
             "null_cause":"N/A — common for small/new businesses not in vendor databases.",
             "badges":[],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/config/BackgroundTab/fieldConfigs.tsx",303,319,"fieldKey: num_employees"),
             "sql":"SELECT JSON_EXTRACT_PATH_TEXT(value, 'value') AS employees, JSON_EXTRACT_PATH_TEXT(value, 'source', 'platformId') AS pid\nFROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name='num_employees';"},
            {"label":"Business Phone Number","fact":"business_phone","api":"/facts/business/:id/details","editable":True,
             "tags":["✏️ Editable","📞 Phone Format"],
             "sources":[("ZoomInfo","pid=24","w=0.8","zi_c_phone"),
                        ("SERP","pid=22","w=0.5","phone from web scraping"),
                        ("Middesk","pid=16","w=2.0","phone_numbers[0]"),
                        ("Equifax","pid=17","w=0.7","contact phone")],
             "rule":"factWithHighestConfidence().",
             "storage":["rds_warehouse_public.facts name='business_phone'"],
             "null_cause":"N/A for new/small businesses not in vendor databases.",
             "badges":[],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/config/BackgroundTab/fieldConfigs.tsx",320,335,"fieldKey: business_phone"),
             "sql":"SELECT JSON_EXTRACT_PATH_TEXT(value, 'value') FROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name='business_phone';"},
            {"label":"Business Email","fact":"email","api":"/facts/business/:id/kyb","editable":True,
             "tags":["✏️ Editable","📧 Email"],
             "sources":[("Equifax","pid=17","w=0.7","efx_email — primary vendor for business email")],
             "rule":"factWithHighestConfidence(). Equifax is primary source.",
             "storage":["rds_warehouse_public.facts name='email'"],
             "null_cause":"N/A — business email rarely in commercial databases.",
             "badges":[],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/config/BackgroundTab/fieldConfigs.tsx",336,349,"fieldKey: email"),
             "sql":"SELECT JSON_EXTRACT_PATH_TEXT(value, 'value') FROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name='email';"},
            {"label":"Minority Business Enterprise","fact":"minority_owned","api":"/facts/business/:id/kyb","editable":True,
             "tags":["✏️ Editable","📋 Dropdown (Yes/No/N/A)","🏢 Equifax"],
             "sources":[("Equifax","pid=17","w=0.7","efx_minority_business_enterprise flag")],
             "rule":"factWithHighestConfidence(). Dropdown: Yes/No/N/A.",
             "storage":["rds_warehouse_public.facts name='minority_owned'"],
             "null_cause":"N/A when Equifax does not have this flag for the business.",
             "badges":[],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/config/BackgroundTab/fieldConfigs.tsx",350,363,"fieldKey: minority_owned"),
             "sql":"SELECT name, JSON_EXTRACT_PATH_TEXT(value, 'value') FROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name IN ('minority_owned','woman_owned','veteran_owned');"},
            {"label":"Woman-Owned Business","fact":"woman_owned","api":"/facts/business/:id/kyb","editable":True,
             "tags":["✏️ Editable","📋 Dropdown (Yes/No/N/A)","🏢 Equifax"],
             "sources":[("Equifax","pid=17","w=0.7","efx_woman_owned_business flag")],
             "rule":"Same as minority_owned.","storage":["rds_warehouse_public.facts name='woman_owned'"],"null_cause":"N/A when Equifax does not have flag.","badges":[],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/config/BackgroundTab/fieldConfigs.tsx",364,377,"fieldKey: woman_owned"),
             "sql":"SELECT JSON_EXTRACT_PATH_TEXT(value, 'value') FROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name='woman_owned';"},
            {"label":"Veteran-Owned Business","fact":"veteran_owned","api":"/facts/business/:id/kyb","editable":True,
             "tags":["✏️ Editable","📋 Dropdown (Yes/No/N/A)","🏢 Equifax"],
             "sources":[("Equifax","pid=17","w=0.7","efx_veteran_owned_business flag")],
             "rule":"Same as minority_owned.","storage":["rds_warehouse_public.facts name='veteran_owned'"],"null_cause":"N/A when Equifax does not have flag.","badges":[],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/config/BackgroundTab/fieldConfigs.tsx",378,391,"fieldKey: veteran_owned"),
             "sql":"SELECT JSON_EXTRACT_PATH_TEXT(value, 'value') FROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name='veteran_owned';"},
        ]
        for f in BDET:
            render_field(f)

    elif bg_section == "🏭 Industry Card":
        st.markdown("### 🏭 Industry Card")
        card("""API: <code>GET /facts/business/:id/details</code> → <code>factsBusinessDetails</code><br>
        Source: <code>fieldConfigs.tsx L392-L469</code> — all 5 Industry field definitions.<br>
        Industry Name + NAICS Code are editable. Description fields are read-only (derived).""")

        # ── Deep explanation panel shown above INDUSTRY fields ──────────
        with st.expander("📚 Deep Dive: How the Fact Engine Selects NAICS Winners — Full Workflow", expanded=False):
            st.markdown("### Complete NAICS Winner Selection — Step-by-Step (from source code)")
            steps = [
                ("Step 1", "isValidFactValue() Filter — runs BEFORE any rule",
                 "rules.ts + factEngine.ts L162-165",
                 "Filters out vendors that returned null/empty NAICS. Only non-empty NAICS codes compete.\n"
                 "isValidFactValue() → FALSE (excluded) for: undefined, '', [], {}.\n"
                 "isValidFactValue() → TRUE (valid candidate) for: any non-empty string including '561499', null, 0, false.\n"
                 "If 0 vendors pass this filter → validOptions=[] → AI enrichment fires.", GREEN),
                ("Step 2", "manualOverride() — ALWAYS prepended first",
                 "rules.ts L109-123",
                 "factEngine.ts L178: rules.unshift(manualOverride) — analyst override beats everything.\n"
                 "If analyst set NAICS via PATCH /facts/override → wins unconditionally, no model needed.\n"
                 "Stored in fact JSONB: override: { value, userId, timestamp }", AMBER),
                ("Step 3", "factWithHighestConfidence() — main selection",
                 "rules.ts L35-57",
                 "Iterates all validOptions, picks highest confidence.\n"
                 "EACH VENDOR COMPUTES ITS OWN CONFIDENCE (not one shared model):\n"
                 "  • Middesk: 0.15 base + 0.20×(name task) + 0.20×(TIN task) + 0.20×(address task) + 0.20×(SoS task) = max 0.95\n"
                 "  • ZI/OC/EFX: match.index / MAX_CONFIDENCE_INDEX(55) = 0–1.0 (XGBoost entity-match)\n"
                 "  • AI: self-reported HIGH/MED/LOW from GPT-5-mini\n"
                 "IF |conf_A - conf_B| > 0.05 → higher confidence wins outright\n"
                 "IF |conf_A - conf_B| ≤ 0.05 → TIE → go to Step 4", BLUE),
                ("Step 4", "weightedFactSelector() — Tie-Break (only when confidences within 5%)",
                 "rules.ts L61-74",
                 "TIE-BREAK uses source WEIGHT (static config in sources.ts), NOT confidence:\n"
                 "  Middesk=2.0 > OC=0.9 > ZI=0.8 = Trulioo=0.8 > EFX=0.7 > SERP=0.3 > Applicant=0.2 > AI=0.1\n"
                 "WHY these weights? Middesk highest = direct SoS+IRS data. AI lowest = last resort.\n"
                 "EFX comment in sources.ts: 'relies upon manual files being ingested at some unknown cadence'.\n"
                 "On exact tie (same weight): left/primary fact wins (primaryFactWeight >= otherFactWeight)", PURPLE),
                ("Step 5", "Rule 4 — NO Minimum Confidence Cutoff",
                 "rules.ts (confirmed by absence)",
                 "NO code says 'if confidence < X, reject this vendor.'\n"
                 "Even confidence=0.05 is valid if it is the ONLY source that returned a NAICS code.\n"
                 "Confidence determines WHO WINS, never WHETHER a vendor is eligible.\n"
                 "Low-confidence NAICS is stored and shown — confidence visible in fact's source.confidence field.", RED),
                ("Step 6", "AI Enrichment Trigger",
                 "aiNaicsEnrichment.ts L61-64",
                 "FIRES WHEN: n_non_AI_sources_with_NAICS < minimumSources(1) AND total_sources < maximumSources(3)\n"
                 "= ALL of OC, ZI, EFX, Middesk, Trulioo, SERP returned null/empty naics_code.\n"
                 "AI receives: business_name + primary_address + (website if available) — NO vendor NAICS in prompt.\n"
                 "System prompt: 'If no evidence → return naics_code 561499 as last resort.'\n"
                 "AI weight=0.1 → AI only wins if ALL other sources have empty NAICS (which is why it fired).", TEAL),
                ("Step 7", "removeNaicsCode() — Final Safety Net (Rule 6)",
                 "aiNaicsEnrichment.ts L215-241",
                 "After AI returns a code: internalGetNaicsCode(code) checks if it exists in core_naics_code table.\n"
                 "IF valid (has naics_label) → code stored as-is.\n"
                 "IF invalid (hallucinated/wrong format) → removeNaicsCode() → code = NAICS_OF_LAST_RESORT = '561499'.\n"
                 "'561499' IS valid in core_naics_code → passes validation → stored and shown to user.\n"
                 "Original invalid code saved in response.naics_removed for audit.", GREY),
            ]
            for step_num, step_name, step_src, step_detail, step_col in steps:
                st.markdown(
                    f'<div class="card" style="border-left-color:{step_col};background:#0A1628;margin-bottom:8px;">'
                    f'<div style="display:flex;align-items:center;gap:10px;margin-bottom:4px;">'
                    f'<span style="color:{step_col};font-weight:900;font-size:1.1rem;">{step_num}</span>'
                    f'<span style="color:#E2E8F0;font-weight:700;">{step_name}</span>'
                    f'<span style="color:#475569;font-size:.75rem;font-family:Courier New;">{step_src}</span>'
                    f'</div>'
                    f'<pre style="color:{step_col};font-size:.80rem;white-space:pre-wrap;background:transparent;border:none;padding:0;margin:0;">{step_detail}</pre>'
                    f'</div>', unsafe_allow_html=True)

            st.markdown("### Complete Scenario Table — What the user ALWAYS sees")
            st.code("""
Scenario                                   | What Fact Engine does              | User sees
-------------------------------------------|-----------------------------------|------------------
All 6 vendors return valid NAICS codes     | factWithHighestConfidence() picks  | Real NAICS code
                                           | highest confidence winner          | (e.g. 722511)

3 vendors return NAICS, 3 return null      | Same rule, 3 candidates           | Real NAICS code
                                           | AI does NOT fire (min=1 met)       | from best of 3

1 vendor returns NAICS, 5 return null      | That vendor wins (Rule 4: no min  | Real NAICS code
                                           | cutoff, even if confidence=0.05)   | even if low conf

0 vendors return NAICS + AI has evidence   | AI fires → finds real NAICS        | Real NAICS code
(website, name keywords)                   | via website/web search             | from AI

0 vendors return NAICS + AI has no evid.  | AI fires → NAICS_OF_LAST_RESORT   | 561499 (always,
                                           | system prompt returns 561499       | never blank)

0 vendors + AI returns invalid/hallucin.  | removeNaicsCode() → 561499        | 561499

Analyst manually set NAICS                 | manualOverride() first → wins      | Analyst's code

CONCLUSION: NAICS is NEVER blank in practice. AI always returns 561499 when it has
no evidence. '-' should not appear unless an integration task failed or is still running.
            """, language=None)

        INDUSTRY = [
            {"label":"Industry Name","fact":"industry","api":"/facts/business/:id/details","editable":True,
             "tags":["✏️ Editable","🔢 NAICS-Derived (calculated)","❌ Not XGBoost","🔗 PostgreSQL Lookup"],
             "sources":[("Derived from NAICS (calculated)","No PID — source:null","confidence=0.9 (hardcoded)","Takes first 2 digits of winning naics_code → internalGetIndustries(sectorCode) → core_business_industries lookup. NOT XGBoost. NOT ML. Pure DB lookup."),
                        ("Trulioo","pid=38","w=0.7","standardizedIndustries[n].industryName — from Trulioo KYB response (secondary)"),
                        ("Applicant","pid=0","w=0.2","industry name submitted at onboarding (fallback only)")],
             "rule":"DEPENDENT FACT (source:null, businessDetails/index.ts L287). Reads resolved naics_code → sectorCode = naics_code.substring(0,2) → internalGetIndustries(sectorCode) → HTTP call to internal API → core_business_industries table lookup → returns industries[0].name. If naics_code=null → fn() returns undefined → Industry Name='-'. The confidence=0.9 is HARDCODED (not XGBoost) — means '90% confident the lookup is correct once we have a NAICS code.'",
             "storage":["rds_warehouse_public.facts name='industry' {name, id}","rds_cases_public.data_businesses.industry → core_business_industries.id"],
             "null_cause":"ONLY '-' when naics_code itself is null/undefined AND internalGetIndustries returned empty. In practice this should not occur because AI returns 561499 (not null). NAICS 561499 sectorCode='56' → 'Administrative and Support and Waste Management and Remediation Services' (always shown, never blank).",
             "badges":[],
             "react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/config/BackgroundTab/fieldConfigs.tsx",393,407,"fieldKey: industry"),
             "sql":"-- How industry is calculated:\nSELECT SUBSTRING(cnc.code, 1, 2) AS sector_code, cbi.name AS industry_name\nFROM core_naics_code cnc\nJOIN core_business_industries cbi ON cbi.sector_code = SUBSTRING(cnc.code, 1, 2)::integer\nWHERE cnc.code = '722511';  -- example\n\n-- Check a business's current industry:\nSELECT f.JSON_EXTRACT_PATH_TEXT(value, 'value') AS industry_fact, cbi.name AS db_industry,\n       (SELECT JSON_EXTRACT_PATH_TEXT(value, 'value') FROM rds_warehouse_public.facts WHERE business_id=f.business_id AND name='naics_code') AS naics\nFROM rds_warehouse_public.facts f\nJOIN rds_cases_public.data_businesses db ON db.id=f.business_id::uuid\nJOIN core_business_industries cbi ON cbi.id=db.industry\nWHERE f.business_id='YOUR-BUSINESS-UUID-HERE' AND f.name='industry';"},
            {"label":"NAICS Code","fact":"naics_code","api":"/facts/business/:id/details","editable":True,
             "tags":["✏️ Editable","7 Sources","🤖 AI Last Resort = 561499","⚙️ 6-Rule Winner Selection"],
             "sources":[
                 ("OpenCorporates","pid=23","w=0.9",
                  "Lineage: warehouse.oc_companies_latest → industry_code_uids field (pipe-delimited: 'us_naics-722511|uk_sic-56101'). Code parses for 'us_naics-' prefix. CONFIDENCE: match.index/55 from XGBoost entity-matching model (entity_matching_20250127 v1). Weight=0.9 because OC has global coverage + authoritative registry data."),
                 ("ZoomInfo","pid=24","w=0.8",
                  "Lineage: zoominfo.comp_standard_global → zi_c_naics6 field (6-digit NAICS). CONFIDENCE: match.index/55 from XGBoost entity-matching. Weight=0.8 — strong US B2B firmographic coverage."),
                 ("Trulioo","pid=38","w=0.7",
                  "Lineage: Trulioo live KYB API → clientData.standardizedIndustries[n].naicsCode. CONFIDENCE: status-based (completed→high, failed→low). Weight=0.7. Preferred for UK/Canada via truliooPreferredRule."),
                 ("Equifax","pid=17","w=0.7",
                  "Lineage: warehouse.equifax_us_latest → efx.primnaicscode field. CONFIDENCE: XGBoost entity-match score. Weight=0.7 — sources.ts comment: 'relies upon manual files being ingested at some unknown cadence' — hence lower weight than OC/ZI."),
                 ("SERP","pid=22","w=0.3",
                  "Lineage: SERP web scraping → businessLegitimacyClassification.naics_code. CONFIDENCE: heuristic/presence-based. Weight=0.3 — web scraping less reliable than structured vendor data."),
                 ("Applicant","pid=0","w=0.2",
                  "Lineage: onboarding form → data_businesses → businessDetails source. Schema enforced: /^\\d{6}$/. CONFIDENCE: always 1.0 for submitted values. Weight=0.2 — merchant may self-report incorrectly."),
                 ("AI Enrichment","pid=31","w=0.1",
                  "FIRES ONLY when all 6 above return null NAICS (minimumSources=1 not met). Receives: business_name + primary_address + website (if available). CONFIDENCE: GPT self-reported HIGH/MED/LOW. Weight=0.1 — last resort. Returns NAICS_OF_LAST_RESORT='561499' when no evidence. removeNaicsCode() replaces hallucinated codes with 561499.")],
             "rule": (
                 "1. manualOverride() — ALWAYS FIRST (analyst override via PATCH /facts/override wins unconditionally)\n"
                 "2. isValidFactValue() — filters out vendors that returned null/empty NAICS BEFORE any rule\n"
                 "3. factWithHighestConfidence() — each vendor's confidence computed by its OWN model:\n"
                 "   • Middesk: 0.15 base + 0.20×(name/TIN/address/SoS tasks) = max 0.95\n"
                 "   • ZI/OC/EFX: match.index / 55 (XGBoost entity-match score)\n"
                 "   • AI: GPT self-reported confidence mapped to numeric\n"
                 "4. weightedFactSelector() TIE-BREAK — fires when |conf_A - conf_B| ≤ 0.05:\n"
                 "   Uses static weight: Middesk(2.0) > OC(0.9) > ZI(0.8)=Trulioo(0.8) > EFX(0.7) > SERP(0.3) > Applicant(0.2) > AI(0.1)\n"
                 "5. Rule 4: NO minimum confidence cutoff — low confidence still wins if it is the only valid candidate\n"
                 "6. removeNaicsCode() — after AI: validates code against core_naics_code table. Invalid → '561499'."
             ),
             "storage":["rds_warehouse_public.facts name='naics_code' {value, source:{platformId,confidence,updatedAt}, alternatives:[]}","rds_cases_public.data_businesses.naics_id → core_naics_code.id",
                        "Pipeline B: datascience.customer_files.primary_naics_code (ZI vs EFX winner-takes-all, separate pipeline)"],
             "null_cause": (
                 "NAICS is NEVER blank in practice:\n"
                 "• If ANY vendor returns a NAICS code → it is shown (even at confidence=0.01)\n"
                 "• If ALL vendors return null → AI fires → returns '561499' (never null)\n"
                 "• If AI hallucinated invalid code → removeNaicsCode() → '561499' (never blank)\n"
                 "• '-' only appears if: (a) AI enrichment task itself crashed, OR (b) integrations not yet complete\n"
                 "• Low confidence does NOT cause '-' — Rule 4 prohibits minimum confidence cutoff\n"
                 "• If winning source has a match but its NAICS field is empty → that source excluded from competition → next source wins → cascade to AI if needed"
             ),
             "badges":[],
             "react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/config/BackgroundTab/fieldConfigs.tsx",408,422,"fieldKey: naics_code"),
             "sql":"-- Fetch raw value text and parse in Python (see Python snippet)\nSELECT name, value, received_at\nFROM rds_warehouse_public.facts\nWHERE business_id = 'YOUR-BUSINESS-UUID-HERE'\n  AND name = 'fact_name';"},
            {"label":"NAICS Description","fact":"naics_description","api":"/facts/business/:id/details","editable":False,
             "tags":["🔒 Read-Only","🔢 Auto-Derived"],
             "sources":[("core_naics_code lookup","calculated","—","label WHERE code = naics_code.value")],
             "rule":"Read-only. Computed from winning naics_code → core_naics_code.label.",
             "storage":["rds_warehouse_public.facts name='naics_description'","core_naics_code.label"],
             "null_cause":"N/A when naics_code is null or not in core_naics_code table.",
             "badges":[],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/config/BackgroundTab/fieldConfigs.tsx",423,438,"fieldKey: naics_description"),
             "sql":"SELECT JSON_EXTRACT_PATH_TEXT(value, 'value') FROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name='naics_description';"},
            {"label":"MCC Code","fact":"mcc_code / mcc_code_found / mcc_code_from_naics","api":"/facts/business/:id/details","editable":True,
             "tags":["✏️ Editable","🤖 AI + Calculated"],
             "sources":[("AI Enrichment (mcc_code_found)","pid=31","—","response.mcc_code directly from GPT (preferred)"),
                        ("Calculated from NAICS","calculated","—","rel_naics_mcc: naics_id → mcc_id")],
             "rule":"mcc_code = mcc_code_found?.value ?? mcc_code_from_naics?.value. AI-provided preferred.",
             "storage":["rds_warehouse_public.facts name='mcc_code'","rds_warehouse_public.facts name='mcc_code_found'","rds_warehouse_public.facts name='mcc_code_from_naics'","rds_cases_public.data_businesses.mcc_id → core_mcc_code"],
             "null_cause":"Almost never null — rel_naics_mcc calculates if NAICS exists. 5614 with 'Fallback MCC...' = AI had no evidence (Gap G5).",
             "badges":[],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/config/BackgroundTab/fieldConfigs.tsx",440,454,"fieldKey: mcc_code"),
             "sql":"SELECT name, JSON_EXTRACT_PATH_TEXT(value, 'value') FROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE'\nAND name IN ('mcc_code','mcc_code_found','mcc_code_from_naics','mcc_description');"},
            {"label":"MCC Description","fact":"mcc_description","api":"/facts/business/:id/details","editable":False,
             "tags":["🔒 Read-Only","⚠️ Known Bug: Fallback Text"],
             "sources":[("AI Enrichment","pid=31","—","AI prompt-generated text OR core_mcc_code.label lookup")],
             "rule":"Read-only (editable=false). mcc_description = AI text OR core_mcc_code.label. KNOWN BUG (Gap G5): shows 'Fallback MCC per instructions...' to customers.",
             "storage":["rds_warehouse_public.facts name='mcc_description'"],
             "null_cause":"Shows 'Fallback MCC per instructions (no industry evidence...)' when AI had no evidence. Should show 'Classification pending...' — fix pending.",
             "badges":[],
             "react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/config/BackgroundTab/fieldConfigs.tsx",455,469,"fieldKey: mcc_description"),
             "sql":"SELECT JSON_EXTRACT_PATH_TEXT(value, 'value') AS mcc_desc FROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name='mcc_description';\n-- If shows 'Fallback MCC per instructions...' → AI enrichment had no vendor evidence"},
        ]
        for f in INDUSTRY:
            render_field(f)


# ════════════════════════════════════════════════════════════════════════════
# TAB 2 — BUSINESS REGISTRATION
# ════════════════════════════════════════════════════════════════════════════
elif tab == "🏛️ Business Registration":
    sh("🏛️ Business Registration Sub-Tab — Every Card, Every Field, Full Lineage")

    card("""<b>Cards in Business Registration:</b><br>
    1. <b>Business Registration card</b> — Business Name + Tax ID (EIN) with ✅/⚠️ Verified badge<br>
    2. <b>Secretary of State Filings card(s)</b> — one card per SoS filing found (OR 'No Registry Data to Display')<br>
    3. <b>Shareholder Document card</b> — document uploaded by applicant for manual TIN verification<br><br>
    <b>API:</b> <code>GET /facts/business/:id/kyb</code><br>
    <b>Sources:</b> <code>BusinessRegistrationTab.tsx</code> + <code>useBusinessRegistrationTabDetails.tsx</code> + <code>SOSBadge.tsx</code> + <code>TinBadge.tsx</code> + <code>EntityJurisdictionCell.tsx</code>
    """)
    st.markdown(src("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/hooks/BusinessRegistrationTab/useBusinessRegistrationTabDetails.tsx",59,94,"taxDetails builder"), unsafe_allow_html=True)

    biz_card = st.radio("Card", ["🏷️ Business Registration Card","📄 SoS Filings Card(s)","📁 Shareholder Document"], horizontal=True)

    if biz_card == "🏷️ Business Registration Card":
        st.markdown("### 🏷️ Business Registration Card — Business Name + Tax ID")
        BIZ_REG = [
            {"label":"Business Name (in Business Registration card)","fact":"legal_name","api":"/facts/business/:id/kyb","editable":True,
             "tags":["✏️ Editable","🔍 Vendor-Discovered"],
             "sources":[("Middesk","pid=16","w=2.0","businessEntityVerification.name — name in SoS records"),
                        ("OpenCorporates","pid=23","w=0.9","firmographic.name"),
                        ("Trulioo","pid=38","w=0.8","clientData.businessName (UK/Canada)")],
             "rule":"factWithHighestConfidence(). Note: this field shows getFactsKybData.data.legal_name.value — the VERIFIED name from public records, not the submitted name.",
             "storage":["rds_warehouse_public.facts name='legal_name'"],
             "null_cause":"Empty when Middesk still processing or could not confirm entity.",
             "badges":[],
             "react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/hooks/BusinessRegistrationTab/useBusinessRegistrationTabDetails.tsx",72,82,"taxDetails Business Name"),
             "sql":"SELECT JSON_EXTRACT_PATH_TEXT(value, 'value') AS legal_name FROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name='legal_name';"},
            {"label":"Tax ID Number (EIN) + ✅/⚠️ Verified badge","fact":"tin + tin_match_boolean","api":"/facts/business/:id/kyb","editable":True,
             "tags":["✏️ Editable","🔐 IRS TIN Match"],
             "sources":[("Applicant","pid=0","w=1.0","EIN submitted at onboarding — always masked after submission"),
                        ("Middesk","pid=16","w=2.0","IRS TIN Match: reviewTasks[key='tin'].status = success/failure"),
                        ("Trulioo","pid=38","w=0.8","TIN-to-name match for UK/Canada")],
             "rule":"tin.value = raw masked EIN. TinBadge reads tin_match_boolean.value: true→Verified(info/blue), false→Unverified(warning). tin_match.value.status='failure'→'Unverified'.",
             "storage":["rds_warehouse_public.facts name='tin_submitted' (masked EIN)","rds_warehouse_public.facts name='tin_match' {status,message,sublabel}","rds_warehouse_public.facts name='tin_match_boolean' (boolean)"],
             "null_cause":"Badge shows Unverified when IRS did not confirm TIN+name match. Causes: wrong EIN, name mismatch, sole proprietor using personal SSN.",
             "badges":[
                 ("✅ Verified","badge-verified","tin_match_boolean.value=true → IRS confirmed EIN+legal name match via Middesk TIN Match service"),
                 ("⚠️ Unverified","badge-unverified","tin_match.value.status='failure' → IRS did not confirm match"),
                 ("⚠️ [status]","badge-unverified","capitalize(tin_match.value.status) when status is neither success nor failure"),
             ],
             "react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/components/TinBadge.tsx",1,48,"TinBadge component"),
             "sql":"SELECT name, JSON_EXTRACT_PATH_TEXT(value, 'value') AS val FROM rds_warehouse_public.facts\nWHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name IN ('tin','tin_submitted','tin_match','tin_match_boolean');",
             "gpn_q":"tin.value is an integer — how does it transform to Verified/Unverified?",
             "gpn_a":"UCM should use tin_match_boolean.value (not tin.value). TRUE=Verified, FALSE=Unverified. Unverified→Fail UCM rule."},
        ]
        for f in BIZ_REG: render_field(f)

    elif biz_card == "📄 SoS Filings Card(s)":
        st.markdown("### 📄 Secretary of State Filings Card(s)")
        card("""<b>Important:</b> This card renders ONCE PER SoS FILING found. If Middesk finds 3 state registrations → 3 cards.<br>
        If Middesk finds NOTHING → shows 'No Registry Data to Display' instead.<br>
        Non-US businesses: card title changes to 'Registration Filing' (SOSFilingCard.tsx).<br>
        <b>Source:</b> <code>useBusinessRegistrationTabDetails.tsx L107-L285</code> — enhancedSosFilingsDetails builder""")
        st.markdown(src("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/components/BusinessRegistrationTab/SOSFilingCard.tsx",1,60,"SOSFilingCard — card per filing"), unsafe_allow_html=True)
        st.markdown(src("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/components/SOSBadge.tsx",1,60,"SOSBadge — badge per filing"), unsafe_allow_html=True)

        SOS_FIELDS = [
            {"label":"SoS Verified / Missing Active Filing badge","fact":"sos_filings[n].active","api":"/facts/business/:id/kyb","editable":False,
             "tags":["🔒 Auto-Badge","📍 Per Filing"],
             "sources":[("Middesk","pid=16","w=2.0","registrations[n].status === 'active' → active=true"),
                        ("OpenCorporates","pid=23","w=0.9","current_status='Active' → true")],
             "rule":"SOSBadge.tsx: isInvalidated→INVALIDATED. sosFiling.active=true→VERIFIED(info/blue,CheckBadgeIcon). sosFiling.active=false→MISSING ACTIVE FILING(error/red,XCircleIcon).",
             "storage":["rds_warehouse_public.facts name='sos_filings' (array, each with active: boolean)","rds_warehouse_public.facts name='sos_active' (boolean)"],
             "null_cause":"No badge (no SoS card) = sos_filings.value is empty → shows 'No Registry Data to Display'.",
             "badges":[
                 ("✅ Verified","badge-verified","sosFiling.active=true: active SoS filing found and in good standing"),
                 ("🔴 Missing Active Filing","badge-missing","sosFiling.active=false: filing found but inactive/dissolved"),
                 ("⚠️ Invalidated","badge-unverified","isInvalidated=true: filing marked invalid by Worth"),
             ],
             "react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/components/SOSBadge.tsx",12,60,"getSosBadgeConfig"),
             "sql":"-- Fetch raw value text and parse in Python (see Python snippet)\nSELECT name, value, received_at\nFROM rds_warehouse_public.facts\nWHERE business_id = 'YOUR-BUSINESS-UUID-HERE'\n  AND name = 'fact_name';"},
            {"label":"Filing Status","fact":"sos_filings[n].active","api":"/facts/business/:id/kyb","editable":False,
             "tags":["🔒 Read-Only"],
             "sources":[("Middesk","pid=16","w=2.0","registrations[n].status → active ? 'Active' : 'Inactive'")],
             "rule":"react: sos.active != null ? capitalize(sos.active ? 'Active' : 'Inactive') : 'N/A'",
             "storage":["Within sos_filings array: active field per filing"],
             "null_cause":"N/A when Middesk did not return a status for this specific filing.",
             "badges":[("Active","badge-verified","registrations[n].status='active'"),("Inactive","badge-missing","registrations[n].status≠'active'"),("N/A","badge-unknown","status field absent in SoS record")],
             "react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/hooks/BusinessRegistrationTab/useBusinessRegistrationTabDetails.tsx",120,130,"Filing Status row"),
             "sql":"-- Fetch raw value text and parse in Python (see Python snippet)\nSELECT name, value, received_at\nFROM rds_warehouse_public.facts\nWHERE business_id = 'YOUR-BUSINESS-UUID-HERE'\n  AND name = 'fact_name';"},
            {"label":"Entity Jurisdiction Type","fact":"sos_filings[n].foreign_domestic","api":"/facts/business/:id/kyb","editable":False,
             "tags":["🔒 Read-Only","🔴 IMPORTANT: 'Primary' is a Worth label"],
             "sources":[("Middesk","pid=16","w=2.0","registrations[n].jurisdiction field → 'domestic'/'foreign'"),
                        ("OpenCorporates","pid=23","w=0.9","home_jurisdiction_code vs jurisdiction_code comparison")],
             "rule":"EntityJurisdictionCell.tsx: JURISDICTION_CONFIG maps domestic→{displayLabel:'Domestic',badgeLabel:'Primary'} and foreign→{displayLabel:'Foreign',badgeLabel:'Secondary'}. THE 'PRIMARY' BADGE IS WORTH'S LABEL, NOT MIDDESK'S.",
             "storage":["Within sos_filings array: foreign_domestic field"],
             "null_cause":"N/A when jurisdiction field absent or could not be determined.",
             "badges":[("Domestic Primary","badge-verified","foreign_domestic='domestic' — business originally formed in this state"),("Foreign Secondary","badge-unverified","foreign_domestic='foreign' — formed elsewhere, registered to operate here")],
             "react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/components/BusinessRegistrationTab/EntityJurisdictionCell.tsx",1,35,"EntityJurisdictionCell"),
             "sql":"-- Fetch raw value text and parse in Python (see Python snippet)\nSELECT name, value, received_at\nFROM rds_warehouse_public.facts\nWHERE business_id = 'YOUR-BUSINESS-UUID-HERE'\n  AND name = 'fact_name';"},
            {"label":"State","fact":"sos_filings[n].state","api":"/facts/business/:id/kyb","editable":False,
             "tags":["🔒 Read-Only"],
             "sources":[("Middesk","pid=16","w=2.0","registrations[n].registration_state — 2-letter code"),
                        ("OpenCorporates","pid=23","w=0.9","jurisdiction_code split → 'us_fl' → 'FL'")],
             "rule":"Shown per filing. Format: 2-letter state code (e.g. 'FL', 'NY').",
             "storage":["Within sos_filings array: state field per filing"],
             "null_cause":"N/A if state absent from SoS record.","badges":[],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/hooks/BusinessRegistrationTab/useBusinessRegistrationTabDetails.tsx",140,160,"State row"),"sql":"-- Fetch raw value text and parse in Python (see Python snippet)\nSELECT name, value, received_at\nFROM rds_warehouse_public.facts\nWHERE business_id = 'YOUR-BUSINESS-UUID-HERE'\n  AND name = 'fact_name';"},
            {"label":"Registration Date","fact":"sos_filings[n].filing_date","api":"/facts/business/:id/kyb","editable":False,
             "tags":["🔒 Read-Only","📅 SoS Registration Date"],
             "sources":[("Middesk","pid=16","w=2.0","registrations[n].registration_date — ISO date from SoS"),
                        ("OpenCorporates","pid=23","w=0.9","incorporation_date")],
             "rule":"Shown per filing. Formatted as MM-DD-YYYY by convertToLocalDate().",
             "storage":["Within sos_filings array: filing_date field"],"null_cause":"N/A if date absent from SoS record.","badges":[],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/hooks/BusinessRegistrationTab/useBusinessRegistrationTabDetails.tsx",163,175,"Registration Date row"),"sql":"-- Fetch raw value text and parse in Python (see Python snippet)\nSELECT name, value, received_at\nFROM rds_warehouse_public.facts\nWHERE business_id = 'YOUR-BUSINESS-UUID-HERE'\n  AND name = 'fact_name';"},
            {"label":"Entity Type","fact":"sos_filings[n].entity_type","api":"/facts/business/:id/kyb","editable":False,
             "tags":["🔒 Read-Only","Normalised"],
             "sources":[("Middesk","pid=16","w=2.0","registrations[n].entity_type — raw SoS string (e.g. 'Llc')"),
                        ("OpenCorporates","pid=23","w=0.9","company_type normalised → llc/corporation/llp/lp/sole proprietorship")],
             "rule":"Shown per filing. Worth normalises: 'Limited Liability Company'→'llc', 'Incorporated'→'corporation'.",
             "storage":["Within sos_filings array: entity_type field"],"null_cause":"N/A if state does not disclose entity type.","badges":[],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/hooks/BusinessRegistrationTab/useBusinessRegistrationTabDetails.tsx",178,195,"Entity Type row"),"sql":"-- Fetch raw value text and parse in Python (see Python snippet)\nSELECT name, value, received_at\nFROM rds_warehouse_public.facts\nWHERE business_id = 'YOUR-BUSINESS-UUID-HERE'\n  AND name = 'fact_name';"},
            {"label":"Corporate Officers","fact":"people.value[] filtered by sos.id","api":"/facts/business/:id/kyb","editable":False,
             "tags":["🔒 Read-Only","Per Filing Filtered"],
             "sources":[("Middesk","pid=16","w=2.0","registrations[n].officers — officers per state filing"),
                        ("OpenCorporates","pid=23","w=0.9","officers array from OC"),
                        ("Trulioo","pid=38","w=0.8","principals/directors")],
             "rule":"people.value filtered by person.source.some(src => src.id === sos.id). Shows only officers for THIS specific filing.",
             "storage":["rds_warehouse_public.facts name='people' (array of {name, titles[], source[]})"],"null_cause":"N/A if state doesn't require officer disclosure.","badges":[],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/hooks/BusinessRegistrationTab/useBusinessRegistrationTabDetails.tsx",197,240,"Corporate Officers row"),"sql":"SELECT JSON_EXTRACT_PATH_TEXT(value, 'value') AS people FROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name='people';"},
            {"label":"Legal Entity Name (within each SoS card)","fact":"sos_filings[n].filing_name","api":"/facts/business/:id/kyb","editable":False,
             "tags":["🔒 Read-Only","Per Filing"],
             "sources":[("Middesk","pid=16","w=2.0","registrations[n].name — name of this specific SoS filing"),
                        ("OpenCorporates","pid=23","w=0.9","OC filing name")],
             "rule":"Shows the name on this specific SoS filing. May differ from legal_name across states.","storage":["Within sos_filings array: filing_name field"],"null_cause":"N/A if absent from SoS record.","badges":[],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/hooks/BusinessRegistrationTab/useBusinessRegistrationTabDetails.tsx",246,260,"Legal Entity Name row"),"sql":"-- Fetch raw value text and parse in Python (see Python snippet)\nSELECT name, value, received_at\nFROM rds_warehouse_public.facts\nWHERE business_id = 'YOUR-BUSINESS-UUID-HERE'\n  AND name = 'fact_name';"},
        ]
        for f in SOS_FIELDS: render_field(f)

        st.markdown("---")
        st.markdown("### ❓ Why 'No Registry Data to Display'?")
        card("""<b>Condition:</b> <code>sos_filings.value = []</code> (empty array)<br><br>
        This appears when Middesk searched all US SoS databases by TIN+name and found ZERO registrations.<br>
        It is NOT a confidence threshold cutoff — if Middesk finds nothing, nothing is shown.<br><br>
        <b>Common reasons:</b><br>
        • Sole proprietor — not required to file with SoS in most states<br>
        • Business registered under different TIN or name variant<br>
        • New business not yet in Middesk's SoS database<br>
        • State not covered by Middesk's SoS search scope<br><br>
        <b>Note:</b> A business can be 'Business Registration ✅ Verified' (TIN matched IRS) AND show 'No Registry Data' (no SoS filing found). These are INDEPENDENT checks.""", "card-amber")
        st.markdown(src("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/hooks/BusinessRegistrationTab/useBusinessRegistrationTabDetails.tsx",107,115,"shouldHideSosForIntegration + empty check"), unsafe_allow_html=True)

    elif biz_card == "📁 Shareholder Document":
        st.markdown("### 📁 Shareholder Document Card")
        card("""Shown when applicant uploaded a document for manual TIN verification.<br>
        Source: <code>getFactsKybData.data.shareholder_document.value</code> → <code>{id, url, file_name}</code><br>
        Used when Middesk TIN match fails — applicant uploads IRS EIN letter or similar document for manual review.""")
        st.code("SELECT JSON_EXTRACT_PATH_TEXT(value, 'value') FROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name='shareholder_document';", language="sql")


# ════════════════════════════════════════════════════════════════════════════
# TAB 3 — CONTACT INFORMATION
# ════════════════════════════════════════════════════════════════════════════
elif tab == "📬 Contact Information":
    sh("📬 Contact Information Sub-Tab — Every Card, Every Field, Full Lineage")

    card("""<b>Cards in Contact Information:</b><br>
    1. <b>Addresses card</b> — submitted + reported addresses with Verified/Unverified badges + Deliverable badge<br>
    2. <b>Business Names card</b> — submitted names + reported names with Verified badge<br>
    3. <b>MATCH card</b> — Mastercard MATCH (fraud database) — currently commented out / not shown<br><br>
    <b>API:</b> <code>GET /facts/business/:id/kyb</code> + <code>GET /facts/business/:id/details</code> + Google Profile API<br>
    <b>Sources:</b> <code>ContactInformationTab.tsx</code> → <code>AddressesCard.tsx</code> + <code>BusinessNamesCard.tsx</code>
    """)
    st.markdown(src("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/components/AddressesCard.tsx",1,60,"AddressesCard — addresses rendering"), unsafe_allow_html=True)

    ci_card = st.radio("Card", ["🗺️ Addresses Card","📛 Business Names Card","🔴 MATCH Card (not active)"], horizontal=True)

    if ci_card == "🗺️ Addresses Card":
        st.markdown("### 🗺️ Addresses Card")
        card("""<b>Important discovery from AddressesCard.tsx:</b><br>
        The card calls THREE APIs simultaneously:<br>
        • <code>useGetFactsBusinessDetails</code> → business_addresses_submitted_strings (submitted addresses)<br>
        • <code>useGetFactsKyb</code> → addresses (reported), addresses_deliverable, address_verification<br>
        • <code>useGetGoogleProfileByBusinessId</code> → Google Profile address_match + business_match<br><br>
        <b>Google Profile logic:</b> googleProfileMatch = (business_match='match found') AND (address_match='match')<br>
        <b>enrichAddressesWithStatusFor360ReportParity</b> enriches both submitted AND reported addresses with status.""")

        ADDR_FIELDS = [
            {"label":"Submitted Address + Business Registration / Google Profile badges","fact":"business_addresses_submitted_strings","api":"/facts/business/:id/details","editable":False,
             "tags":["🔒 Read-Only","📝 Applicant Submitted"],
             "sources":[("Applicant","pid=0","w=1.0","address submitted at onboarding — stored in business_addresses_submitted_strings fact")],
             "rule":"enrichAddressesWithStatusFor360ReportParity() compares submitted address against: kybDeliverableAddresses (USPS) + kybAddressVerification (Middesk) + googleProfileMatch.",
             "storage":["rds_warehouse_public.facts name='business_addresses_submitted_strings'","rds_warehouse_public.facts name='addresses' (reported)","rds_warehouse_public.facts name='addresses_deliverable'","rds_warehouse_public.facts name='address_verification' {status,sublabel,message}"],
             "null_cause":"Empty if applicant did not submit an address.",
             "badges":[
                 ("✅ Business Registration","badge-verified","address_verification.value.status='success' OR address_match_boolean=true — Middesk confirmed address matches public records"),
                 ("⚠️ Business Registration","badge-unverified","address_verification.value.sublabel shows reason — address not confirmed in SoS records"),
                 ("✅ Google Profile","badge-verified","googleProfileMatch=true — Google Business Profile confirms same address"),
                 ("⚠️ Google Profile","badge-unverified","googleProfileMatch=false — Google Business Profile shows different address or not found"),
                 ("✅ Deliverable","badge-deliverable","address in addresses_deliverable — USPS confirmed this is a real mailable address"),
             ],
             "react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/components/AddressesCard.tsx",200,252,"AddressesCard submitted + reported"),
             "sql":"SELECT name, JSON_EXTRACT_PATH_TEXT(value, 'value') FROM rds_warehouse_public.facts\nWHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name IN\n('business_addresses_submitted_strings','addresses','addresses_deliverable','address_verification','address_match_boolean');"},
            {"label":"Reported Addresses + Verified / Deliverable badges","fact":"addresses + addresses_deliverable + address_verification","api":"/facts/business/:id/kyb","editable":False,
             "tags":["🔒 Read-Only","🔍 Multi-Source"],
             "sources":[("Middesk","pid=16","w=2.0","addressSources from businessEntityVerification — SoS registered addresses"),
                        ("OpenCorporates","pid=23","w=0.9","firmographic.registered_address"),
                        ("ZoomInfo","pid=24","w=0.8","zi_c_street/city/state/zip"),
                        ("Equifax","pid=17","w=0.7","efx_address1/city/state"),
                        ("SERP","pid=22","w=0.5","address from web scraping")],
             "rule":"combineFacts() — all reported addresses merged. Deliverable status from USPS check (addresses_deliverable). Business Registration badge from address_verification fact (Middesk address task).",
             "storage":["rds_warehouse_public.facts name='addresses' (array of reported addresses)","rds_warehouse_public.facts name='addresses_deliverable' (array of USPS-verified addresses)"],
             "null_cause":"No Addresses section if addresses.value.length=0. Each address gets own Deliverable badge independently.",
             "badges":[("✅ Deliverable","badge-deliverable","Address in addresses_deliverable array — USPS confirmed mailable"),("✅ Business Registration","badge-verified","address_verification confirms this address"),("⚠️ Business Registration","badge-unverified","Middesk address task failed for this address")],
             "react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/components/AddressesCard.tsx",229,252,"enrichAddressesWithStatusFor360ReportParity"),
             "sql":"SELECT JSON_ARRAY_LENGTH(value, true)  -- or parse in Python AS count, JSON_EXTRACT_PATH_TEXT(value, 'value') AS addrs\nFROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name='addresses';\n\nSELECT JSON_EXTRACT_PATH_TEXT(value, 'value') AS deliverable_addrs FROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name='addresses_deliverable';"},
        ]
        for f in ADDR_FIELDS: render_field(f)

    elif ci_card == "📛 Business Names Card":
        st.markdown("### 📛 Business Names Card")
        BN_FIELDS = [
            {"label":"Submitted Names + Verified badge","fact":"names_submitted + name_match_boolean","api":"/facts/business/:id/kyb","editable":False,
             "tags":["🔒 Read-Only"],
             "sources":[("Applicant","pid=0","w=1.0","Legal name + DBA submitted at onboarding")],
             "rule":"name_match_boolean.value controls the ✅ Verified / Failure badge on the section header. names_submitted shows what the applicant submitted.",
             "storage":["rds_warehouse_public.facts name='names_submitted'","rds_warehouse_public.facts name='name_match_boolean' (boolean)"],
             "null_cause":"Section hidden if names_submitted.value.length = 0.",
             "badges":[("✅ Verified","badge-verified","name_match_boolean=true: submitted name found in SoS/IRS records (Middesk name task success)"),("❌ Failure","badge-missing","name_match_boolean=false: name not confirmed in public records"),("⚠️ [sublabel]","badge-unverified","name_match.value.status='warning': partial match — shows sublabel text")],
             "react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/components/BusinessNamesCard.tsx",1,30,"BusinessNamesCard"),
             "sql":"SELECT name, JSON_EXTRACT_PATH_TEXT(value, 'value') FROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE'\nAND name IN ('names_submitted','names_found','name_match_boolean','name_match');"},
            {"label":"Reported Names (from public records)","fact":"names_found","api":"/facts/business/:id/kyb","editable":False,
             "tags":["🔒 Read-Only","🔍 Multi-Source"],
             "sources":[("Middesk","pid=16","w=2.0","names[] — legal + trade names found in SoS"),
                        ("ZoomInfo","pid=24","w=0.8","zi_c_name, zi_c_tradename"),
                        ("OpenCorporates","pid=23","w=0.9","alternate_names"),
                        ("SERP","pid=22","w=0.5","business name from web scraping")],
             "rule":"combineFacts() — all reported names merged and deduplicated.",
             "storage":["rds_warehouse_public.facts name='names_found' (array)"],"null_cause":"Empty if no vendor found any names.","badges":[],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/components/BusinessNamesCard.tsx",1,30,"reported names section"),"sql":"SELECT JSON_EXTRACT_PATH_TEXT(value, 'value') AS names_found FROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name='names_found';"},
        ]
        for f in BN_FIELDS: render_field(f)

    elif ci_card == "🔴 MATCH Card (not active)":
        st.markdown("### 🔴 MATCH Card")
        card("""<b>Status: Commented out / not yet active</b><br><br>
        From <code>ContactInformationTab.tsx L181</code>:<br>
        <code>{'/* todo: add back once Mastercard Match is implemented */'}</code><br><br>
        The MATCH card was planned to show Mastercard MATCH (Merchant Alert To Control High-risk) results.<br>
        MATCH is a database of merchants terminated for fraud, excessive chargebacks, or data security issues.<br>
        When no MATCH records: "No MATCH records were found. This means the merchant has not been reported..."<br>
        When records found: Shows title + description per MATCH result with ⚠️ Results Found badge.<br><br>
        <b>This card is NOT currently shown to users.</b>""", "card-amber")
        st.markdown(src("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/ContactInformationTab.tsx",181,182,"MATCH card commented out"), unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════════════════
# TAB 4 — WEBSITE
# ════════════════════════════════════════════════════════════════════════════
elif tab == "🌐 Website":
    sh("🌐 Website Sub-Tab — Every Card, Every Field, Full Lineage")

    card("""<b>Cards in Website sub-tab:</b><br>
    1. <b>Website Details card</b> — URL (editable) + Creation Date, Expiration Date, Parked Domain, Status (read-only)<br>
    2. <b>Website Pages</b> — screenshot thumbnails per discovered page<br><br>
    <b>API:</b> <code>GET /verification/businesses/:id/website-data</code> (website details) + <code>GET /facts/business/:id/details</code> (website URL)<br>
    <b>Sources:</b> <code>WebsiteTab.tsx</code> + <code>WebsiteTab/fieldConfigs.tsx</code> + <code>useWebsiteNonEditableFields.tsx</code><br><br>
    <b>Key discovery:</b> When website URL is edited (isWebsiteDirty=true) → Creation Date, Expiration Date, Parked Domain, Status ALL show N/A until re-verified.""")

    WEB_FIELDS = [
        {"label":"Website URL","fact":"website","api":"/facts/business/:id/details","editable":True,
         "tags":["✏️ Editable","🔗 External Link Icon","💡 Suggestions"],
         "sources":[("ZoomInfo","pid=24","w=0.8","zi_c_url from firmographic data"),
                    ("businessDetails","pid=0","w=1.0","official_website submitted at onboarding"),
                    ("SERP","pid=22","w=0.5","domain found via web scraping"),
                    ("Verdata","pid=35","w=0.6","seller.domain_name"),
                    ("Equifax","pid=17","w=0.7","efx_web field")],
         "rule":"factWithHighestConfidence(). Custom renderer renderWebsiteField() shows ExternalLink icon when value exists. formatUrl() formats for display.",
         "storage":["rds_warehouse_public.facts name='website'","rds_warehouse_public.facts name='website_found'"],
         "null_cause":"N/A when no website submitted and no vendor found a domain.",
         "badges":[],
         "react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/config/WebsiteTab/fieldConfigs.tsx",1,60,"website fieldConfig"),
         "sql":"SELECT name, JSON_EXTRACT_PATH_TEXT(value, 'value') FROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name IN ('website','website_found');"},
        {"label":"Creation Date (WHOIS)","fact":"websiteData.domain.creation_date","api":"/verification/businesses/:id/website-data","editable":False,
         "tags":["🔒 Read-Only","📅 WHOIS Data","⚠️ N/A when URL edited"],
         "sources":[("SERP","pid=22","w=0.5","WHOIS lookup on found domain → creation_date"),
                    ("Verdata","pid=35","w=0.6","domain age check via WHOIS")],
         "rule":"useWebsiteNonEditableFields: isWebsiteDirty=true → N/A. websiteData.domain.creation_date → formatSourceDate().",
         "storage":["integration_data.request_response (SERP/Verdata WHOIS data)","websiteData.domain.creation_date via /website-data endpoint"],
         "null_cause":"N/A when: (1) no website found, (2) WHOIS privacy protection enabled, (3) URL was just edited (isWebsiteDirty=true).",
         "badges":[],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/hooks/WebsiteTab/useWebsiteNonEditableFields.tsx",60,80,"Creation Date field"),
         "sql":"-- Website data not in facts table — comes from /website-data endpoint\n-- Check SERP raw response:\nSELECT response FROM integration_data.request_response\nWHERE business_id='YOUR-BUSINESS-UUID-HERE' AND platform_id=22 ORDER BY requested_at DESC LIMIT 1;"},
        {"label":"Expiration Date (WHOIS)","fact":"websiteData.domain.expiration_date","api":"/verification/businesses/:id/website-data","editable":False,
         "tags":["🔒 Read-Only","📅 WHOIS Data","⚠️ N/A when URL edited"],
         "sources":[("SERP","pid=22","w=0.5","WHOIS expiration date"),("Verdata","pid=35","w=0.6","domain expiration")],
         "rule":"Same as Creation Date. isWebsiteDirty=true → N/A.","storage":["Via /website-data endpoint"],"null_cause":"N/A same reasons as Creation Date.","badges":[],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/hooks/WebsiteTab/useWebsiteNonEditableFields.tsx",80,95,"Expiration Date field"),"sql":"-- Same query as Creation Date above"},
        {"label":"Parked Domain","fact":"websiteData.parked","api":"/verification/businesses/:id/website-data","editable":False,
         "tags":["🔒 Read-Only","⚠️ N/A when URL edited"],
         "sources":[("SERP","pid=22","w=0.5","SERP scraping domain parking analysis")],
         "rule":"websiteData.parked → true/false/null. isWebsiteDirty=true → N/A.","storage":["Via /website-data endpoint"],"null_cause":"'No' = not parked or unknown. N/A when URL edited.","badges":[("No","badge-verified","parked=false — domain is actively used"),("Yes","badge-unverified","parked=true — domain not in active use")],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/hooks/WebsiteTab/useWebsiteNonEditableFields.tsx",95,110,"Parked Domain field"),"sql":"-- Same SERP raw response query as above"},
        {"label":"Status + badge","fact":"websiteData.status","api":"/verification/businesses/:id/website-data","editable":False,
         "tags":["🔒 Read-Only","🔵 Badge","⚠️ N/A when URL edited"],
         "sources":[("SERP","pid=22","w=0.5","website status check: online/offline/unknown"),
                    ("Verdata","pid=35","w=0.6","website availability check")],
         "rule":"status='online'→CheckCircleIcon badge (verified). status=other→secondary badge. isWebsiteDirty=true→'Unknown' badge.",
         "storage":["Via /website-data endpoint"],"null_cause":"'Unknown' = verification not complete or URL just edited.",
         "badges":[("✅ Online","badge-online","status='online' — website is live and accessible"),("❌ Offline","badge-offline","status='offline' — website not reachable"),("Unknown","badge-unknown","status unknown or URL was just edited")],
         "react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/hooks/WebsiteTab/useWebsiteNonEditableFields.tsx",110,128,"Status field"),
         "sql":"-- Check Verdata raw response:\nSELECT response FROM integration_data.request_response WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND platform_id=35 ORDER BY requested_at DESC LIMIT 1;"},
    ]
    for f in WEB_FIELDS: render_field(f)


# ════════════════════════════════════════════════════════════════════════════
# TAB 5 — WATCHLISTS
# ════════════════════════════════════════════════════════════════════════════
elif tab == "🔍 Watchlists":
    sh("🔍 Watchlists Sub-Tab — Every Card, Every Field, Full Lineage")

    card("""<b>Cards in Watchlists sub-tab:</b><br>
    1. <b>Watchlists Scanned card</b> (right panel) — shows all 14 lists with hit status per list<br>
    2. <b>Watchlist Hit card per entity</b> — one card per scanned entity (business names + officers)<br>
    3. <b>No Hits state</b> — WatchlistHitNullState shows green "No Hits Found"<br><br>
    <b>API:</b> <code>GET /facts/business/:id/kyb</code> → <code>watchlist.value.metadata</code><br>
    <b>IMPORTANT:</b> ALL hits (business + person) are NOW consolidated in <code>watchlist.value.metadata</code>.<br>
    No separate <code>/people/watchlist</code> endpoint needed (changed in BEST-65 via <code>trulioo_advanced_watchlist_results</code>).<br>
    <b>Source:</b> <code>WatchlistsTab.tsx</code> + <code>WatchlistsScannedCard.tsx</code> + <code>WatchlistHitCard.tsx</code>
    """)
    st.markdown(src("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/WatchlistsTab/WatchlistsTab.tsx",186,215,"Backend consolidates all hits now"), unsafe_allow_html=True)

    w_card = st.radio("Card", ["📋 Watchlists Scanned Card","🎯 Hit Cards (per entity)"], horizontal=True)

    if w_card == "📋 Watchlists Scanned Card":
        st.markdown("### 📋 Watchlists Scanned Card (right panel)")
        card("""Shows all 14 watchlists grouped by agency. For each list: hit badge or 'No Hits' badge.
        Lists are hardcoded in <code>Watchlists.ts</code> constant — the same 14 lists for all businesses.
        Source: <code>WatchlistsScannedCard.tsx</code> — renders WatchlistHitItem or WatchlistNoHitItem per list.""")

        st.markdown("#### The 14 Watchlists Checked:")
        lists_html = '<table class="t"><tr><th>Agency</th><th>List Name</th><th>Source</th></tr>'
        watchlists = [
            ("OFAC","Specially Designated Nationals (SDN)","Middesk + Trulioo"),
            ("OFAC","Capta List Foreign Sanctions Evaders","Middesk + Trulioo"),
            ("OFAC","Non-SDN Menu-Based Sanctions","Middesk + Trulioo"),
            ("OFAC","Non-SDN Iranian Sanctions","Middesk + Trulioo"),
            ("OFAC","Non-SDN Chinese Military-Industrial Complex","Middesk + Trulioo"),
            ("OFAC","Non-SDN Palestinian Legislative Council List","Middesk + Trulioo"),
            ("OFAC","Sectoral Sanctions Identifications List","Middesk + Trulioo"),
            ("BIS","Entity List","Middesk + Trulioo"),
            ("BIS","Denied Persons List","Middesk + Trulioo"),
            ("BIS","Military End User","Middesk + Trulioo"),
            ("BIS","Unverified List","Middesk + Trulioo"),
            ("State Dept","ITAR Debarred","Middesk + Trulioo"),
            ("State Dept","Nonproliferation Sanctions","Middesk + Trulioo"),
            ("Other","Adverse Media (→ Public Records tab)","Trulioo — FILTERED OUT from this tab"),
        ]
        for agency, name, source in watchlists:
            c = "#F87171" if "FILTERED" in source else "#94A3B8"
            lists_html += f"<tr><td style='color:#60A5FA;font-weight:600'>{agency}</td><td style='color:#E2E8F0'>{name}</td><td style='color:{c};font-size:.82rem'>{source}</td></tr>"
        st.markdown(lists_html + "</table>", unsafe_allow_html=True)
        card("⚠️ <b>Adverse media hits are FILTERED OUT from this tab.</b> combineWatchlistMetadata: filteredMetadata = allMetadata.filter(hit => hit.type !== WATCHLIST_HIT_TYPE.ADVERSE_MEDIA). Adverse media → Public Records tab.", "card-amber")

        WLIST_FIELDS = [
            {"label":"Watchlist hit count badge + entity grouping","fact":"watchlist.value.metadata","api":"/facts/business/:id/kyb","editable":False,
             "tags":["🔒 Read-Only","All Entities"],
             "sources":[("Middesk","pid=16","—","reviewTasks[type='watchlist'] — OFAC/BIS/State Dept checks"),
                        ("Trulioo","pid=38","—","business + person screening via trulioo_advanced_watchlist_results"),
                        ("Combined","—","—","combineWatchlistMetadata() deduplicates by type|title|entity_name|url")],
             "rule":"combineWatchlistMetadata rule. Groups hits by entity_name via groupWatchlistHitsByEntityName(). Shows WatchlistHitCard (hits found) or WatchlistHitNullState (no hits) per entity.",
             "storage":["rds_warehouse_public.facts name='watchlist' {metadata:[{type,metadata:{title,agency},entity_name,url,entity_type}],message:''}","rds_warehouse_public.facts name='watchlist_hits' (integer count)","rds_warehouse_public.facts name='watchlist_raw' (full merged before filtering)"],
             "null_cause":"'No Hits Found' = metadata array is empty after deduplication. This is the NORMAL expected state. Section hidden for Canadian businesses (excludedCountriesForVerification).",
             "badges":[("✅ No Hits","badge-nohits","metadata.length=0 — watchlist scan completed with no matches"),("🔴 N Hits Found","badge-hits","metadata.length>0 — N unique hits found after deduplication")],
             "react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/WatchlistsTab/WatchlistsTab.tsx",190,215,"groupWatchlistHitsByEntityName"),
             "sql":"-- Fetch raw value text and parse in Python (see Python snippet)\nSELECT name, value, received_at\nFROM rds_warehouse_public.facts\nWHERE business_id = 'YOUR-BUSINESS-UUID-HERE'\n  AND name = 'fact_name';"},
        ]
        for f in WLIST_FIELDS: render_field(f)

    elif w_card == "🎯 Hit Cards (per entity)":
        st.markdown("### 🎯 Watchlist Hit Card (per entity)")
        card("""One WatchlistHitCard is rendered per entity name (business names + officer/director names).<br>
        Each card shows: entity name + 'No Hits' OR 'N Hits Found' badge + individual hit details.<br>
        Entity names come from: <code>names_submitted</code> (business) + <code>people.value[]</code> (officers).<br>
        Hits are matched to entities via <code>groupWatchlistHitsByEntityName()</code>.""")

        card("""<b>Each Hit shows:</b><br>
        • <b>List name</b> (agency abbreviation — e.g. 'SDN', 'Entity List')<br>
        • <b>Entity name</b> matched on the list<br>
        • <b>Country</b> of the watchlist<br>
        • <b>Source URL</b> — link to the actual watchlist entry (if available; some lists are CSV-only)<br><br>
        <b>Hit data structure:</b> <code>{type, metadata:{title, agency}, entity_name, url, entity_type:'BUSINESS'|'PERSON', list_country}</code>""")

        st.code("""-- Step 1: fetch raw watchlist value (CONFIRMED WORKING)
SELECT name, value, received_at
FROM rds_warehouse_public.facts
WHERE business_id = 'YOUR-BUSINESS-UUID-HERE'
  AND name = 'watchlist_raw';

-- Step 2: parse hits in Python (see Python snippet below)
-- import json
-- fact = json.loads(value_text)
-- outer = json.loads(fact['value']) if isinstance(fact['value'], str) else fact['value']
-- hits = outer.get('metadata', [])
-- for hit in hits:
--     print(hit.get('type'), hit.get('entity_name'), hit.get('url'))
--     print('  List:', hit.get('metadata', {}).get('title'))""", language="sql")


# ════════════════════════════════════════════════════════════════════════════
# TAB 6 — KYB API FIELDS (UCM)
# ════════════════════════════════════════════════════════════════════════════
elif tab == "🔑 KYB API Fields":
    sh("🔑 KYB API Field Names — UCM Working Session Lineage")

    card(f"""<b>Source:</b> <code>[GPN version] Worth Field Outputs_Working Session Disussions_020426.xlsx</code> — <code>[WS] UCM Q/A</code> tab column E.<br>
    <b>Total API fields loaded:</b> <b>{len(UCM_FIELDS)}</b> fields from the spreadsheet.<br>
    For each field: full lineage (Admin Portal location, fact name, vendors, confidence models, winner rule, storage tables, null/blank scenarios, badges/states, SQL to verify, Python to load, UCM Q&A).""")

    if not UCM_FIELDS:
        card("⚠️ Spreadsheet not found. Ensure the Admin-Portal-KYB repo is cloned at /tmp/Admin-Portal-KYB/", "card-red")
        st.code("git clone git@github.com:wecsleyprates-design/Admin-Portal-KYB.git /tmp/Admin-Portal-KYB", language="bash")
    else:
        sections = sorted(set(f['section'] for f in UCM_FIELDS if f['section']))
        selected_section = st.radio("Filter by Section", ["All"] + sections, horizontal=True)
        filtered = UCM_FIELDS if selected_section == "All" else [f for f in UCM_FIELDS if f['section'] == selected_section]

        sel_field = st.selectbox("Select API Field", [f['api_field'] for f in filtered])
        ucm = next((f for f in filtered if f['api_field'] == sel_field), None)

        # Full deep field catalog for all 28 UCM fields
        UCM_DEEP = {
            "tin.value": {"label":"Tax ID (EIN) + Verified Badge","fact":"tin / tin_match / tin_match_boolean","api":"/facts/business/:id/kyb","editable":True,"tags":["✏️ Editable","🔐 IRS TIN Match","UCM: use tin_match_boolean NOT tin.value"],"sources":[("Applicant","pid=0","w=1.0","EIN integer submitted at onboarding — always masked after submission. This IS tin.value (raw integer). NOT what UCM should use for verification."),("Middesk","pid=16","w=2.0","IRS TIN Match service: reviewTasks[key='tin'].status. Returns {status:'success'|'failure',message,sublabel}. THIS drives the Verified/Unverified badge via tin_match_boolean."),("Trulioo","pid=38","w=0.8","TIN-to-name match for UK/Canada businesses only. Falls back for US. truliooPreferredRule applies for GB/CA.")],"rule":"tin.value = raw masked EIN (applicant submitted). UCM MUST use tin_match_boolean.value (boolean) NOT tin.value (integer). tin_match_boolean = (tin_match.value.status === 'success'). TinBadge.tsx: true→✅Verified(info/blue), false→⚠️Unverified(warning). Winner: factWithHighestConfidence() — Middesk wins (w=2.0) for US.","storage":["rds_warehouse_public.facts name='tin_submitted' (masked EIN string)","rds_warehouse_public.facts name='tin_match' {status,message,sublabel}","rds_warehouse_public.facts name='tin_match_boolean' (boolean: status==='success')","integration_data.request_response (pid=16 Middesk raw IRS response)"],"null_cause":"Unverified badge: IRS did not confirm TIN+name match. Causes: wrong EIN, name mismatch, sole proprietor using personal SSN, EIN not yet in IRS records. Blank tin.value: applicant did not submit (extremely rare).","badges":[("✅ Verified","badge-verified","tin_match_boolean=true → IRS confirmed EIN+legal name are a valid registered match"),("⚠️ Unverified","badge-unverified","tin_match_boolean=false → IRS could not confirm match"),("⚠️ [status]","badge-unverified","capitalize(tin_match.value.status) when status is neither success nor failure (e.g. 'in_review')")],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/components/TinBadge.tsx",1,48,"TinBadge exact badge logic"),"sql":"SELECT name, JSON_EXTRACT_PATH_TEXT(value, 'value') AS val, JSON_EXTRACT_PATH_TEXT(value, 'source', 'platformId') AS pid\nFROM rds_warehouse_public.facts\nWHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name IN ('tin','tin_submitted','tin_match','tin_match_boolean');","gpn_q":"tin.value is an integer — how does it transform to Verified/Unverified? Are you talking about tin_match_boolean.value?","gpn_a":"UCM should use tin_match_boolean.value (includes true or false). tin.value is the raw tax id. Unverified maps to False, Verified maps to True. UCM Rule: Unverified → Fail the rule. All other results → Pass the rule."},
            "watchlist.hits.value": {"label":"Watchlist Hit Count","fact":"watchlist_hits / watchlist.value.metadata.length","api":"/facts/business/:id/kyb","editable":False,"tags":["🔒 Read-Only","14 Lists Scanned","All entities consolidated"],"sources":[("Middesk","pid=16","—","reviewTasks[type='watchlist'] — OFAC, BIS, State Dept lists for the business entity"),("Trulioo","pid=38","—","Business + person watchlist screening via trulioo_advanced_watchlist_results (ALL hits consolidated since BEST-65)"),("Combined","combineWatchlistMetadata()","—","Merges ALL hits, deduplicates by type|title|entity_name|url key. Adverse media FILTERED OUT (→ Public Records tab).")],"rule":"watchlist.hits.value = watchlist.value.metadata.length (count after deduplication). WatchlistsTab.tsx: badge = metadata.length>0 ? 'N Hits Found'(red) : 'No Hits'(green). groupWatchlistHitsByEntityName() groups hits by entity name across business names + officers. IMPORTANT: since BEST-65, NO separate /people/watchlist endpoint — all consolidated.","storage":["rds_warehouse_public.facts name='watchlist_hits' (integer count)","rds_warehouse_public.facts name='watchlist' {metadata:[{type,entity_name,url}],message:''}","rds_warehouse_public.facts name='watchlist_raw' (full pre-filter data)"],"null_cause":"0 hits = watchlist scan ran and found nothing (NORMAL expected state). NULL hit_count = watchlist scan not yet completed (integrations still processing). Section hidden for Canadian businesses (excludedCountriesForVerification).","badges":[("✅ No Hits","badge-nohits","metadata.length=0 — all 14 lists scanned, no matches found"),("🔴 N Hits Found","badge-hits","metadata.length>0 — N unique hits across OFAC/BIS/State Dept after deduplication")],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/WatchlistsTab/WatchlistsTab.tsx",190,215,"groupWatchlistHitsByEntityName logic"),"sql":"SELECT JSON_EXTRACT_PATH_TEXT(value, 'value', 'metadata') FROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name='watchlist';\n\n-- Count hits:\nSELECT JSON_ARRAY_LENGTH(value, true)  -- parse in Python instead AS hit_count\nFROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name='watchlist_raw';"},
            "watchlist.value.metadata": {"label":"Watchlist Hit Detail Metadata","fact":"watchlist.value.metadata[]","api":"/facts/business/:id/kyb","editable":False,"tags":["🔒 Read-Only","Array per Hit","Adverse Media Excluded"],"sources":[("Middesk+Trulioo combined","combineWatchlistMetadata()","—","rules.ts L273-344: deduplicates by type|metadata.title|entity_name|url. Filters out ADVERSE_MEDIA type (→ Public Records tab).")],"rule":"combineWatchlistMetadata() rule. Each hit: {type, metadata:{title,agency}, entity_name, url, entity_type:'BUSINESS'|'PERSON', list_country}. Dedup key: type|title/agency|entity_name|url.","storage":["rds_warehouse_public.facts name='watchlist_raw' (full merged before adverse-media filter)"],"null_cause":"Empty array = no hits found. Adverse media hits present in watchlist_raw but EXCLUDED from watchlist.value.metadata (filtered out by combineWatchlistMetadata).","badges":[],"react_src":("SIC_UK","integration-service/lib/facts/rules.ts",273,344,"combineWatchlistMetadata deduplication and adverse media filter"),"sql":"-- Fetch raw value text and parse in Python (see Python snippet)\nSELECT name, value, received_at\nFROM rds_warehouse_public.facts\nWHERE business_id = 'YOUR-BUSINESS-UUID-HERE'\n  AND name = 'fact_name';"},
            "legal_name.value": {"label":"Legal Business Name (Business Registration card)","fact":"legal_name","api":"/facts/business/:id/kyb","editable":True,"tags":["✏️ Editable","🔍 Vendor-Discovered","NOT applicant-submitted"],"sources":[("Middesk","pid=16","w=2.0","businessEntityVerification.name — name found in SoS records via IRS TIN match"),("OpenCorporates","pid=23","w=0.9","firmographic.name — registered name in global OC registry"),("Trulioo","pid=38","w=0.8","clientData.businessName for UK/Canada")],"rule":"factWithHighestConfidence() — Middesk wins (w=2.0) for US. legal_name.value is the VERIFIED name from public records. IMPORTANT: NOT pre-filled by Worth — Global/UCM sends Legal Name + Tax ID. Worth compares submitted name against IRS via Middesk.","storage":["rds_warehouse_public.facts name='legal_name'"],"null_cause":"Empty when: (1) Middesk still processing integrations, (2) Middesk found no entity in SoS. NOT a failure — means entity lookup not yet complete.","badges":[],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/hooks/BusinessRegistrationTab/useBusinessRegistrationTabDetails.tsx",72,82,"taxDetails Business Name"),"sql":"SELECT JSON_EXTRACT_PATH_TEXT(value, 'value') AS legal_name, JSON_EXTRACT_PATH_TEXT(value, 'source', 'platformId') AS pid\nFROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name='legal_name';","gpn_q":"Worth Discussion: We don't see company_name in KYB documentation. UCM found legal_name, but not company_name.value. What UCM rule is needed?","gpn_a":"Per Gavin, the field is legal_name. Per Ben, this is NOT pre-filled. Global sends Legal Name & Tax ID. Worth compares against IRS. Tax ID Verification & Legal Name use the SAME UCM Rule. No Action Needed."},
            "dba_found.value[n]": {"label":"DBA Names (Contact Information)","fact":"dba_found (array)","api":"/facts/business/:id/kyb","editable":True,"tags":["✏️ Editable","Array of DBAs","combineFacts merges all"],"sources":[("Middesk","pid=16","w=2.0","names[] where submitted=false — trade names Middesk found in SoS"),("ZoomInfo","pid=24","w=0.8","zi_c_tradename field"),("OpenCorporates","pid=23","w=0.9","alternate_names from OC registry"),("SERP","pid=22","w=0.5","DBA extracted from web scraping"),("Applicant","pid=0","w=1.0","DBA submitted at onboarding")],"rule":"combineFacts() — merges DBA names from ALL sources into deduplicated array. Verification via name_match_boolean.value (TRUE=Verified, FALSE=Unverified). Name comparison uses name_match_boolean not dba_found directly.","storage":["rds_warehouse_public.facts name='dba_found' (array of discovered DBAs)","rds_warehouse_public.facts name='dba' (submitted DBA from applicant)"],"null_cause":"Empty array = business operates only under legal name. Very common — NOT an error. Section only renders when names_submitted.value.length > 0.","badges":[("✅ Verified","badge-verified","name_match_boolean=true: submitted name/DBA found in SoS/public records"),("❌ Failure","badge-missing","name_match_boolean=false: name not confirmed")],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/components/BusinessNamesCard.tsx",1,30,"BusinessNamesCard"),"sql":"SELECT name, JSON_EXTRACT_PATH_TEXT(value, 'value') FROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name IN ('dba','dba_found');\nSELECT JSON_EXTRACT_PATH_TEXT(value, 'value') AS name_match_boolean FROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name='name_match_boolean';","gpn_q":"How does dba_found.value[n] (array) transform to Verified/Unverified? DBA Verified: Rule passes. DBA Unverified: Rule fails.","gpn_a":"name_match_boolean.value governs DBA Names, alternate names. True=Verified, False=Unverified. Worth 360 Report shows Verified or Unverified. Logic uses data including confidence score."},
            "google_profile.address": {"label":"Google Business Profile — Address","fact":"address_verification (Google component)","api":"/facts/business/:id/kyb + Google Places API","editable":False,"tags":["🔒 Read-Only","SERP→Google Places","Match/No Match"],"sources":[("SERP","pid=22","w=0.5","SERP scraper finds google_place_id → calls Google Places API → gets address"),("Google Places API","pid=29","w=0.6","address_match field from business listing")],"rule":"AddressesCard.tsx: googleProfileMatch = (googleProfileData.data.business_match.toLowerCase()==='match found') AND (googleProfileData.data.address_match.toLowerCase()==='match'). enrichAddressesWithStatusFor360ReportParity() applies to submitted AND reported addresses.","storage":["rds_warehouse_public.facts name='address_verification' {status,sublabel,message}","rds_warehouse_public.facts name='address_match_boolean' (boolean)","integration_data.request_response (pid=22 SERP, pid=29 Google)"],"null_cause":"Google Profile badge shows 'Unverified' when: (1) no Google Business Profile found, (2) SERP found no Place ID, (3) Google address doesn't match submitted. Many businesses have no Google listing — NOT an error.","badges":[("✅ Google Profile","badge-verified","googleProfileMatch=true: Google BProfile confirms same business name AND same address"),("⚠️ Google Profile","badge-unverified","googleProfileMatch=false: Google not found or shows different address/name")],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/components/AddressesCard.tsx",218,228,"googleProfileMatch logic"),"sql":"SELECT name, JSON_EXTRACT_PATH_TEXT(value, 'value') FROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name IN ('address_verification','address_match_boolean','google_place_id');","gpn_q":"What is the Google-Profile endpoint? What UCM rule is needed? Need to display on Worth 360 Report.","gpn_a":"Action Item: Gavin to provide Google Profile documentation. UCM Team can investigate Worth 360 mapping to find these details. Note: This info is on the Worth 360 Report."},
            "google_profile.business_name": {"label":"Google Business Profile — Business Name","fact":"names_found + name_match_boolean","api":"/facts/business/:id/kyb","editable":False,"tags":["🔒 Read-Only","SERP→Google Places"],"sources":[("SERP+Google Places","pid=22,29","—","Google Business Profile name field. Compared against submitted name for Match/No Match.")],"rule":"names_found uses combineFacts (all sources merged). name_match_boolean governs Verified badge. Google name contributes to name match evaluation.","storage":["rds_warehouse_public.facts name='names_found' (array)","rds_warehouse_public.facts name='name_match_boolean' (boolean)"],"null_cause":"No Google Business Profile found — SERP could not find a Google Place ID for this business.","badges":[("✅ Verified","badge-verified","name_match_boolean=true: submitted business name matches public records including Google"),("❌ Failure","badge-missing","name_match_boolean=false: name not confirmed")],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/components/BusinessNamesCard.tsx",1,30,"Business Names section"),"sql":"SELECT JSON_EXTRACT_PATH_TEXT(value, 'value') FROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name IN ('names_found','name_match_boolean');"},
            "mcc_code": {"label":"MCC Code (Background tab)","fact":"mcc_code / mcc_code_found / mcc_code_from_naics","api":"/facts/business/:id/details","editable":True,"tags":["✏️ Editable","🤖 AI + Calculated","3 internal facts"],"sources":[("AI Enrichment (mcc_code_found)","pid=31","—","response.mcc_code directly from GPT-5-mini (preferred). Fires when all vendor NAICS are null."),("Calculated from NAICS (mcc_code_from_naics)","calculated","—","rel_naics_mcc lookup: winning naics_id → mcc_id. Dependency on naics_code fact."),("Combined (mcc_code)","combined","—","mcc_code = mcc_code_found?.value ?? mcc_code_from_naics?.value. AI-provided preferred.")],"rule":"mcc_code is NOT from a competitive vendor selection. It is: (1) AI-provided MCC if AI enrichment ran, OR (2) calculated from winning NAICS via rel_naics_mcc table. When AI returns 5614 (last resort), mcc_description = 'Fallback MCC per instructions...' — KNOWN BUG (Gap G5), editable=false so analyst cannot fix.","storage":["rds_warehouse_public.facts name='mcc_code'","rds_warehouse_public.facts name='mcc_code_found' (AI direct)","rds_warehouse_public.facts name='mcc_code_from_naics' (calculated)","rds_cases_public.data_businesses.mcc_id → core_mcc_code.id"],"null_cause":"Almost never null — rel_naics_mcc calculates MCC when NAICS exists. 5614 shown with 'Fallback MCC...' when AI had no evidence (Gap G5 — mcc_description is read-only).","badges":[],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/config/BackgroundTab/fieldConfigs.tsx",440,469,"mcc_code + mcc_description fieldConfigs"),"sql":"SELECT name, JSON_EXTRACT_PATH_TEXT(value, 'value') FROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name IN ('mcc_code','mcc_code_found','mcc_code_from_naics','mcc_description');","gpn_q":"MCC code is needed for UCM. Where does it come from?","gpn_a":"MCC comes from AI enrichment (NAICS→MCC map + AI-returned MCC). 5614 is the AI fallback when no industry evidence. 'Fallback MCC per instructions...' description is a known bug."},
            "people.value[n].name, people.value[n].titles[n]": {"label":"Corporate Officers (Business Registration)","fact":"people.value[] filtered by sos.id","api":"/facts/business/:id/kyb","editable":False,"tags":["🔒 Read-Only","Per Filing Filtered","combineFacts"],"sources":[("Middesk","pid=16","w=2.0","registrations[n].officers — officers per state filing, matched to specific sos.id"),("OpenCorporates","pid=23","w=0.9","officers array from OC company record"),("Trulioo","pid=38","w=0.8","principals/directors from Trulioo KYB response")],"rule":"combineFacts() — all people from all sources merged. Business Registration tab shows: people.value.filter(person => person.source.some(src => src.id === sos.id)). Each SoS filing card shows ONLY officers linked to THAT specific filing's sos.id.","storage":["rds_warehouse_public.facts name='people' (array of {name, titles[], source[], jurisdictions[]})"],"null_cause":"N/A per filing when no officers linked to that sos.id. Common when state doesn't require officer disclosure. Entire people section empty if no SoS filing found.","badges":[],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/hooks/BusinessRegistrationTab/useBusinessRegistrationTabDetails.tsx",197,240,"Corporate Officers row — filtered by sos.id"),"sql":"SELECT JSON_EXTRACT_PATH_TEXT(value, 'value') AS people FROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name='people';"},
            "sos_active.value": {"label":"SoS Filing Status (Active/Inactive)","fact":"sos_filings[n].active / sos_active","api":"/facts/business/:id/kyb","editable":False,"tags":["🔒 Read-Only","Per Filing","Drives SoS badge"],"sources":[("Middesk","pid=16","w=2.0","registrations[n].status === 'active' → active=true. Middesk queries SoS databases by TIN+name."),("OpenCorporates","pid=23","w=0.9","current_status: 'Active'→true, 'Dissolved'→false")],"rule":"sos_active = any(sos_filings[n].active === true). SOSBadge.tsx: sosFiling.active=true→VERIFIED(info/blue,CheckBadgeIcon,tooltip='An active filing was found'). sosFiling.active=false→MISSING ACTIVE FILING(error/red). sos_active=null→N/A. SOSFilingCard title: US='Secretary of State Filings', non-US='Registration Filing'.","storage":["rds_warehouse_public.facts name='sos_filings' (array of filings, each with active: boolean)","rds_warehouse_public.facts name='sos_active' (boolean: any filing active?)","rds_warehouse_public.facts name='sos_match_boolean'"],"null_cause":"'No Registry Data to Display' = sos_filings.value=[] (empty array). Middesk searched all 50 US SoS databases by TIN+name and found ZERO registrations. NOT a confidence threshold — if Middesk finds nothing, nothing is shown.","badges":[("✅ Verified","badge-verified","sosFiling.active=true: active SoS filing found and in good standing"),("🔴 Missing Active Filing","badge-missing","sosFiling.active=false: filing found but inactive/dissolved"),("⚠️ Invalidated","badge-unverified","isInvalidated=true: filing marked invalid")],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/components/SOSBadge.tsx",12,60,"SOSBadge exact logic"),"sql":"-- Fetch raw value text and parse in Python (see Python snippet)\nSELECT name, value, received_at\nFROM rds_warehouse_public.facts\nWHERE business_id = 'YOUR-BUSINESS-UUID-HERE'\n  AND name = 'fact_name';"},
            "sos_filings.value[n].state": {"label":"SoS Filing State","fact":"sos_filings[n].state","api":"/facts/business/:id/kyb","editable":False,"tags":["🔒 Read-Only","Per Filing","2-letter code"],"sources":[("Middesk","pid=16","w=2.0","registrations[n].registration_state — 2-letter US state code"),("OpenCorporates","pid=23","w=0.9","jurisdiction_code split by '_' → second part uppercased: 'us_fl'→'FL'")],"rule":"Shown per filing card. Format: 2-letter state code. EntityJurisdictionCell.tsx maps foreign_domestic to Domestic(Primary) or Foreign(Secondary) badge — 'Primary' is Worth's label, NOT Middesk's.","storage":["Within sos_filings array: state field"],"null_cause":"N/A if registration_state absent from SoS record (uncommon).","badges":[],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/components/BusinessRegistrationTab/EntityJurisdictionCell.tsx",1,35,"Domestic/Foreign + Primary/Secondary badges"),"sql":"-- Fetch raw value text and parse in Python (see Python snippet)\nSELECT name, value, received_at\nFROM rds_warehouse_public.facts\nWHERE business_id = 'YOUR-BUSINESS-UUID-HERE'\n  AND name = 'fact_name';"},
            "sos_filings.value[n].filing_date": {"label":"SoS Registration Date","fact":"sos_filings[n].filing_date","api":"/facts/business/:id/kyb","editable":False,"tags":["🔒 Read-Only","ISO date → MM-DD-YYYY"],"sources":[("Middesk","pid=16","w=2.0","registrations[n].registration_date — ISO date string from SoS database"),("OpenCorporates","pid=23","w=0.9","incorporation_date from OC record")],"rule":"Formatted by convertToLocalDate(new Date(sos.filing_date), 'MM-DD-YYYY') in index.tsx L300. useBusinessRegistrationTabDetails.tsx: formatSourceDate() for display.","storage":["Within sos_filings array: filing_date field per filing"],"null_cause":"N/A if date absent from SoS registry record. Shows 'N/A' for that specific field.","badges":[],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/hooks/BusinessRegistrationTab/useBusinessRegistrationTabDetails.tsx",163,175,"Registration Date row"),"sql":"-- Fetch raw value text and parse in Python (see Python snippet)\nSELECT name, value, received_at\nFROM rds_warehouse_public.facts\nWHERE business_id = 'YOUR-BUSINESS-UUID-HERE'\n  AND name = 'fact_name';"},
            "sos_filings.value[n].entity_type": {"label":"SoS Entity Type (Corporation Type)","fact":"sos_filings[n].entity_type","api":"/facts/business/:id/kyb","editable":False,"tags":["🔒 Read-Only","Normalised","Per Filing"],"sources":[("Middesk","pid=16","w=2.0","registrations[n].entity_type — raw string from SoS (e.g. 'Llc')"),("OpenCorporates","pid=23","w=0.9","company_type normalised: 'Limited Liability Company'→'llc', 'Incorporated'→'corporation'")],"rule":"Worth normalises entity_type: 'Limited Liability Company'→'llc', 'Incorporated'→'corporation', 'Limited Liability Partnership'→'llp'. useBusinessRegistrationTabDetails.tsx L178-195.","storage":["Within sos_filings array: entity_type field"],"null_cause":"N/A if state doesn't disclose entity type. Different from Background tab Corporation Type (which is a separate independently-selected fact).","badges":[],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/hooks/BusinessRegistrationTab/useBusinessRegistrationTabDetails.tsx",178,195,"Entity Type row"),"sql":"-- Fetch raw value text and parse in Python (see Python snippet)\nSELECT name, value, received_at\nFROM rds_warehouse_public.facts\nWHERE business_id = 'YOUR-BUSINESS-UUID-HERE'\n  AND name = 'fact_name';"},
            "status": {"label":"Case Status (Case Results panel)","fact":"data_cases.status → core_case_statuses","api":"GET /cases/:id (case management API)","editable":True,"tags":["✏️ Analyst can change","Pending→Under Review/Approved","Not from facts table"],"sources":[("Worth Score engine","calculated","—","Auto-Approved when all UCM rules pass AND score above configured threshold"),("Manual","analyst","—","Analyst can change status via admin portal (PATCH /cases/:id)"),("Middesk","pid=16","—","businessEntityVerification.status feeds verification_status fact (separate from case status)")],"rule":"Case status lives in rds_cases_public.data_cases.status (FK to core_case_statuses). NOT a Fact Engine fact. Transitions: Pending→Under Review (rules fail) or Auto-Approved (all pass). 'Pending' = initial state while integrations run (not null). Worth Score shows '-' until calculation completes. case-management.ts getCaseByIDQuery L1555-1574 joins data_cases + core_case_statuses.","storage":["rds_cases_public.data_cases.status → core_case_statuses.id","core_case_statuses: {id, code:'pending'|'under_review'|'auto_approved'|'archived', label}"],"null_cause":"Shows 'Pending' (with loading bars) initially — not null. '-' for Worth Score means not yet calculated. 'Pending' disappears when all integrations complete and score is calculated.","badges":[("Pending","badge-unknown","Integrations still running, Worth Score not yet calculated"),("Under Review","badge-unverified","One or more UCM rules failed — analyst review required"),("Auto-Approved","badge-verified","All rules passed AND score above threshold"),("Archived","badge-unknown","Case closed — cannot reopen without action")],"react_src":("SIC_UK","case-service/src/api/v1/modules/case-management/case-management.ts",1555,1574,"getCaseByIDQuery — joins data_cases + core_case_statuses"),"sql":"SELECT dc.id, cs.code AS status, cs.label, dc.created_at\nFROM rds_cases_public.data_cases dc\nJOIN core_case_statuses cs ON cs.id = dc.status\nWHERE dc.business_id = 'YOUR-BUSINESS-UUID-HERE';"},
            "domain.creation_date": {"label":"Website Domain Creation Date (Website tab)","fact":"websiteData.domain.creation_date","api":"/verification/businesses/:id/website-data","editable":False,"tags":["🔒 Read-Only","WHOIS Data","N/A when URL edited"],"sources":[("SERP","pid=22","w=0.5","SERP scraper performs WHOIS lookup on found domain → creation_date field"),("Verdata","pid=35","w=0.6","domain age check via WHOIS")],"rule":"useWebsiteNonEditableFields.tsx: isWebsiteDirty=true → all non-editable fields show N/A (cleared when URL is edited). websiteData.domain.creation_date → formatSourceDate() for display. IMPORTANT: when analyst edits the website URL, all derived fields (Creation Date, Expiration Date, Parked Domain, Status) immediately show N/A until re-verification runs.","storage":["Via /verification/businesses/:id/website-data endpoint — not stored in facts table","integration_data.request_response (pid=22 SERP, pid=35 Verdata WHOIS data)"],"null_cause":"N/A when: (1) no website found, (2) WHOIS privacy protection enabled, (3) URL just edited (isWebsiteDirty=true). Many businesses have no website — N/A is common and NOT an error.","badges":[],"react_src":("ADMIN_PORTAL","microsites-main/packages/case/src/page/Cases/CaseDetails/Tabs/KYB/hooks/WebsiteTab/useWebsiteNonEditableFields.tsx",60,80,"Creation Date non-editable field"),"sql":"-- Not in facts table — check SERP raw response:\nSELECT response FROM integration_data.request_response WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND platform_id=22 ORDER BY requested_at DESC LIMIT 1;"},
        }

        if ucm:
            # Build a render_field-compatible dict from UCM data + deep catalog
            api_field = ucm['api_field']
            deep = UCM_DEEP.get(api_field)

            st.markdown(f"## `{api_field}`")
            m1, m2, m3, m4 = st.columns(4)
            m1.markdown(f"**Section:** {ucm['section'] or '—'}")
            m2.markdown(f"**Data Type:** {ucm['data_type'] or '—'}")
            m3.markdown(f"**W360:** {'✅ Yes' if ucm.get('w360')=='Yes' else '❌ No'}")
            m4.markdown(f"**Requires Transform:** {ucm.get('requires_transform') or '—'}")
            if ucm.get('description'):
                card(f"<b>Description (from UCM Working Session):</b> {ucm['description']}")
            if ucm.get('decisional'):
                card(f"<b>Decisional/Info Only:</b> {ucm['decisional']}")

            if deep:
                # Use the shared render_field pattern with deep data
                render_field({
                    "label": deep["label"],
                    "fact": deep["fact"],
                    "api": deep["api"],
                    "editable": deep.get("editable"),
                    "tags": deep.get("tags", []),
                    "sources": deep.get("sources", []),
                    "rule": deep.get("rule"),
                    "storage": deep.get("storage", []),
                    "null_cause": deep.get("null_cause"),
                    "badges": deep.get("badges", []),
                    "react_src": deep.get("react_src"),
                    "sql": deep.get("sql"),
                    "gpn_q": ucm.get("gpn_questions"),
                    "gpn_a": ucm.get("decisions"),
                })
            else:
                # Fallback for fields not yet in deep catalog
                if ucm.get("gpn_questions"):
                    st.markdown("### ❓ UCM Working Session Question")
                    card(ucm["gpn_questions"], "card-amber")
                if ucm.get("decisions"):
                    st.markdown("### ✅ Confirmed Decisions")
                    card(ucm["decisions"], "card-green")
                # Generic SQL
                fact_key = api_field.replace(".value","").replace("[n]","").replace("[]","").split(",")[0].strip()
                st.markdown("### 🔍 Verify in Database")
                st.code(f"SELECT name, value FROM rds_warehouse_public.facts WHERE business_id='YOUR-BUSINESS-UUID-HERE' AND name='{fact_key}';", language="sql")
                card(f"⚠️ Deep lineage for '{api_field}' not yet mapped. Use the 🤖 AI Agent tab to ask: 'Where does {api_field} come from?'", "card-amber")


# ════════════════════════════════════════════════════════════════════════════
# TAB 7 — AI AGENT
# ════════════════════════════════════════════════════════════════════════════
elif tab == "🤖 AI Agent":
    sh("🤖 KYB Intelligence Agent — Ask Anything About the Admin Portal")

    # ── How schema/source knowledge stays current ───────────────────────
    with st.expander("🔄 How Does the App Know the Correct Table Names, Column Names & Schema?", expanded=False):
        st.markdown("""
### Where Redshift & PostgreSQL Schema Knowledge Comes From

The app's knowledge of database table names, column names, and data types comes from **three
complementary sources**, each updated differently:

---

#### Source 1 — `api-docs/redshift-schema.md` (auto-generated, always indexed by RAG)

This file contains **every table, every column, and every data type** for all Worth AI
production tables. It is generated by running:

```bash
python3 Admin-Portal-KYB-App/sync_redshift_schema.py
```

**When you have Redshift/PostgreSQL credentials**, it connects live and fetches from
`information_schema.columns` — meaning it reflects the *exact current state of the database*,
including any new columns or renamed tables.

**Without credentials**, it uses the confirmed schemas provided directly (from `SELECT *
FROM information_schema.columns WHERE table_name = ...` queries you ran and shared).

This file is automatically re-generated during every **`sync-sources.yml` GitHub Actions run**
(Monday 6am UTC or manual trigger). If `REDSHIFT_*` secrets are configured in the repo,
the workflow connects live.

---

#### Source 2 — `warehouse-service/customer_table.sql` (auto-synced via GitHub Actions)

Pipeline B's winner-takes-all SQL is synced from the private `joinworth/warehouse-service`
repo. This SQL contains the **authoritative list of all `datascience.customer_files` columns**
and how they are computed (ZI vs EFX `CASE WHEN` logic). Every column in Pipeline B's output
table is documented in this one file.

---

#### Source 3 — Hardcoded in `kyb_portal_app.py` (confirmed manually, rarely changes)

Some schema facts were confirmed directly from your `information_schema.columns` queries and
embedded directly into the app:

| Schema fact | Where in app | How to update |
|---|---|---|
| `rds_warehouse_public.facts` column list | Database Reference panel | Edit `kyb_portal_app.py` `pg_tbl` block |
| `value` is `varchar` (not `jsonb`) | Every SQL code block | Already confirmed — will not change |
| `similarity_index` 0-55 formula | SQL/Python generators | Edit `generate_sql()` comments |
| Platform IDs (16=Middesk, 24=ZI, ...) | AI Agent system prompt | Edit `AGENT_SYSTEM_PROMPT` constant |

---

#### When a schema changes — what to do

| What changed | What to run | Files updated |
|---|---|---|
| New column added to `rds_warehouse_public.facts` | `python3 sync_redshift_schema.py` | `api-docs/redshift-schema.md` |
| New Redshift table added | `python3 sync_redshift_schema.py` (add table to `TABLES` list first) | `api-docs/redshift-schema.md` |
| Pipeline B SQL logic changed | GitHub Actions `sync-sources.yml` runs automatically | `warehouse-service/customer_table.sql` |
| New fact field added to `integration-service` | GitHub Actions `sync-sources.yml` | `integration-service/lib/facts/` |
| New Admin Portal component added | GitHub Actions `sync-sources.yml` | `microsites-main/packages/case/src/` |

After any of the above, rebuild the RAG index:

```bash
cd Admin-Portal-KYB-App
python3 kyb_rag_builder.py
git add api-docs/redshift-schema.md kyb_rag_index.json
git commit -m "sync: updated schema/source knowledge"
git push
```

The AI Agent will then use the updated knowledge on next restart.

---

#### How to verify what the AI Agent currently knows about a table

Ask the agent:
> *"What columns does `rds_warehouse_public.facts` have? Where did you find this?"*

The agent will cite the source chunk from `api-docs/redshift-schema.md` (or the source code file)
and tell you the exact line numbers it retrieved from.
""")

    # ── DB connection helper panel ──────────────────────────────────────
    with st.expander("🗄️  Complete Database Reference — All Tables, All Columns, All Queries", expanded=False):
        st.info(
            "**Source of truth for all column names:** `Admin-Portal-KYB-App/api-docs/redshift-schema.md` — "
            "auto-generated by `sync_redshift_schema.py`. Run it after any schema change, then rebuild the "
            "RAG index with `python3 kyb_rag_builder.py`. See the '🔄 How Does the App Know...' panel above "
            "for full instructions.",
            icon="🔄"
        )
        st.markdown("### Connection Templates")
        col_pg, col_rs = st.columns(2)
        with col_pg:
            st.markdown("**PostgreSQL (RDS) — port 5432**")
            st.code("""import psycopg2

pg = psycopg2.connect(
    host='your-rds-endpoint.rds.amazonaws.com',
    port=5432,
    dbname='postgres',
    user='your_pg_username',
    password='your_pg_password',
    sslmode='require'
)""", language="python")
        with col_rs:
            st.markdown("**Redshift — port 5439**")
            st.code("""import psycopg2

rs = psycopg2.connect(
    host='your-cluster.redshift.amazonaws.com',
    port=5439,
    dbname='dev',
    user='your_rs_username',
    password='your_rs_password',
    sslmode='require'
)""", language="python")

        st.markdown("---")
        st.markdown("### PostgreSQL Tables — Confirmed from source code")
        pg_tbl = """<table class="t"><tr><th>Schema.Table</th><th>Key columns (all confirmed)</th><th>Source file</th></tr>
        <tr><td><code>rds_warehouse_public.facts</code></td>
            <td><code>business_id</code> VARCHAR (PK), <code>name</code> VARCHAR (PK), <code>value</code> JSONB, <code>received_at</code> TIMESTAMPTZ<br>
            JSONB structure: <code>{"value":..., "source":{"platformId":N,"confidence":X,"updatedAt":""}, "override":{"value":"","userId":""}, "alternatives":[...]}</code></td>
            <td>facts.py L8-16</td></tr>
        <tr><td><code>rds_cases_public.data_businesses</code></td>
            <td><code>id</code> UUID, <code>name</code> VARCHAR, <code>naics_id</code>→core_naics_code, <code>mcc_id</code>→core_mcc_code, <code>industry</code>→core_business_industries, <code>tin</code> (masked)</td>
            <td>case-management.ts L1555</td></tr>
        <tr><td><code>rds_cases_public.data_cases</code></td>
            <td><code>id</code> UUID, <code>business_id</code> FK, <code>status</code>→core_case_statuses, <code>created_at</code>, <code>customer_id</code></td>
            <td>case-management.ts L1555</td></tr>
        <tr><td><code>rds_cases_public.data_owners</code></td>
            <td><code>id</code>, <code>first_name</code>, <code>last_name</code>, <code>email</code>, <code>mobile</code>, <code>date_of_birth</code> (encrypted), <code>ssn</code> (masked last 4)</td>
            <td>case-management.ts</td></tr>
        <tr><td><code>integration_data.request_response</code></td>
            <td><code>business_id</code>, <code>platform_id</code> INT, <code>response</code> JSONB (full raw vendor response), <code>requested_at</code></td>
            <td>sources.ts L49</td></tr>
        <tr><td><code>core_naics_code</code></td><td><code>id</code>, <code>code</code> VARCHAR(6), <code>label</code> VARCHAR</td><td>case-management.ts</td></tr>
        <tr><td><code>core_mcc_code</code></td><td><code>id</code>, <code>code</code> VARCHAR(4), <code>label</code> VARCHAR</td><td>case-management.ts</td></tr>
        <tr><td><code>core_business_industries</code></td><td><code>id</code>, <code>name</code> VARCHAR, <code>sector_code</code> INT (2-digit NAICS prefix)</td><td>businessDetails/index.ts L297</td></tr>
        <tr><td><code>rel_naics_mcc</code></td><td><code>naics_id</code>→core_naics_code, <code>mcc_id</code>→core_mcc_code</td><td>businessDetails/index.ts L408</td></tr>
        </table>"""
        st.markdown(pg_tbl, unsafe_allow_html=True)

        st.markdown("---")
        st.markdown("### Redshift Tables — All columns confirmed from schemas you provided")

        rs_tables = [
            ("datascience.customer_files", "Pipeline B output (winner-takes-all: ZI vs EFX)",
             ["business_id", "primary_naics_code INT (ZI or EFX winner)", "mcc_code", "mcc_desc",
              "worth_score", "match_confidence FLOAT", "zi_match_confidence FLOAT", "efx_match_confidence FLOAT",
              "company_name (from ZI or EFX winner)", "tax_id", "employee_count INT", "year_established INT",
              "name_verification", "address_verification", "tin_verification", "watchlist_verification",
              "sos_domestic_verification", "sos_match_verification",
              "corporate_filing_business_name", "corporate_filing_filling_date",
              "corporate_filing_incorporation_state", "corporate_filing_corporation_type",
              "corporate_filing_secretary_of_state_status", "primary_naics_description",
              "phone_number", "latitude FLOAT", "longitude FLOAT", "gov_flag", "nonprofit_flag", "edu_flag",
              "sos_online_firm_coverage", "average_rating", "total_review_count"],
             "customer_table.sql"),
            ("zoominfo.comp_standard_global", "ZoomInfo firmographic data — primary NAICS source",
             ["zi_c_location_id BIGINT", "zi_c_company_id BIGINT", "zi_c_name VARCHAR", "zi_c_url VARCHAR",
              "zi_c_street / zi_c_city / zi_c_state / zi_c_zip / zi_c_country",
              "zi_c_employees INT", "zi_c_revenue BIGINT", "zi_c_phone VARCHAR",
              "zi_c_naics2 / zi_c_naics4 / zi_c_naics6 VARCHAR (6-digit NAICS)",
              "zi_c_naics_top3 VARCHAR (pipe-delimited top 3)", "zi_c_naics_confidence_score FLOAT",
              "zi_c_naics_top3_confidence_scores VARCHAR",
              "zi_c_sic2 / zi_c_sic3 / zi_c_sic4 / zi_c_sic_top3 VARCHAR",
              "zi_c_industry_primary / zi_c_sub_industry_primary VARCHAR",
              "zi_c_year_founded VARCHAR", "zi_c_ein VARCHAR",
              "zi_c_linkedin_url / zi_c_facebook_url / zi_c_twitter_url"],
             "information_schema.columns (you provided)"),
            ("warehouse.equifax_us_latest", "Equifax firmographic data — NAICS/SIC/employee source",
             ["efx_id BIGINT", "efx_name / efx_legal_name VARCHAR",
              "efx_address / efx_city / efx_state / efx_zipcode / efx_ctryname",
              "efx_primnaicscode BIGINT (primary NAICS)", "efx_secnaics1-4 BIGINT (secondary NAICS codes)",
              "efx_primnaicsdesc / efx_secnaicsdesc1-2 VARCHAR",
              "efx_primsic BIGINT / efx_secsic1-4 BIGINT (SIC codes)",
              "efx_corpempcnt BIGINT (corporate employee count)", "efx_locempcnt BIGINT (location employees)",
              "efx_corpamount / efx_locamount BIGINT (revenue)",
              "efx_yrest BIGINT (year established)",
              "efx_web VARCHAR (website)", "efx_phone / efx_faxphone BIGINT",
              "efx_email VARCHAR", "efx_mbe / efx_wbe / efx_vet VARCHAR (minority/woman/veteran flags)",
              "efx_nonprofit / efx_gov / efx_edu VARCHAR",
              "efx_ceoname / efx_cioname / efx_cfoname VARCHAR"],
             "information_schema.columns (you provided)"),
            ("warehouse.oc_companies_latest", "OpenCorporates global registry — NAICS + international coverage",
             ["company_number VARCHAR", "jurisdiction_code VARCHAR (e.g. 'us_fl', 'gb')",
              "name / normalised_name VARCHAR", "company_type VARCHAR",
              "current_status VARCHAR (Active/Dissolved/etc.)",
              "incorporation_date VARCHAR", "dissolution_date VARCHAR",
              "industry_code_uids VARCHAR (pipe-delimited: 'us_naics-722511|uk_sic-56101')",
              "number_of_employees VARCHAR",
              "registered_address.street_address / .locality / .region / .postal_code / .country",
              "home_jurisdiction_code / home_jurisdiction_company_number",
              "inactive BOOLEAN", "has_been_liquidated BOOLEAN",
              "latest_accounts_cash / _assets / _liabilities FLOAT"],
             "information_schema.columns (you provided)"),
            ("datascience.zoominfo_matches_custom_inc_ml", "ZI entity match results — which businesses matched ZI records",
             ["business_id VARCHAR", "zi_c_company_id BIGINT", "zi_c_location_id BIGINT",
              "zi_es_location_id VARCHAR", "zi_probability FLOAT (XGBoost confidence: 0-1)",
              "similarity_index INT (0-55: match.index / 55 = confidence)"],
             "smb_zoominfo_standardized_joined.sql + your schema"),
            ("datascience.efx_matches_custom_inc_ml", "EFX entity match results — which businesses matched Equifax records",
             ["business_id VARCHAR", "efx_id BIGINT",
              "efx_probability FLOAT (XGBoost confidence 0-1, NULL if from similarity_index)",
              "similarity_index INT (0-55, NULL if from XGBoost — same MAX_CONFIDENCE_INDEX=55)"],
             "efx_match_select.sql"),
            ("datascience.oc_matches_custom_inc_ml", "OC entity match results — which businesses matched OpenCorporates records",
             ["business_id VARCHAR", "company_number VARCHAR", "jurisdiction_code VARCHAR",
              "oc_probability FLOAT (XGBoost confidence 0-1, NULL if from similarity_index)",
              "similarity_index INT (0-55, NULL if from XGBoost)"],
             "oc_match_select.sql"),
        ]
        for tname, purpose, cols, source in rs_tables:
            st.markdown(
                f'<div class="card" style="background:#0E1E38;border-left-color:#60A5FA;margin-bottom:8px;">'
                f'<div style="color:#60A5FA;font-weight:700;font-family:Courier New;">{tname}</div>'
                f'<div style="color:#94A3B8;font-size:.8rem;margin-bottom:4px;">{purpose} · Source: <code>{source}</code></div>'
                f'<div style="color:#CBD5E1;font-size:.78rem;">' +
                " · ".join(f"<code>{c}</code>" for c in cols) +
                "</div></div>",
                unsafe_allow_html=True
            )

        st.markdown("---")
        st.markdown("### ✅ Complete ready-to-run queries (using confirmed column names)")

        card("⚠️ <b>Confirmed working pattern:</b> The <code>value</code> column is stored as <code>character varying</code> (text JSON). "
             "Use <code>SELECT name, value, received_at</code> to fetch the raw text, then parse with <code>JSON_EXTRACT_PATH_TEXT(value, 'key')</code> or "
             "parse in Python with <code>json.loads(value_text)</code>. "
             "Never use <code>-&gt;&gt;</code> or <code>-&gt;</code> operators directly — they require native jsonb type.", "card-amber")

        st.markdown("#### Query 1 — Full KYB facts for one business (PostgreSQL — CONFIRMED WORKING)")
        st.code("""-- CONFIRMED WORKING — replace UUID with your business_id
-- value column is varchar (Redshift) — use JSON_EXTRACT_PATH_TEXT(value, 'key') — NO cast

SELECT
    name,
    JSON_EXTRACT_PATH_TEXT(value, 'value')                AS fact_value,
    JSON_EXTRACT_PATH_TEXT(value, 'source', 'platformId') AS winning_pid,
    JSON_EXTRACT_PATH_TEXT(value, 'source', 'confidence') AS confidence,
    JSON_EXTRACT_PATH_TEXT(value, 'override', 'value')    AS analyst_override,
    received_at
FROM rds_warehouse_public.facts
WHERE business_id = 'YOUR-BUSINESS-UUID-HERE'  -- replace
  AND name IN ('naics_code','mcc_code','industry','naics_description','mcc_description',
               'business_name','legal_name','tin_match_boolean','sos_active')
ORDER BY name;""", language="sql")

        st.markdown("#### Query 2 — All vendor NAICS responses (PostgreSQL — parse alternatives in Python)")
        st.code("""-- Fetch the raw value text — parse alternatives[] in Python (see Python tab)
-- The alternatives array cannot be unnested easily on varchar JSON columns

SELECT name, value, received_at
FROM rds_warehouse_public.facts
WHERE business_id = 'YOUR-BUSINESS-UUID-HERE'  -- replace
  AND name = 'naics_code';

-- Then in Python:
-- import json
-- fact = json.loads(row['value'])
-- for alt in fact.get('alternatives', []):
--     print(alt.get('platformId'), alt.get('value'), alt.get('confidence'))""", language="sql")

        st.markdown("#### Query 3 — ZoomInfo raw NAICS data for a business (Redshift)")
        st.code("""-- Shows raw ZI firmographic data for a matched business
-- Requires: business_id from your case, run in Redshift
SELECT
    m.business_id,
    m.zi_c_location_id,
    m.zi_probability              AS zi_match_confidence,
    m.similarity_index,
    m.similarity_index / 55.0     AS similarity_as_confidence,  -- MAX_CONFIDENCE_INDEX = 55
    z.zi_c_name                   AS zi_company_name,
    z.zi_c_naics6                 AS zi_naics_code,
    z.zi_c_naics_confidence_score AS zi_naics_confidence,
    z.zi_c_naics_top3             AS zi_naics_top3,
    z.zi_c_naics_top3_confidence_scores,
    z.zi_c_sic4                   AS zi_sic_code,
    z.zi_c_employees              AS zi_employees,
    z.zi_c_revenue                AS zi_revenue,
    z.zi_c_url                    AS zi_website,
    z.zi_c_state                  AS zi_state,
    z.zi_c_city                   AS zi_city,
    z.zi_c_year_founded           AS zi_year_founded,
    z.zi_c_ein                    AS zi_ein
FROM datascience.zoominfo_matches_custom_inc_ml m
JOIN zoominfo.comp_standard_global z
    ON z.zi_c_location_id = m.zi_c_location_id
WHERE m.business_id = 'YOUR-BUSINESS-UUID-HERE';""", language="sql")

        st.markdown("#### Query 4 — Equifax raw NAICS data for a business (Redshift)")
        st.code("""-- Shows raw Equifax firmographic data for a matched business
SELECT
    m.business_id,
    m.efx_id,
    m.efx_probability             AS efx_match_confidence,
    m.similarity_index,
    m.similarity_index / 55.0     AS similarity_as_confidence,
    e.efx_name                    AS efx_company_name,
    e.efx_legal_name,
    e.efx_primnaicscode           AS efx_primary_naics,
    e.efx_primnaicsdesc           AS efx_primary_naics_desc,
    e.efx_secnaics1               AS efx_secondary_naics1,
    e.efx_secnaics2               AS efx_secondary_naics2,
    e.efx_primsic                 AS efx_primary_sic,
    e.efx_primsicdesc             AS efx_primary_sic_desc,
    e.efx_corpempcnt              AS efx_corp_employees,
    e.efx_corpamount              AS efx_corp_revenue,
    e.efx_yrest                   AS efx_year_established,
    e.efx_web                     AS efx_website,
    e.efx_state,
    e.efx_mbe                     AS minority_owned,
    e.efx_wbe                     AS woman_owned,
    e.efx_vet                     AS veteran_owned,
    e.efx_email
FROM datascience.efx_matches_custom_inc_ml m
JOIN warehouse.equifax_us_latest e
    ON e.efx_id = m.efx_id
WHERE m.business_id = 'YOUR-BUSINESS-UUID-HERE';""", language="sql")

        st.markdown("#### Query 5 — OpenCorporates raw data for a business (Redshift)")
        st.code("""-- Shows OC registry data. industry_code_uids contains ALL taxonomies pipe-delimited:
-- e.g. 'us_naics-722511|uk_sic-56101|ca_naics-722511'
SELECT
    m.business_id,
    m.company_number,
    m.jurisdiction_code,
    m.oc_probability              AS oc_match_confidence,
    m.similarity_index,
    o.name                        AS oc_company_name,
    o.company_type,
    o.current_status,             -- 'Active', 'Dissolved', etc.
    o.incorporation_date,
    o.industry_code_uids,         -- parse for 'us_naics-XXXXXX'
    o."registered_address.region" AS registered_state,
    o."registered_address.country" AS registered_country,
    o.home_jurisdiction_code,
    o.inactive,
    o.has_been_liquidated,
    -- Extract NAICS from pipe-delimited industry_code_uids:
    SPLIT_PART(
        REGEXP_SUBSTR(industry_code_uids, 'us_naics-[0-9]+'),
        '-', 2
    ) AS extracted_us_naics
FROM datascience.oc_matches_custom_inc_ml m
JOIN warehouse.oc_companies_latest o
    ON o.company_number   = m.company_number
   AND o.jurisdiction_code = m.jurisdiction_code
WHERE m.business_id = 'YOUR-BUSINESS-UUID-HERE';""", language="sql")

        st.markdown("#### Query 6 — Pipeline B output: who won NAICS (ZI vs EFX) (Redshift)")
        st.code("""-- Shows Pipeline B winner-takes-all result
-- WHEN zi_match_confidence > efx_match_confidence THEN ZI NAICS ELSE EFX NAICS
SELECT
    business_id,
    primary_naics_code,
    mcc_code,
    mcc_desc,
    zi_match_confidence,
    efx_match_confidence,
    match_confidence,             -- = MAX(zi_conf, efx_conf)
    CASE
        WHEN zi_match_confidence > efx_match_confidence THEN 'ZoomInfo (pid=24)'
        ELSE 'Equifax (pid=17)'
    END                           AS pipeline_b_winner,
    employee_count,
    year_established,
    worth_score,
    company_name,
    tax_id,
    name_verification,
    address_verification,
    tin_verification,
    sos_match_verification,
    corporate_filing_business_name,
    corporate_filing_secretary_of_state_status
FROM datascience.customer_files
WHERE business_id = 'YOUR-BUSINESS-UUID-HERE';""", language="sql")

        st.markdown("#### Query 7 — Full Python: load all 6 vendor responses + Pipeline B (ready to run)")
        st.code("""import psycopg2
import pandas as pd

BID = 'YOUR-BUSINESS-UUID-HERE'  # replace with real UUID

# PostgreSQL connection (RDS)
pg = psycopg2.connect(
    host='your-rds-endpoint.rds.amazonaws.com', port=5432,
    dbname='postgres', user='your_pg_user', password='your_pg_pass',
    sslmode='require'
)

# Redshift connection
rs = psycopg2.connect(
    host='your-cluster.redshift.amazonaws.com', port=5439,
    dbname='dev', user='your_rs_user', password='your_rs_pass',
    sslmode='require'
)

# 1. Winning NAICS from Pipeline A (PostgreSQL facts table)
SQL_FACTS = (
    "SELECT name, JSON_EXTRACT_PATH_TEXT(value, 'value') AS value, "
    "JSON_EXTRACT_PATH_TEXT(value, 'source', 'platformId') AS winning_pid, "
    "JSON_EXTRACT_PATH_TEXT(value, 'source', 'confidence') AS confidence, "
    "JSON_EXTRACT_PATH_TEXT(value, 'override', 'value') AS analyst_override "
    "FROM rds_warehouse_public.facts "
    "WHERE business_id = %(bid)s "
    "AND name IN ('naics_code','mcc_code','mcc_description','industry','business_name','legal_name') "
    "ORDER BY name"
)
df_facts = pd.read_sql(SQL_FACTS, pg, params={'bid': BID})
print("=== Pipeline A (Facts table) ==="); print(df_facts.to_string())

# 2. All vendor alternatives for NAICS — fetch raw and parse in Python
SQL_ALTS = (
    "SELECT name, value FROM rds_warehouse_public.facts "
    "WHERE business_id = %(bid)s AND name = 'naics_code'"
)
row_naics = pd.read_sql(SQL_ALTS, pg, params={'bid': BID})
df_alts = row_naics  # parse alternatives below
import json as _json
if not row_naics.empty:
    _fact = _json.loads(row_naics['value'].iloc[0])
    _alts = _fact.get('alternatives', [])
    print("\\n=== All vendor NAICS alternatives ===")
    for _a in _alts:
        print(f"  PID {_a.get('platformId'):>3}  conf={str(_a.get('confidence','N/A')):>6}  naics={_a.get('value')}")
else:
    print("\\n=== No NAICS fact found ===")
print(df_alts.to_string())

# 3. Pipeline B (Redshift)
SQL_PIPB = (
    "SELECT primary_naics_code, zi_match_confidence, efx_match_confidence, "
    "CASE WHEN zi_match_confidence > efx_match_confidence THEN 'ZoomInfo' ELSE 'Equifax' END AS winner, "
    "worth_score, employee_count FROM datascience.customer_files WHERE business_id = %(bid)s"
)
df_pipb = pd.read_sql(SQL_PIPB, rs, params={'bid': BID})
print("\\n=== Pipeline B ==="); print(df_pipb.to_string())

# 4. ZI raw NAICS (Redshift)
SQL_ZI = (
    "SELECT m.zi_probability, m.similarity_index, z.zi_c_naics6, "
    "z.zi_c_naics_confidence_score, z.zi_c_naics_top3, z.zi_c_name, z.zi_c_employees "
    "FROM datascience.zoominfo_matches_custom_inc_ml m "
    "JOIN zoominfo.comp_standard_global z ON z.zi_c_location_id = m.zi_c_location_id "
    "WHERE m.business_id = %(bid)s"
)
df_zi = pd.read_sql(SQL_ZI, rs, params={'bid': BID})
print("\\n=== ZoomInfo raw NAICS ==="); print(df_zi.to_string())

# 5. EFX raw NAICS (Redshift)
SQL_EFX = (
    "SELECT m.efx_probability, m.similarity_index, e.efx_primnaicscode, "
    "e.efx_primnaicsdesc, e.efx_corpempcnt, e.efx_name, e.efx_mbe, e.efx_wbe, e.efx_vet "
    "FROM datascience.efx_matches_custom_inc_ml m "
    "JOIN warehouse.equifax_us_latest e ON e.efx_id = m.efx_id "
    "WHERE m.business_id = %(bid)s"
)
df_efx = pd.read_sql(SQL_EFX, rs, params={'bid': BID})
print("\\n=== Equifax raw NAICS ==="); print(df_efx.to_string())

pg.close(); rs.close()""", language="python")

    import os
    # Read key: (1) Streamlit Cloud secrets, (2) env var, (3) local secrets.toml
    OPENAI_API_KEY = ""
    try:
        OPENAI_API_KEY = st.secrets.get("OPENAI_API_KEY", "")
    except Exception:
        pass
    if not OPENAI_API_KEY:
        OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "")
    if not OPENAI_API_KEY:
        try:
            import tomllib, pathlib
            for sp in [pathlib.Path.home()/".streamlit"/"secrets.toml",
                       pathlib.Path(__file__).parent/".streamlit"/"secrets.toml"]:
                if sp.exists():
                    with open(sp,"rb") as f: sec = tomllib.load(f)
                    OPENAI_API_KEY = sec.get("OPENAI_API_KEY","")
                    if OPENAI_API_KEY: break
        except Exception: pass

    card(f"""<b>RAG Knowledge Base:</b> {rag_index['N']} chunks from 25 source files<br>
    Frontend (Admin Portal): fieldConfigs.tsx, SOSBadge.tsx, TinBadge.tsx, EntityJurisdictionCell.tsx,
    AddressesCard.tsx, WatchlistsTab.tsx, WebsiteTab.tsx, useBusinessRegistrationTabDetails.tsx, useWebsiteNonEditableFields.tsx...<br>
    Backend (SIC-UK-Codes): kyb/index.ts, businessDetails/index.ts, aiNaicsEnrichment.ts, rules.ts, sources.ts, report.ts, customer_table.sql...<br>
    <b>AI provides:</b> exact file+line citations, SQL queries to verify data, Python code to load facts.
    """, "card-teal")

    if not OPENAI_API_KEY:
        card("⚠️ No OPENAI_API_KEY. Run: OPENAI_API_KEY='sk-...' python3 -m streamlit run kyb_portal_app.py", "card-red")

    st.markdown("#### Quick Questions:")
    suggestions = [
        "Why does 'Business Registration ✅ Verified' appear with 'No Registry Data to Display'?",
        "What produces the Domestic Primary badge on Entity Jurisdiction Type?",
        "Why do all SoS fields show N/A when I edit the business name?",
        "Where does NAICS code come from and what are its 7 sources?",
        "Why does MCC description show 'Fallback MCC per instructions'?",
        "Which API endpoint powers the Watchlists tab and how are hits grouped?",
        "What is the difference between tin.value and tin_match_boolean?",
        "Why do Creation Date, Status go N/A when I edit the website URL?",
        "Where is watchlist.value.metadata stored and how do I query it?",
        "What happens when the analyst edits a field — where is the override stored?",
        "Write me SQL to see all fact values for a business_id",
        "Write Python code to load the sos_filings fact from the database",
    ]
    cols = st.columns(3)
    for i, q in enumerate(suggestions):
        if cols[i%3].button(q, key=f"q_{i}", use_container_width=True):
            st.session_state.setdefault("agent_history", [])
            st.session_state["agent_history"].append({"role":"user","content":q})
            st.session_state["pending_q"] = q

    st.markdown("---")
    if "agent_history" not in st.session_state:
        st.session_state["agent_history"] = []

    for msg in st.session_state["agent_history"]:
        if msg["role"] == "user":
            st.markdown(f'<div style="background:#1A2744;border:1px solid #2D4070;border-radius:8px;padding:10px 14px;margin:6px 0;color:#E2E8F0;"><b>You:</b> {msg["content"]}</div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div style="background:#0C2218;border:1px solid #059669;border-radius:8px;padding:12px 16px;margin:6px 0;"><b style="color:#A7F3D0;">🤖 Agent:</b><br><br><span style="color:#E2E8F0;white-space:pre-wrap;">{msg["content"]}</span></div>', unsafe_allow_html=True)
            if msg.get("sources"):
                with st.expander("📚 Source code chunks used"):
                    for s in msg["sources"][:5]:
                        url = f"{REPO_URLS.get(s.get('source_type','SIC_UK'),'')}/{s['path']}#L{s['line_start']}-L{s['line_end']}"
                        tag = "🖥️ Frontend" if s.get("source_type")=="ADMIN_PORTAL" else "⚙️ Backend"
                        st.markdown(f'<div class="src-ref">{tag} <a href="{url}" target="_blank" style="color:#60A5FA;">🔗 {s["path"]} L{s["line_start"]}–{s["line_end"]}</a></div>', unsafe_allow_html=True)
                        st.code(s["text"][:500]+("…" if len(s["text"])>500 else ""), language=None)

    st.markdown("#### 📎 Attach a screenshot:")
    uploaded = st.file_uploader("Upload screenshot or file", type=["png","jpg","jpeg","webp","pdf","xlsx"], key="upload")
    if uploaded and uploaded.type.startswith("image/"):
        st.image(uploaded, caption=f"📸 {uploaded.name}", use_column_width=True)
        uploaded.seek(0)

    user_q = st.chat_input("Ask about any field, badge, card, rule, or lineage...")
    pending = st.session_state.pop("pending_q", None)
    if pending: user_q = pending

    if user_q:
        full_q = user_q + (f" [Screenshot: {uploaded.name}]" if uploaded else "")
        st.session_state["agent_history"].append({"role":"user","content":full_q})

        with st.spinner("Searching 613 code chunks..."):
            retrieved = rag_search(rag_index, full_q, top_k=8)
            context = "\n".join(
                f"--- [{'Frontend' if c.get('source_type')=='ADMIN_PORTAL' else 'Backend'}] {c['path']} L{c['line_start']}–{c['line_end']} ---\n[{c.get('description','')[:150]}]\n{c['text'][:1400]}\n"
                for c in retrieved)

            image_content = None
            if uploaded and uploaded.type.startswith("image/"):
                img_b64 = base64.b64encode(uploaded.read()).decode()
                image_content = {"type":"image_url","image_url":{"url":f"data:{uploaded.type};base64,{img_b64}"}}

            history_gpt = [{"role":m["role"],"content":m["content"]} for m in st.session_state["agent_history"][:-1][-6:]]

            system_prompt = """You are a precise technical assistant for Worth AI's admin portal KYB tab.
You answer questions about every field, card, badge, and lineage in: Background, Business Registration, Contact Information, Website, Watchlists sub-tabs.

CRITICAL FACTS from source code:
- 'Domestic Primary'/'Foreign Secondary' badge: EntityJurisdictionCell.tsx JURISDICTION_CONFIG maps domestic→Primary (Worth label, NOT Middesk)
- SOSBadge: sosFiling.active=true→Verified(info/blue). sosFiling.active=false→Missing Active Filing(error/red)
- TinBadge: tin_match_boolean.value=true→Verified. 'Verified' = IRS TIN match, NOT SoS match
- hasDirtyBusinessRegistrationFields=true → ALL SoS fields show N/A
- isWebsiteDirty=true → Creation Date, Expiration Date, Parked Domain, Status all show N/A
- WatchlistsTab: ALL hits consolidated in watchlist.value.metadata (no separate people endpoint since BEST-65)
- VALUE_NOT_AVAILABLE = 'N/A' — React constant, not a backend concept
- Rule 4: NO minimum confidence cutoff for field display
- mcc_description 'Fallback MCC per instructions...' = known bug (Gap G5), editable=false so analyst cannot fix

RULES:
1. ALWAYS cite [FILE: path L123-L145] for every claim.
2. For screenshot analysis: identify every field, badge, card shown.
3. ALWAYS provide SQL and Python — use ONLY these CONFIRMED-WORKING Redshift patterns:

   DATABASE = AMAZON REDSHIFT. value column = varchar (text JSON). port=5439.
   ✅ WORKS: SELECT name, value, received_at FROM rds_warehouse_public.facts WHERE business_id='uuid' AND name IN ('fact1','fact2')
   ✅ WORKS: JSON_EXTRACT_PATH_TEXT(value, 'key')  — scalar extraction, NO cast
   ✅ WORKS: JSON_EXTRACT_PATH_TEXT(value, 'key1', 'key2')  — nested extraction (e.g. source.platformId)
   ❌ FAILS: value::json  →  type "json" does not exist
   ❌ FAILS: value->>'key'  →  operator does not exist
   ❌ FAILS: value->'key'  →  operator does not exist
   ALWAYS use json.loads(value_text) in Python — never str.split() or regex on JSON

   SCALAR FACT (single value like business_name, naics_code, tin_match_boolean):
     SELECT name,
            JSON_EXTRACT_PATH_TEXT(value, 'value')                AS fact_value,
            JSON_EXTRACT_PATH_TEXT(value, 'source', 'platformId') AS winning_vendor,
            JSON_EXTRACT_PATH_TEXT(value, 'source', 'confidence') AS vendor_conf,
            JSON_EXTRACT_PATH_TEXT(value, 'override', 'value')    AS analyst_override,
            received_at
     FROM rds_warehouse_public.facts
     WHERE business_id = 'YOUR-BUSINESS-UUID-HERE'
       AND name IN ('fact_name');
     -- winning_vendor: 16=Middesk 24=ZoomInfo 23=OC 17=Equifax 38=Trulioo 22=SERP 31=AI

   NESTED OBJECT FACTS (tin_match, address_verification, name_match, idv_status):
     SELECT JSON_EXTRACT_PATH_TEXT(value, 'value', 'status') AS status,
            JSON_EXTRACT_PATH_TEXT(value, 'value', 'message') AS message
     FROM rds_warehouse_public.facts
     WHERE business_id = 'YOUR-BUSINESS-UUID-HERE' AND name = 'tin_match';

   ARRAY FACTS (sos_filings, addresses, people, watchlist, dba_found):
     SELECT name, value, received_at FROM rds_warehouse_public.facts
     WHERE business_id = 'YOUR-BUSINESS-UUID-HERE' AND name = 'sos_filings';
     -- Then in Python: fact=json.loads(value); filings=fact.get('value',[])

   PIPELINE B (datascience.customer_files — plain columns, no JSON):
     SELECT primary_naics_code, mcc_code, zi_match_confidence, efx_match_confidence, worth_score
     FROM datascience.customer_files WHERE business_id = 'YOUR-BUSINESS-UUID-HERE';

4. PYTHON pattern — ALWAYS use this structure:
   import psycopg2, json
   conn = psycopg2.connect(host='your-cluster.redshift.amazonaws.com', port=5439,
       dbname='dev', user='u', password='pw', sslmode='require')
   BID = 'YOUR-BUSINESS-UUID-HERE'
   cur = conn.cursor()
   cur.execute('SELECT name, value, received_at FROM rds_warehouse_public.facts '
               'WHERE business_id = %s AND name IN (%s,%s)', (BID, 'fact1', 'fact2'))
   for name, value_text, ts in cur.fetchall():
       fact = json.loads(value_text)  # value is varchar — ALWAYS use json.loads()
       val = fact.get('value')
       src = fact.get('source', {})
       pid_map = {'16':'Middesk','24':'ZoomInfo','23':'OC','17':'Equifax',
                  '38':'Trulioo','22':'SERP','31':'AI'}
       vendor = pid_map.get(str(src.get('platformId','')), str(src.get('platformId')))
       print(f'{name}: value={val} | vendor={vendor} | conf={src.get("confidence")} | override={fact.get("override")}')
   cur.close(); conn.close()
5. Structure answers: 📍WHERE → 📡WHICH API → 🔧WHICH FACT → 📦WHICH SOURCE → 🗄️STORAGE → ✅VERIFY SQL → 📝WHY BLANK
6. Do NOT hallucinate. Only use provided code chunks.
7. End: "✅ Click source chunks below to verify." """

            user_msg: list = [{"type":"text","text":f"Question: {full_q}\n\nCode chunks:\n{context}"}]
            if image_content: user_msg.insert(0, image_content)

            if not OPENAI_API_KEY:
                answer = "⚠️ No OPENAI_API_KEY.\n\nRun: OPENAI_API_KEY='sk-...' python3 -m streamlit run kyb_portal_app.py\n\n--- Chunks (no GPT) ---\n" + "\n".join(f"📄 {c['path']} L{c['line_start']}:\n{c['text'][:300]}" for c in retrieved[:3])
            else:
                try:
                    from openai import OpenAI
                    client = OpenAI(api_key=OPENAI_API_KEY)
                    resp = client.chat.completions.create(
                        model="gpt-4o-mini",
                        messages=[{"role":"system","content":system_prompt},*history_gpt,{"role":"user","content":user_msg}],
                        temperature=0.1, max_tokens=1500)
                    answer = resp.choices[0].message.content
                except Exception as e:
                    answer = f"⚠️ Error: {e}\n\n" + "\n".join(f"📄 {c['path']} L{c['line_start']}:\n{c['text'][:300]}" for c in retrieved[:3])

        st.session_state["agent_history"].append({"role":"assistant","content":answer,"sources":retrieved})
        st.rerun()

    if st.button("🗑️ Clear"):
        st.session_state["agent_history"] = []
        st.rerun()
